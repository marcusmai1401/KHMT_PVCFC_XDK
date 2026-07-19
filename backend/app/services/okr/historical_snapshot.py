from __future__ import annotations

from dataclasses import dataclass, field
from io import BytesIO
import hashlib
import json
from pathlib import Path
import re
import unicodedata
from typing import Any
from zipfile import ZipFile
import xml.etree.ElementTree as ET

from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models.domain import HistoricalSnapshotModel
from app.services.okr.team_normalizer import normalize_team_label
from app.services.repositories import make_id, model_to_dict


DATA_BLOCK_RANGES = {
    "o2_bddk": ("data", "A3:E18"),
    "o2_scdx": ("data", "A21:E35"),
    "stop_by_team": ("data", "A67:E70"),
    "stop_by_month": ("data", "A72:D84"),
    "training": ("data", "A98:N107"),
    "competency": ("data", "A135:B142"),
    "vhdn_running": ("data", "A86:E89"),
    "vhdn_sports": ("data", "A91:E94"),
    "sk_initiatives": ("data", "A110:B114"),
}

UNCONFIRMED_BLOCKS = [
    {
        "source_range": "data!A117:D127",
        "observed_label": "Tuần 14-22 backlog",
        "candidate_kr_codes": ["O2.KR3"],
        "mapping_status": "needs_confirmation",
        "reason": "Weekly backlog block needs business confirmation before UI aggregation.",
    },
]

UNCONFIRMED_WARNING_BLOCKS = [
    {
        "source_range": "data!A43:E62",
        "observed_label": "BDĐK NPK",
        "candidate_kr_codes": ["O2.KR2"],
        "mapping_status": "needs_confirmation",
        "reason": "BDĐK NPK block is preserved as an import warning until business mapping is confirmed.",
    },
]

SAP_COMPLIANCE_MANIFEST_PATH = Path(__file__).resolve().parents[2] / "data" / "sap_compliance_snapshots.json"

DRAWING_NAMESPACES = {
    "a": "http://schemas.openxmlformats.org/drawingml/2006/main",
    "xdr": "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing",
}

# Heading of a KR narrative text box, e.g. "KR 04 Mở rộng..." or "KR 7. Thiết kế...".
_KR_HEAD = re.compile(r"^KR\s*0*(\d{1,2})\b", re.IGNORECASE)
# Objective header text box, e.g. "O1: ... - Mục tiêu: 0 vụ - Kết quả: 0 vụ".
_OBJ_HEAD = re.compile(r"^(O[1-6])\s*[:.]")
_OBJ_TARGET = re.compile(r"Mục tiêu\s*:?\s*(.+?)\s*[-–]\s*Kết quả\s*:?\s*(.+)$")
# Short chart labels / pillar tags that must not be mistaken for narrative prose.
_NARRATIVE_SKIP = {
    "am", "pm", "et", "fi", "oi", "she", "muc tieu", "ket qua", "thuc hien",
    "ke", "hoach", "ke hoach", "luy ke", "so lan to chuc", "tcdk", "tbch", "tbdl", "tbhtdk",
}


@dataclass
class HistoricalSnapshotImportResult:
    imported_count: int = 0
    updated_count: int = 0
    skipped_duplicates: int = 0
    months_covered: set[int] = field(default_factory=set)
    source_file_hash: str = ""
    warnings: list[dict[str, Any]] = field(default_factory=list)

    def to_dict(self) -> dict[str, Any]:
        return {
            "imported_count": self.imported_count,
            "updated_count": self.updated_count,
            "skipped_duplicates": self.skipped_duplicates,
            "months_covered": sorted(self.months_covered),
            "source_file_hash": self.source_file_hash,
            "warnings": self.warnings,
        }


def _warning(reason: str, source_range: str | None = None, severity: str = "MEDIUM") -> dict[str, Any]:
    return {
        "warning_type": "HISTORICAL_SNAPSHOT_PARSE_WARNING",
        "severity": severity,
        "source_cell": {"source_range": source_range} if source_range else None,
        "reason": reason,
        "admin_action": "PENDING",
    }


def _source_year(workbook: Any) -> int:
    for sheet_name, coordinate in [("Dashboard", "A20"), ("Dashboard", "A1")]:
        if sheet_name not in workbook.sheetnames:
            continue
        value = str(workbook[sheet_name][coordinate].value or "")
        match = re.search(r"(20\d{2})", value)
        if match:
            return int(match.group(1))
    return 2026


def _source_month(workbook: Any) -> int | None:
    for sheet_name, coordinate in [("Dashboard", "A1"), ("Dashboard", "A20")]:
        if sheet_name not in workbook.sheetnames:
            continue
        value = str(workbook[sheet_name][coordinate].value or "")
        match = re.search(r"(?:THÁNG|T)[\s._-]*(1[0-2]|0?[1-9])", value.upper())
        if match:
            return int(match.group(1))
    return None


def _validate_sap_compliance_payload(payload: dict[str, Any]) -> None:
    backlog_total = int(payload.get("backlog_total") or 0)
    totals = payload.get("totals") or {}
    rates = payload.get("rates") or {}
    supervisors = payload.get("supervisors") or []
    if backlog_total <= 0:
        raise ValueError("SAP backlog_total must be greater than zero")
    if not isinstance(supervisors, list) or not supervisors:
        raise ValueError("SAP supervisor breakdown is missing")

    total_keys = ("overdue_wo", "unconfirmed_wo", "violating_wo")
    normalized_totals = {key: int(totals.get(key) or 0) for key in total_keys}
    if any(value < 0 for value in normalized_totals.values()):
        raise ValueError("SAP WO totals cannot be negative")
    if normalized_totals["overdue_wo"] + normalized_totals["unconfirmed_wo"] != normalized_totals["violating_wo"]:
        raise ValueError("SAP violating WO total must equal overdue plus unconfirmed WO")
    if normalized_totals["violating_wo"] > backlog_total:
        raise ValueError("SAP violating WO total cannot exceed total backlog")

    supervisor_totals = {key: 0 for key in total_keys}
    names: set[str] = set()
    for supervisor in supervisors:
        name = str(supervisor.get("name") or "").strip()
        if not name:
            raise ValueError("SAP supervisor name is required")
        if name in names:
            raise ValueError(f"Duplicate SAP supervisor: {name}")
        names.add(name)
        overdue = int(supervisor.get("overdue_wo") or 0)
        unconfirmed = int(supervisor.get("unconfirmed_wo") or 0)
        violating = int(supervisor.get("violating_wo") or 0)
        if min(overdue, unconfirmed, violating) < 0:
            raise ValueError(f"SAP WO counts cannot be negative for supervisor {name}")
        if overdue + unconfirmed != violating:
            raise ValueError(f"SAP WO subtotal is inconsistent for supervisor {name}")
        supervisor_totals["overdue_wo"] += overdue
        supervisor_totals["unconfirmed_wo"] += unconfirmed
        supervisor_totals["violating_wo"] += violating
    if supervisor_totals != normalized_totals:
        raise ValueError("SAP supervisor totals do not reconcile with report totals")

    rate_pairs = {
        "overdue_share": normalized_totals["overdue_wo"],
        "unconfirmed_share": normalized_totals["unconfirmed_wo"],
        "violation_share": normalized_totals["violating_wo"],
    }
    for key, numerator in rate_pairs.items():
        rate = float(rates.get(key))
        if abs(rate - (numerator / backlog_total)) > 0.0006:
            raise ValueError(f"SAP rate {key} does not reconcile with total backlog")


def extract_sap_compliance_payload(
    workbook_bytes: bytes,
    *,
    month: int,
    year: int,
    manifest: dict[str, Any] | None = None,
) -> dict[str, Any] | None:
    registry = manifest
    if registry is None:
        registry = json.loads(SAP_COMPLIANCE_MANIFEST_PATH.read_text(encoding="utf-8"))
    period_key = f"{year:04d}-{month:02d}"
    configured = registry.get(period_key)
    if configured is None:
        return None
    payload = dict(configured)
    expected_hash = str(payload.get("source_image_sha256") or "").lower()
    if not re.fullmatch(r"[0-9a-f]{64}", expected_hash):
        raise ValueError(f"SAP manifest {period_key} has an invalid source image SHA-256")

    matched_media_name: str | None = None
    matched_size = 0
    with ZipFile(BytesIO(workbook_bytes)) as archive:
        for media_name in archive.namelist():
            if not media_name.startswith("xl/media/"):
                continue
            media_bytes = archive.read(media_name)
            if hashlib.sha256(media_bytes).hexdigest() == expected_hash:
                matched_media_name = media_name
                matched_size = len(media_bytes)
                break
    if matched_media_name is None:
        raise ValueError(f"SAP source image for {period_key} is not present or does not match the audited image")
    expected_size = int(payload.get("source_image_size_bytes") or 0)
    if expected_size and matched_size != expected_size:
        raise ValueError(f"SAP source image size for {period_key} does not match the audited image")

    _validate_sap_compliance_payload(payload)
    return {
        "block_type": "sap_compliance",
        **payload,
        "period": {"month": month, "year": year, "label": payload.get("period_label") or f"T{month}/{year}"},
        "source_image_file": matched_media_name,
    }


def _dashboard_month_columns() -> list[tuple[int, int]]:
    return [(month, 6 + ((month - 1) * 2)) for month in range(1, 13)]


def _upsert_snapshot(
    db: Session,
    result: HistoricalSnapshotImportResult,
    *,
    source_file_name: str,
    source_sheet: str,
    source_range: str,
    source_label: str | None,
    team: str,
    month: int,
    year: int,
    monthly_assessment: str | None = None,
    chart_payload: dict[str, Any] | None = None,
    warnings: list[dict[str, Any]] | None = None,
    imported_by: str,
) -> None:
    existing = db.execute(
        select(HistoricalSnapshotModel).where(
            HistoricalSnapshotModel.source_file_hash == result.source_file_hash,
            HistoricalSnapshotModel.team == team,
            HistoricalSnapshotModel.month == month,
            HistoricalSnapshotModel.year == year,
            HistoricalSnapshotModel.source_range == source_range,
        )
    ).scalar_one_or_none()
    if existing is not None:
        next_payload = chart_payload or {}
        next_warnings = warnings or []
        changed = False
        if existing.monthly_assessment != monthly_assessment:
            existing.monthly_assessment = monthly_assessment
            changed = True
        if existing.chart_payload != next_payload:
            existing.chart_payload = next_payload
            changed = True
        if existing.warnings != next_warnings:
            existing.warnings = next_warnings
            changed = True
        if changed:
            result.updated_count += 1
        else:
            result.skipped_duplicates += 1
        return
    db.add(
        HistoricalSnapshotModel(
            id=make_id("snap"),
            source_file_name=source_file_name,
            source_file_hash=result.source_file_hash,
            source_sheet=source_sheet,
            source_range=source_range,
            source_label=source_label,
            team=team,
            month=month,
            year=year,
            monthly_assessment=monthly_assessment,
            kr_statuses={},
            chart_payload=chart_payload or {},
            warnings=warnings or [],
            imported_by=imported_by,
            is_historical_snapshot=True,
        )
    )
    result.imported_count += 1
    if month:
        result.months_covered.add(month)


def find_dashboard_header_row(sheet: Any) -> int | None:
    for row in range(1, sheet.max_row + 1):
        first_cell = str(sheet.cell(row, 1).value or "").strip().lower()
        if first_cell == "đội/tổ" or first_cell == "doi/to":
            return row
    return None


def find_dashboard_team_rows(sheet: Any) -> list[tuple[int, str]]:
    header_row = find_dashboard_header_row(sheet)
    if header_row is None:
        rows: list[tuple[int, str]] = []
        for row in range(1, sheet.max_row + 1):
            source_label = str(sheet.cell(row, 1).value or "").strip()
            team, _original_label = normalize_team_label(source_label)
            if team:
                rows.append((row, team))
        return rows
    rows: list[tuple[int, str]] = []
    for row in range(header_row + 1, sheet.max_row + 1):
        source_label = str(sheet.cell(row, 1).value or "").strip()
        if not source_label:
            if rows:
                break
            continue
        team, _original_label = normalize_team_label(source_label)
        if team:
            rows.append((row, team))
            continue
        if rows:
            break
    return rows


def _parse_dashboard_history(db: Session, workbook: Any, result: HistoricalSnapshotImportResult, *, source_file_name: str, imported_by: str) -> None:
    if "Dashboard" not in workbook.sheetnames:
        raise ValueError("Workbook is missing required sheet: Dashboard")
    sheet = workbook["Dashboard"]
    year = _source_year(workbook)
    header_row = find_dashboard_header_row(sheet)
    team_rows = find_dashboard_team_rows(sheet)
    if not team_rows:
        result.warnings.append(_warning("No dashboard team rows found below Đội/Tổ header", "Dashboard"))
    for row, team in team_rows:
        source_label = str(sheet.cell(row, 1).value or "").strip()
        _team, original_label = normalize_team_label(source_label)
        for month, col in _dashboard_month_columns():
            assessment = str(sheet.cell(row, col).value or "").strip()
            if not assessment:
                continue
            _upsert_snapshot(
                db,
                result,
                source_file_name=source_file_name,
                source_sheet="Dashboard",
                source_range=f"Dashboard!{sheet.cell(row, col).coordinate}",
                source_label=original_label,
                team=team,
                month=month,
                year=year,
                monthly_assessment=assessment,
                imported_by=imported_by,
            )


def _range_rows(sheet: Any, range_text: str) -> list[dict[str, Any]]:
    min_col, min_row, max_col, max_row = range_boundaries(range_text)
    rows = []
    for row_idx in range(min_row, max_row + 1):
        values = [sheet.cell(row_idx, col_idx).value for col_idx in range(min_col, max_col + 1)]
        if not any(value is not None and str(value).strip() for value in values):
            continue
        rows.append(
            {
                "source_row": row_idx,
                "label": values[0],
                "value": values[1] if len(values) > 1 else None,
                "values": values,
            }
        )
    return rows


def _plain_text(value: Any) -> str:
    text = unicodedata.normalize("NFD", str(value or ""))
    normalized = "".join(char for char in text if unicodedata.category(char) != "Mn")
    return normalized.replace("đ", "d").replace("Đ", "D").lower()


def _extract_dashboard_note_blocks(workbook_bytes: bytes) -> dict[str, list[str]]:
    notes: dict[str, list[str]] = {}
    try:
        with ZipFile(BytesIO(workbook_bytes)) as archive:
            drawing_names = [
                name
                for name in archive.namelist()
                if name.startswith("xl/drawings/") and name.endswith(".xml")
            ]
            for drawing_name in drawing_names:
                root = ET.fromstring(archive.read(drawing_name))
                for anchor in root:
                    paragraphs = []
                    for paragraph in anchor.findall(".//a:p", DRAWING_NAMESPACES):
                        text = "".join(node.text or "" for node in paragraph.findall(".//a:t", DRAWING_NAMESPACES)).strip()
                        if text:
                            paragraphs.append(text)
                    if not paragraphs:
                        continue
                    heading = _plain_text(paragraphs[0])
                    if "kr02" in heading and "bao cao cong tac bddk" in heading:
                        notes["o2_bddk"] = paragraphs[1:]
                    elif "kr03" in heading and "bao cao cong tac sua chua dot xuat" in heading:
                        notes["o2_scdx"] = paragraphs[1:]
    except Exception:
        return {}
    return notes


def _narrative_boxes(workbook_bytes: bytes) -> list[dict[str, Any]]:
    """Read every Dashboard drawing text box as {col, row, lines}."""
    boxes: list[dict[str, Any]] = []
    with ZipFile(BytesIO(workbook_bytes)) as archive:
        drawing_names = [
            name
            for name in archive.namelist()
            if name.startswith("xl/drawings/") and name.endswith(".xml")
        ]
        for drawing_name in drawing_names:
            root = ET.fromstring(archive.read(drawing_name))
            for anchor in root:
                col = anchor.find(".//xdr:from/xdr:col", DRAWING_NAMESPACES)
                row = anchor.find(".//xdr:from/xdr:row", DRAWING_NAMESPACES)
                if col is None or row is None:
                    continue
                lines: list[str] = []
                for paragraph in anchor.findall(".//a:p", DRAWING_NAMESPACES):
                    text = "".join(node.text or "" for node in paragraph.findall(".//a:t", DRAWING_NAMESPACES)).strip()
                    if text:
                        lines.append(text)
                if lines:
                    boxes.append({"col": int(col.text), "row": int(row.text), "lines": lines})
    return boxes


def _narrative_band(box: dict[str, Any]) -> str | None:
    """Map a text box position to the objective whose column/row band it sits in."""
    col, row = box["col"], box["row"]
    if col <= 13:
        if row < 60:
            return "O1"
        if row < 143:
            return "O2"
        if row < 167:
            return "O3"
        return "O4"
    if row < 146:
        return "O5"
    return "O6"


def _is_narrative_line(line: str) -> bool:
    """True when a paragraph reads like report prose rather than a chart label."""
    text = line.strip()
    if not text or _plain_text(text) in _NARRATIVE_SKIP:
        return False
    if re.match(r"^[\d.,]+%?$", text):  # pure number / percent (chart data label)
        return False
    if re.match(r"^\d+%\s*:", text):  # radar legend "25%: ..."
        return False
    if text.startswith(("-", "+", "•", "*")):
        return True
    if re.match(r"^(TB|TC|Đội|Tổ|FS|ĐH|KH|Không|Trong|Hoàn|Đang|Front|Xây)", text):
        return True
    return len(text) >= 18 and " " in text


# Standalone chart-axis fragments, pillar tags and bare labels that must never
# enter the textual report (they only make sense next to their chart).
_REPORT_AXIS_NOISE = {
    "thuc hien", "ke", "hoach", "ke hoach", "luy ke", "so lan to chuc",
    "et", "am", "pm", "fi", "oi", "she", "tcdk", "tbch", "tbdl", "tbhtdk",
    "tieu chi danh gia", "muc tieu", "ket qua",
}
# A text box whose heading already renders as a dedicated chart is skipped so the
# report does not duplicate it: the per-team additions ("Hạng mục phát sinh") that
# the user does not want, and the STOP programme headers (shown as O3 charts).
_REPORT_SKIP_HEADINGS = ("hang muc phat sinh", "chuong trinh stop")


def _is_report_noise(line: str) -> bool:
    text = line.strip()
    if not text:
        return True
    if re.match(r"^[\d.,]+\s*%?$", text):  # pure number / percent (chart data label)
        return True
    if re.match(r"^\d+%\s*:", text):  # radar legend "25%: Xây dựng"
        return True
    return _plain_text(text) in _REPORT_AXIS_NOISE


def _kr_number(code: str) -> int:
    try:
        return int(code.split(".KR", 1)[1])
    except (IndexError, ValueError):
        return 999


def _build_objective_report(boxes: list[dict[str, Any]]) -> dict[str, Any]:
    """Group every Dashboard text box into a verbatim, per-objective report.

    Each objective gets its KR headings (with any in-box progress lines kept
    word-for-word) plus a list of loose ``notes``. The team-specific "Hạng mục
    phát sinh" blocks and already-charted STOP headers are dropped. This is what
    lets the web dashboard reproduce the full Excel narrative without rewording.
    """
    report: dict[str, Any] = {}
    for band in ("O1", "O2", "O3", "O4", "O5", "O6"):
        band_boxes = sorted(
            (box for box in boxes if _narrative_band(box) == band),
            key=lambda box: (box["row"], box["col"]),
        )
        groups: dict[str, dict[str, Any]] = {}
        notes: list[str] = []
        # KR a *heading-less* box attaches to. In Feb/Mar/Apr the Excel splits a
        # KR heading and its progress into separate boxes stacked vertically; we
        # carry the last unambiguous heading so the body lands under its KR
        # instead of in the loose-notes bucket. Reset when a box carries several
        # different KR headings (e.g. Feb lists KR02/04/05/06/03 in one box), so
        # we never guess which KR an orphan body belongs to.
        carry_code: str | None = None
        for box in band_boxes:
            if any(skip in _plain_text(box["lines"][0]) for skip in _REPORT_SKIP_HEADINGS):
                continue
            current: dict[str, Any] | None = None
            heading_codes: list[str] = []
            for line in box["lines"]:
                head = _KR_HEAD.match(line)
                if head:
                    code = f"{band}.KR{int(head.group(1))}"
                    if code not in heading_codes:
                        heading_codes.append(code)
                    existing = groups.get(code)
                    if existing is not None:
                        # A repeated heading for a KR already seen (e.g. the
                        # "BÁO CÁO LŨY KẾ" cumulative chart header next to the base
                        # one): keep writing body lines into the original group but
                        # drop the duplicate heading text. NOTE: only true duplicates
                        # are dropped — a first-seen heading whose status contains
                        # "Lũy kế" (e.g. KR10 "… - Lũy kế 186/201 HM") is preserved.
                        current = existing
                    else:
                        current = {"code": code, "title": line.strip(), "lines": []}
                        groups[code] = current
                    continue
                if _OBJ_HEAD.match(line):
                    current = None
                    continue
                if _is_report_noise(line):
                    continue
                target = current
                if target is None and not heading_codes and carry_code in groups:
                    target = groups[carry_code]
                if target is not None:
                    target["lines"].append(line.strip())
                else:
                    notes.append(line.strip())
            if heading_codes:
                carry_code = heading_codes[0] if len(heading_codes) == 1 else None
        krs = sorted(groups.values(), key=lambda group: _kr_number(group["code"]))
        if krs or notes:
            report[band] = {"krs": krs, "notes": notes}
    return report


def extract_dashboard_narratives(workbook_bytes: bytes) -> dict[str, Any]:
    """Extract the free-text narrative content of the Dashboard sheet (drawing text boxes).

    These paragraphs are not stored in cells, so they are invisible to the cell/data
    parsers. We recover them here so the web dashboard can reproduce the objective
    target/result lines, the per-KR progress notes (O4/O5), the "Số lần tổ chức" counts
    (O6) and the discipline-violation note exactly as they appear in Excel.
    """
    result: dict[str, Any] = {
        "objectives": {},
        "kr_details": {},
        "violations": [],
        "o6_counts": {},
        "extras": {},
        "report": {},
    }
    try:
        boxes = _narrative_boxes(workbook_bytes)
    except Exception:
        return result
    result["report"] = _build_objective_report(boxes)

    for box in boxes:
        first = box["lines"][0]
        obj_match = _OBJ_HEAD.match(first)
        if obj_match:
            head = " ".join(box["lines"])
            entry: dict[str, Any] = {"full": head}
            target = _OBJ_TARGET.search(head)
            if target:
                entry["target"] = target.group(1).strip()
                entry["result"] = target.group(2).strip()
            result["objectives"][obj_match.group(1)] = entry
        if "ghi nhan vi pham" in _plain_text(first):
            result["violations"] = box["lines"]

    for band in ("O4", "O5"):
        band_boxes = sorted(
            (b for b in boxes if _narrative_band(b) == band),
            key=lambda b: (b["col"], b["row"]),
        )
        titles = [
            {**b, "kr": int(_KR_HEAD.match(b["lines"][0]).group(1))}
            for b in band_boxes
            if _KR_HEAD.match(b["lines"][0])
        ]

        def _add(code: str, lines: list[str]) -> None:
            kept = [ln for ln in lines if _is_narrative_line(ln)]
            if kept:
                result["kr_details"].setdefault(code, []).extend(kept)

        for box in band_boxes:
            head = _KR_HEAD.match(box["lines"][0])
            if head:
                _add(f"{band}.KR{int(head.group(1))}", box["lines"][1:])
                continue
            candidates = [t for t in titles if t["row"] <= box["row"]]
            if not candidates:
                continue
            owner = max(candidates, key=lambda t: t["row"])
            _add(f"{band}.KR{owner['kr']}", box["lines"])

    for box in boxes:
        lines = box["lines"]
        if _narrative_band(box) == "O6" and len(lines) == 4 and lines[1] == "Lũy kế" and lines[2] == "Mục tiêu":
            if box["row"] < 210:
                key = "running" if box["col"] < 35 else "sports"
            else:
                key = "culture"
            result["o6_counts"][key] = {"actual": lines[0], "target": lines[3]}

    for box in boxes:
        first = box["lines"][0]
        if "vị trí chức danh" in first:
            result["extras"].setdefault("competency_positions", first)
        if re.match(r"^\s*\d+\s*CBCNV", first):  # e.g. "33 CBCNV (KTV)", not the KR2 title
            result["extras"].setdefault("cbcnv", first)
        if "Thanh toán lần" in first or "Thành tích lần" in first:
            result["extras"].setdefault("training_payment", first)
    return result


def _parse_data_blocks(
    db: Session,
    workbook: Any,
    result: HistoricalSnapshotImportResult,
    *,
    source_file_name: str,
    imported_by: str,
    note_blocks: dict[str, list[str]] | None = None,
) -> None:
    if "data" not in workbook.sheetnames:
        result.warnings.append(_warning("Workbook is missing optional sheet: data", "data"))
        return
    sheet = workbook["data"]
    year = _source_year(workbook)
    source_month = _source_month(workbook)
    for block_type, (sheet_name, range_text) in DATA_BLOCK_RANGES.items():
        source_range = f"{sheet_name}!{range_text}"
        try:
            rows = _range_rows(sheet, range_text)
            warnings: list[dict[str, Any]] = []
            if block_type == "competency" and not rows and source_month in {1, 2}:
                warnings.append(
                    _warning(
                        "Competency data is expected to be absent before T3",
                        source_range,
                        "LOW",
                    )
                )
            _upsert_snapshot(
                db,
                result,
                source_file_name=source_file_name,
                source_sheet=sheet_name,
                source_range=source_range,
                source_label=block_type,
                team="__CHARTS__",
                month=0,
                year=year,
                chart_payload={
                    "block_type": block_type,
                    "rows": rows,
                    "notes": list((note_blocks or {}).get(block_type, [])),
                },
                warnings=warnings,
                imported_by=imported_by,
            )
        except Exception as exc:
            result.warnings.append(_warning(f"Cannot parse {source_range}: {exc}", source_range))

    _upsert_snapshot(
        db,
        result,
        source_file_name=source_file_name,
        source_sheet="data",
        source_range="data!unconfirmed_blocks",
        source_label="unconfirmed_blocks",
        team="__SOURCE_REFERENCES__",
        month=0,
        year=year,
        chart_payload={"block_type": "unconfirmed_blocks", "items": UNCONFIRMED_BLOCKS},
        warnings=[],
        imported_by=imported_by,
    )
    for item in UNCONFIRMED_WARNING_BLOCKS:
        result.warnings.append(
            _warning(
                f"{item['observed_label']} is not imported as authoritative dashboard data: {item['reason']}",
                item["source_range"],
                "LOW",
            )
        )


def _parse_sap_compliance(
    db: Session,
    workbook: Any,
    workbook_bytes: bytes,
    result: HistoricalSnapshotImportResult,
    *,
    source_file_name: str,
    imported_by: str,
) -> None:
    source_month = _source_month(workbook)
    if source_month is None:
        return
    year = _source_year(workbook)
    try:
        payload = extract_sap_compliance_payload(workbook_bytes, month=source_month, year=year)
    except Exception as exc:
        result.warnings.append(
            _warning(
                f"Cannot import SAP compliance report: {exc}",
                "Dashboard!SAP_COMPLIANCE",
                "HIGH",
            )
        )
        return
    if payload is None:
        return
    _upsert_snapshot(
        db,
        result,
        source_file_name=source_file_name,
        source_sheet="Dashboard",
        source_range="Dashboard!SAP_COMPLIANCE",
        source_label=str(payload.get("title") or "SAP compliance"),
        team="__CHARTS__",
        month=0,
        year=year,
        chart_payload=payload,
        warnings=[],
        imported_by=imported_by,
    )


def import_historical_snapshot(
    db: Session,
    workbook_bytes: bytes,
    *,
    source_file_name: str,
    imported_by: str,
) -> dict[str, Any]:
    result = HistoricalSnapshotImportResult(source_file_hash=hashlib.sha256(workbook_bytes).hexdigest())
    try:
        workbook = load_workbook(BytesIO(workbook_bytes), read_only=False, data_only=True, keep_links=False)
    except Exception as exc:
        raise ValueError(f"Invalid Excel workbook: {exc}") from exc

    try:
        _parse_dashboard_history(db, workbook, result, source_file_name=source_file_name, imported_by=imported_by)
    except ValueError as exc:
        result.warnings.append(_warning(str(exc), "Dashboard", "HIGH"))
    _parse_data_blocks(
        db,
        workbook,
        result,
        source_file_name=source_file_name,
        imported_by=imported_by,
        note_blocks=_extract_dashboard_note_blocks(workbook_bytes),
    )
    _parse_sap_compliance(
        db,
        workbook,
        workbook_bytes,
        result,
        source_file_name=source_file_name,
        imported_by=imported_by,
    )
    try:
        narratives = extract_dashboard_narratives(workbook_bytes)
        _upsert_snapshot(
            db,
            result,
            source_file_name=source_file_name,
            source_sheet="Dashboard",
            source_range="Dashboard!narratives",
            source_label="dashboard_narratives",
            team="__CHARTS__",
            month=0,
            year=_source_year(workbook),
            chart_payload={"block_type": "dashboard_narratives", **narratives},
            warnings=[],
            imported_by=imported_by,
        )
    except Exception as exc:  # pragma: no cover - narrative extraction must never break import
        result.warnings.append(_warning(f"Cannot parse Dashboard narratives: {exc}", "Dashboard!narratives", "LOW"))
    db.flush()
    return result.to_dict()


def snapshots_to_dicts(records: list[HistoricalSnapshotModel]) -> list[dict[str, Any]]:
    return [model_to_dict(record) for record in records]
