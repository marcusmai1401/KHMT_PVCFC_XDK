from __future__ import annotations

from datetime import datetime, timezone
from pathlib import Path
import argparse
import re

from openpyxl import load_workbook
from sqlalchemy import and_, or_, select

from app.db.session import create_session
from app.models.domain import SKCTKTModel
from app.services.fi.workflow import SKStatus
from app.services.repositories import make_id


SHEET_TEAM = {
    "TBCH": "TBCH",
    "TBĐ": "TBĐL",
    "TBHTĐK": "TBHTĐK",
    "TC- ĐK": "TCĐK",
}
SHEET_KHMT_COLUMN = {
    "TBĐ": 14,
}
SHEET_LEADER_CONCLUSION_COLUMN: dict[str, int | None] = {
    "TBĐ": None,
}


def clean(value) -> str:
    return str(value or "").strip()


def parse_month(value: str) -> int | None:
    text = clean(value).lower()
    if not text:
        return None
    match = re.search(r"(?:tháng|thang|t)?\s*(1[0-2]|0?[1-9])\s*[./-]\s*20\d{2}", text)
    if match:
        return int(match.group(1))
    match = re.search(r"(?:tháng|thang|t)\s*(1[0-2]|0?[1-9])\b", text)
    if match:
        return int(match.group(1))
    if re.fullmatch(r"1[0-2]|0?[1-9]", text):
        return int(text)
    return None


def status_from_review(review: str) -> str:
    text = clean(review).lower()
    if not text:
        return SKStatus.SUBMITTED.value
    # Negative checks must run before the positive "đồng ý" check.
    rejected_markers = [
        "không đồng ý",
        "khong dong y",
        "khồng đồng ý",
        "khong dong ý",
        "không đạt",
        "khong dat",
        "không dat",
    ]
    if any(marker in text for marker in rejected_markers):
        return SKStatus.REJECTED.value
    if "xem xét sau" in text or "xem xet sau" in text:
        return SKStatus.DEFERRED.value
    if "đồng ý" in text or "dong y" in text:
        return SKStatus.APPROVED.value
    return SKStatus.SUBMITTED.value


def has_data(sheet, row: int) -> bool:
    author = clean(sheet.cell(row, 4).value)
    title = clean(sheet.cell(row, 5).value)
    content = clean(sheet.cell(row, 6).value)
    if author.lower() in {"họ và tên tác giả chính", "họ và tên tác giả"}:
        return False
    return bool(author and title and content)


def parse_sheet(workbook_path: Path, sheet_name: str, year: int) -> list[dict]:
    workbook = load_workbook(workbook_path, read_only=False, data_only=True, keep_links=False)
    sheet = workbook[sheet_name]
    rows: list[dict] = []
    current_registration_month: int | None = None
    khmt_column = SHEET_KHMT_COLUMN.get(sheet_name, 15)
    leader_column = SHEET_LEADER_CONCLUSION_COLUMN.get(sheet_name, 14)
    for row in range(1, sheet.max_row + 1):
        candidate_month = parse_month(clean(sheet.cell(row, 1).value))
        if candidate_month is not None:
            current_registration_month = candidate_month
        if not has_data(sheet, row):
            continue
        month_raw = clean(sheet.cell(row, 1).value)
        review = clean(sheet.cell(row, 13).value)
        note_raw = clean(sheet.cell(row, 12).value)
        khmt_raw = clean(sheet.cell(row, khmt_column).value)
        completion_plan = clean(sheet.cell(row, 11).value)
        khmt_month = parse_month(khmt_raw)
        registration_month = parse_month(month_raw) or khmt_month or parse_month(note_raw) or current_registration_month
        if registration_month is not None:
            current_registration_month = registration_month
        rows.append(
            {
                "source_row": row,
                "month_raw": month_raw,
                "registration_month": registration_month,
                "registration_year": year,
                "author_name": clean(sheet.cell(row, 4).value),
                "title": clean(sheet.cell(row, 5).value),
                "content_description": clean(sheet.cell(row, 6).value),
                "completion_plan": completion_plan,
                "review": review,
                "leader_conclusion": clean(sheet.cell(row, leader_column).value) if leader_column else "",
                "khmt_raw": khmt_raw,
                "khmt_month": khmt_month,
                "status": status_from_review(review),
            }
        )
    return rows


def import_rows(
    rows: list[dict],
    *,
    source_label: str,
    sheet_name: str,
    team: str,
    year: int,
    imported_by: str,
) -> tuple[int, int]:
    inserted = 0
    updated = 0
    with create_session() as db:
        for item in rows:
            sk_code = f"HIST-{team}-{sheet_name}-{item['source_row']}"
            record = db.execute(
                select(SKCTKTModel).where(
                    or_(
                        SKCTKTModel.sk_code == sk_code,
                        and_(
                            SKCTKTModel.bm01_source_file == source_label,
                            SKCTKTModel.bm01_source_sheet == sheet_name,
                            SKCTKTModel.bm01_source_row == item["source_row"],
                        ),
                    )
                )
            ).scalar_one_or_none()
            created_at = datetime(year, item["registration_month"], 1, tzinfo=timezone.utc)
            now = datetime.now(timezone.utc)
            is_approved = item["status"] in {SKStatus.APPROVED.value, SKStatus.COMPLETED.value}
            history = [
                {
                    "from_status": "Legacy",
                    "to_status": item["status"],
                    "changed_by": imported_by,
                    "changed_at": now.isoformat(),
                    "reason": item["review"] or "Legacy BM01 chưa có dữ liệu xét duyệt",
                    "comments": {
                        "registration_month": item["registration_month"],
                        "registration_year": year,
                        "month_raw": item["month_raw"],
                        "khmt_raw": item["khmt_raw"],
                    },
                }
            ]
            values = {
                "sk_code": sk_code,
                "title": item["title"],
                "author_name": item["author_name"],
                "author_user_id": "historical-import",
                "team": team,
                "content_description": item["content_description"],
                "completion_plan": item["completion_plan"],
                "status": item["status"],
                "status_history": history,
                "fi_coordinator_comments": item["review"] or None,
                "workshop_leader_conclusion": item["leader_conclusion"] or None,
                "decision_note": None,
                "consider_for_khmt": bool(is_approved and item["khmt_month"]),
                "khmt_month": item["khmt_month"] if is_approved else None,
                "khmt_year": year if is_approved and item["khmt_month"] else None,
                "is_public": is_approved,
                "is_counted_for_okr": bool(is_approved and item["khmt_month"]),
                "is_historical_import": True,
                "bm01_source_file": source_label,
                "bm01_source_sheet": sheet_name,
                "bm01_source_row": item["source_row"],
                "bm01_raw_conclusion": item["review"],
                "created_at": created_at,
                "updated_at": now,
                "submitted_at": created_at,
                "reviewed_at": now
                if item["status"] in {SKStatus.APPROVED.value, SKStatus.REJECTED.value, SKStatus.DEFERRED.value}
                else None,
                "approved_at": now if item["status"] == SKStatus.APPROVED.value else None,
                "completed_at": None,
            }
            if record is None:
                db.add(SKCTKTModel(id=make_id("sk"), **values))
                inserted += 1
            else:
                for key, value in values.items():
                    setattr(record, key, value)
                updated += 1
        db.commit()
    return inserted, updated


def main() -> int:
    parser = argparse.ArgumentParser(description="Import one legacy BM01 sheet into SK-CTKT records")
    parser.add_argument("workbook")
    parser.add_argument("--sheet", required=True, choices=sorted(SHEET_TEAM))
    parser.add_argument("--year", type=int, default=2026)
    parser.add_argument("--expected-count", type=int)
    parser.add_argument("--source-label", default="FI xlsx/BM 01 Dang ky - Danh gia SK _Rev1.xlsx")
    parser.add_argument("--imported-by", default="bm01-legacy-import")
    parser.add_argument("--no-commit", action="store_true")
    args = parser.parse_args()

    team = SHEET_TEAM[args.sheet]
    rows = parse_sheet(Path(args.workbook), args.sheet, args.year)
    if args.expected_count is not None and len(rows) != args.expected_count:
        raise SystemExit(f"Expected {args.expected_count} rows, got {len(rows)}. Abort.")
    missing_month_rows = [item["source_row"] for item in rows if item["registration_month"] is None]
    if missing_month_rows:
        raise SystemExit(f"Rows without registration month: {missing_month_rows}. Abort.")

    if args.no_commit:
        inserted = updated = 0
    else:
        inserted, updated = import_rows(
            rows,
            source_label=args.source_label,
            sheet_name=args.sheet,
            team=team,
            year=args.year,
            imported_by=args.imported_by,
        )

    print(f"Imported {args.sheet} -> {team}: inserted={inserted}, updated={updated}, total={len(rows)}")
    for item in rows:
        print(
            f"row={item['source_row']:02d} "
            f"month=T{item['registration_month']}/{args.year} "
            f"status={item['status']:<9} "
            f"author={item['author_name'][:24]} "
            f"title={item['title'][:70]}"
        )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
