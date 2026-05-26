from io import BytesIO
import json

from openpyxl import Workbook, load_workbook

from app.db.session import create_session
from app.models.domain import SKCTKTModel


def _login(client, user_id: str, password: str) -> dict[str, str]:
    response = client.post("/api/v1/auth/login", json={"user_id": user_id, "password": password})
    assert response.status_code == 200, response.text
    return {"Authorization": f"Bearer {response.json()['access_token']}"}


def _create_user(client, admin_headers, user_id: str, role: str = "Team_Account", team: str = "TBCH") -> dict[str, str]:
    response = client.post(
        "/api/v1/admin/users",
        headers=admin_headers,
        json={
            "id": user_id,
            "display_name": user_id,
            "full_name": user_id,
            "password": "pw",
            "role": role,
            "team": team,
        },
    )
    assert response.status_code == 200, response.text
    return _login(client, user_id, "pw")


def _xlsx_bytes(title: str = "TBCH") -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = title
    ws["N1"] = "Tình hình thực hiện"
    ws["O1"] = "KR Assessment"
    ws["P1"] = "Notes"
    ws["B2"] = "O2.KR1"
    ws["N2"] = "Hoàn thành 10/12 hạng mục đạt 83%"
    ws["O2"] = "Hoàn thành"
    ws["P2"] = "Tồn khách quan"
    ws["B3"] = "O5.KR13"
    ws["N3"] = "2 sáng kiến được công nhận"
    ws["O3"] = "Hoàn thành"
    ws["B5"] = "Kỷ luật"
    ws["C5"] = "OK"
    ws["B6"] = "Đánh giá chung"
    ws["C6"] = "Hoàn thành"
    output = BytesIO()
    wb.save(output)
    return output.getvalue()


def test_auth_rejects_no_token_and_invalid_login(client):
    assert client.get("/api/v1/admin/users").status_code == 401
    response = client.post("/api/v1/auth/login", json={"user_id": "admin", "password": "wrong"})
    assert response.status_code == 401


def test_guest_test_login_is_admin_sandbox_and_does_not_write_production(client):
    response = client.post("/api/v1/auth/login", json={"user_id": "test", "password": "PVCFC@123"})
    assert response.status_code == 200, response.text
    payload = response.json()
    assert payload["role"] == "Admin"
    assert payload["display_name"] == "Khách kiểm thử - Quản trị"
    headers = {"Authorization": f"Bearer {payload['access_token']}"}

    profile = client.get("/api/v1/auth/me", headers=headers)
    assert profile.status_code == 200, profile.text
    assert profile.json()["sandbox"] is True
    assert profile.json()["user_id"] == "test"

    with create_session() as db:
        before_count = db.query(SKCTKTModel).count()

    created = client.post(
        "/api/v1/fi/sk-ctkt",
        headers=headers,
        json={
            "author_name": "Khách kiểm thử",
            "team": "TBCH",
            "title": "Sandbox only",
            "content_description": "Không ghi vào production",
            "completion_plan": "T6/2026",
        },
    )
    assert created.status_code == 200, created.text

    with create_session() as db:
        assert db.query(SKCTKTModel).count() == before_count
        assert db.get(SKCTKTModel, created.json()["id"]) is None


def test_user_cannot_submit_another_users_sk(client, admin_headers):
    user_one = _create_user(client, admin_headers, "u1")
    user_two = _create_user(client, admin_headers, "u2")
    created = client.post(
        "/api/v1/fi/sk-ctkt",
        headers=user_one,
        json={
            "author_name": "A",
            "team": "TBCH",
            "title": "T",
            "content_description": "C",
            "completion_plan": "P",
        },
    )
    assert created.status_code == 200
    record_id = created.json()["id"]
    assert client.post(f"/api/v1/fi/sk-ctkt/{record_id}/submit", headers=user_two, json={}).status_code == 403
    assert client.post(f"/api/v1/fi/sk-ctkt/{record_id}/submit", headers=user_one, json={}).status_code == 200


def test_team_account_create_is_locked_to_own_team(client):
    team_headers = _login(client, "TBCH", "tbch-pass")
    created = client.post(
        "/api/v1/fi/sk-ctkt",
        headers=team_headers,
        json={
            "author_name": "Đội TBCH",
            "team": "TBĐL",
            "author_user_id": "someone-else",
            "title": "Không được khai đội khác",
            "content_description": "Nội dung",
            "completion_plan": "T6/2026",
        },
    )
    assert created.status_code == 200, created.text
    assert created.json()["team"] == "TBCH"
    assert created.json()["author_user_id"] == "TBCH"
    assert created.json()["author_name"] == "TBCH"
    assert created.json()["sk_code"].startswith("FI/")
    assert "-TBCH-" in created.json()["sk_code"]


def test_fi_team_draft_is_private_until_submitted(client, admin_headers):
    team_headers = _login(client, "TBCH", "tbch-pass")
    fi_headers = _login(client, "fi", "fi-pass")
    leader_headers = _login(client, "leader", "leader-pass")
    created = client.post(
        "/api/v1/fi/sk-ctkt",
        headers=team_headers,
        json={
            "author_name": "Đội TBCH",
            "team": "TBCH",
            "title": "Bản nháp chỉ đội thấy",
            "content_description": "Nội dung đang soạn",
            "completion_plan": "T6/2026",
        },
    )
    assert created.status_code == 200, created.text
    record_id = created.json()["id"]

    admin_list = client.get("/api/v1/fi/sk-ctkt", headers=admin_headers)
    assert admin_list.status_code == 200, admin_list.text
    assert record_id not in {item["id"] for item in admin_list.json()}
    assert client.get(f"/api/v1/fi/sk-ctkt/{record_id}", headers=admin_headers).status_code == 403
    public_draft = client.get("/api/v1/fi/sk-ctkt/public", headers=team_headers)
    assert public_draft.status_code == 200
    assert record_id not in {item["id"] for item in public_draft.json()}

    team_list = client.get("/api/v1/fi/sk-ctkt", headers=team_headers)
    assert record_id in {item["id"] for item in team_list.json()}
    submitted = client.post(f"/api/v1/fi/sk-ctkt/{record_id}/submit", headers=team_headers, json={})
    assert submitted.status_code == 200, submitted.text

    admin_list_after_submit = client.get("/api/v1/fi/sk-ctkt", headers=admin_headers)
    assert record_id in {item["id"] for item in admin_list_after_submit.json()}
    fi_list_after_submit = client.get("/api/v1/fi/sk-ctkt", headers=fi_headers)
    assert record_id in {item["id"] for item in fi_list_after_submit.json()}
    leader_list_before_approve = client.get("/api/v1/fi/sk-ctkt", headers=leader_headers)
    assert record_id in {item["id"] for item in leader_list_before_approve.json()}
    assert client.get(f"/api/v1/fi/sk-ctkt/{record_id}", headers=leader_headers).status_code == 200
    public_after_submit = client.get("/api/v1/fi/sk-ctkt/public", headers=team_headers)
    assert record_id in {item["id"] for item in public_after_submit.json()}

    approved = client.post(f"/api/v1/fi/sk-ctkt/{record_id}/approve", headers=fi_headers, json={"note": "Đủ điều kiện"})
    assert approved.status_code == 200, approved.text
    assert approved.json()["decision_note"] == "Đủ điều kiện"
    assert approved.json()["workshop_leader_conclusion"] is None
    leader_list_after_approve = client.get("/api/v1/fi/sk-ctkt", headers=leader_headers)
    assert record_id in {item["id"] for item in leader_list_after_approve.json()}
    assert client.get(f"/api/v1/fi/sk-ctkt/{record_id}", headers=leader_headers).status_code == 200


def test_fi_reject_requires_note_and_delete_is_admin_only(client, admin_headers):
    team_headers = _login(client, "TBĐL", "tbdl-pass")
    fi_headers = _login(client, "fi", "fi-pass")
    created = client.post(
        "/api/v1/fi/sk-ctkt",
        headers=team_headers,
        json={
            "author_name": "Đội TBĐL",
            "team": "TBĐL",
            "title": "Cần FI xử lý",
            "content_description": "Nội dung",
            "completion_plan": "T6/2026",
        },
    )
    assert created.status_code == 200, created.text
    record_id = created.json()["id"]
    assert client.post(f"/api/v1/fi/sk-ctkt/{record_id}/submit", headers=team_headers, json={}).status_code == 200

    missing_note = client.post(f"/api/v1/fi/sk-ctkt/{record_id}/reject", headers=fi_headers, json={})
    assert missing_note.status_code == 400
    rejected = client.post(f"/api/v1/fi/sk-ctkt/{record_id}/reject", headers=fi_headers, json={"note": "Trùng ý tưởng"})
    assert rejected.status_code == 200, rejected.text
    assert rejected.json()["status"] == "Rejected"
    assert rejected.json()["decision_note"] == "Trùng ý tưởng"
    assert rejected.json()["workshop_leader_conclusion"] is None

    assert client.delete(f"/api/v1/fi/sk-ctkt/{record_id}", headers=team_headers).status_code == 403
    assert client.delete(f"/api/v1/fi/sk-ctkt/{record_id}", headers=fi_headers).status_code == 403
    deleted = client.delete(f"/api/v1/fi/sk-ctkt/{record_id}", headers=admin_headers)
    assert deleted.status_code == 200
    assert deleted.json() == {"deleted": True}


def test_team_account_can_delete_own_draft_only(client):
    team_headers = _login(client, "TBCH", "tbch-pass")
    other_team_headers = _login(client, "TBĐL", "tbdl-pass")
    created = client.post(
        "/api/v1/fi/sk-ctkt",
        headers=team_headers,
        json={
            "author_name": "Đội TBCH",
            "team": "TBCH",
            "title": "Có thể xóa bản nháp",
            "content_description": "Nội dung",
            "completion_plan": "T6/2026",
        },
    )
    assert created.status_code == 200, created.text
    record_id = created.json()["id"]

    other_delete = client.delete(f"/api/v1/fi/sk-ctkt/{record_id}", headers=other_team_headers)
    assert other_delete.status_code == 403

    own_delete = client.delete(f"/api/v1/fi/sk-ctkt/{record_id}", headers=team_headers)
    assert own_delete.status_code == 200
    assert own_delete.json() == {"deleted": True}
    assert client.get(f"/api/v1/fi/sk-ctkt/{record_id}", headers=team_headers).status_code == 404

    submitted = client.post(
        "/api/v1/fi/sk-ctkt",
        headers=team_headers,
        json={
            "author_name": "Đội TBCH",
            "team": "TBCH",
            "title": "Không xóa sau khi gửi",
            "content_description": "Nội dung",
            "completion_plan": "T6/2026",
        },
    )
    submitted_id = submitted.json()["id"]
    assert client.post(f"/api/v1/fi/sk-ctkt/{submitted_id}/submit", headers=team_headers, json={}).status_code == 200
    assert client.delete(f"/api/v1/fi/sk-ctkt/{submitted_id}", headers=team_headers).status_code == 403


def test_public_sk_is_internal_and_excludes_drafts(client):
    team_headers = _login(client, "TCĐK", "tcdk-pass")
    other_team_headers = _login(client, "TBCH", "tbch-pass")
    created = client.post(
        "/api/v1/fi/sk-ctkt",
        headers=team_headers,
        json={
            "author_name": "Tổ TCĐK",
            "team": "TCĐK",
            "title": "Public internal check",
            "content_description": "Nội dung",
            "completion_plan": "T6/2026",
        },
    )
    assert created.status_code == 200, created.text
    record_id = created.json()["id"]
    assert client.get("/api/v1/fi/sk-ctkt/public").status_code == 401
    assert record_id not in {item["id"] for item in client.get("/api/v1/fi/sk-ctkt/public", headers=team_headers).json()}
    assert client.post(f"/api/v1/fi/sk-ctkt/{record_id}/submit", headers=team_headers, json={}).status_code == 200
    assert record_id in {item["id"] for item in client.get("/api/v1/fi/sk-ctkt/public", headers=team_headers).json()}
    cross_team_detail = client.get(f"/api/v1/fi/sk-ctkt/{record_id}", headers=other_team_headers)
    assert cross_team_detail.status_code == 200, cross_team_detail.text
    assert cross_team_detail.json()["team"] == "TCĐK"


def test_legacy_sk_is_history_and_can_be_reviewed_from_history(client, admin_headers):
    team_headers = _login(client, "TBCH", "tbch-pass")
    other_team_headers = _login(client, "TBĐL", "tbdl-pass")
    fi_headers = _login(client, "fi", "fi-pass")
    with create_session() as db:
        db.add(
            SKCTKTModel(
                id="sk-legacy",
                sk_code="HIST-TBCH-TBCH-99",
                title="Legacy BM01",
                author_name="Tác giả cũ",
                author_user_id="historical-import",
                team="TBCH",
                content_description="Nội dung legacy",
                completion_plan="T6/2026",
                status="Submitted",
                status_history=[],
                is_public=False,
                is_counted_for_okr=False,
                is_historical_import=True,
                bm01_source_file="FI xlsx/BM 01 Dang ky - Danh gia SK _Rev1.xlsx",
                bm01_source_sheet="TBCH",
                bm01_source_row=99,
                bm01_raw_conclusion="",
            )
        )
        db.commit()

    processing = client.get("/api/v1/fi/sk-ctkt", headers=team_headers)
    assert processing.status_code == 200
    assert "sk-legacy" not in {item["id"] for item in processing.json()}

    processing_with_history = client.get("/api/v1/fi/sk-ctkt?include_historical=true", headers=team_headers)
    assert "sk-legacy" in {item["id"] for item in processing_with_history.json()}

    legacy_history = client.get("/api/v1/fi/sk-ctkt/public?historical=true&team=TBCH", headers=team_headers)
    assert legacy_history.status_code == 200
    assert "sk-legacy" in {item["id"] for item in legacy_history.json()}

    current_public = client.get("/api/v1/fi/sk-ctkt/public?historical=false&team=TBCH", headers=team_headers)
    assert "sk-legacy" not in {item["id"] for item in current_public.json()}

    content_update = client.put(
        "/api/v1/fi/sk-ctkt/sk-legacy",
        headers=team_headers,
        json={"content_description": "Nội dung đã cập nhật", "completion_plan": "Dự kiến 07/2026"},
    )
    assert content_update.status_code == 403

    protected_update = client.put(
        "/api/v1/fi/sk-ctkt/sk-legacy",
        headers=team_headers,
        json={"title": "Không được đổi tên"},
    )
    assert protected_update.status_code == 403

    transition = client.post("/api/v1/fi/sk-ctkt/sk-legacy/approve", headers=fi_headers, json={})
    assert transition.status_code == 200
    assert transition.json()["status"] == "Approved"
    assert transition.json()["is_historical_import"] is True

    foreign_khmt = client.post(
        "/api/v1/fi/sk-ctkt/sk-legacy/assign-khmt",
        headers=other_team_headers,
        json={"month": 7, "year": 2026},
    )
    assert foreign_khmt.status_code == 403

    khmt = client.post(
        "/api/v1/fi/sk-ctkt/sk-legacy/assign-khmt",
        headers=team_headers,
        json={"month": 7, "year": 2026},
    )
    assert khmt.status_code == 200, khmt.text
    assert khmt.json()["consider_for_khmt"] is True
    assert khmt.json()["khmt_month"] == 7


def test_okr_duplicate_requires_confirmation_and_export_is_valid_xlsx(client, admin_headers):
    data = _xlsx_bytes()
    response = client.post(
        "/api/v1/okr/reports/upload",
        headers=admin_headers,
        files={"file": ("Bao cao TBCH T4 2026.xlsx", data, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert response.status_code == 200, response.text
    duplicate = client.post(
        "/api/v1/okr/reports/upload",
        headers=admin_headers,
        files={"file": ("Bao cao TBCH T4 2026.xlsx", data, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert duplicate.status_code == 409
    confirmed = client.post(
        "/api/v1/okr/reports/upload",
        headers=admin_headers,
        data={"confirm_replace": "true"},
        files={"file": ("Bao cao TBCH T4 2026.xlsx", data, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert confirmed.status_code == 200
    export = client.post("/api/v1/okr/dashboard/export", headers=admin_headers)
    assert export.status_code == 200, export.text
    wb = load_workbook(BytesIO(export.content), read_only=True)
    assert {"data", "Dashboard"}.issubset(wb.sheetnames)


def test_json_endpoints_serialize_nested_dates_and_json_fields(client, admin_headers):
    data = _xlsx_bytes()
    upload = client.post(
        "/api/v1/okr/reports/upload",
        headers=admin_headers,
        files={"file": ("Bao cao TBCH T4 2026.xlsx", data, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert upload.status_code == 200, upload.text

    endpoints = [
        "/api/v1/okr/reports",
        "/api/v1/okr/warnings",
        "/api/v1/okr/dashboard/4/2026",
        "/api/v1/admin/audit-log",
        "/api/v1/notifications",
    ]
    for endpoint in endpoints:
        response = client.get(endpoint, headers=admin_headers)
        assert response.status_code == 200, response.text
        json.dumps(response.json(), ensure_ascii=False)

    assert upload.json()["team_level"]["discipline_status"] == "OK"


def test_warnings_can_be_scoped_to_current_report_period(client, admin_headers):
    for month in (3, 4):
        upload = client.post(
            "/api/v1/okr/reports/upload",
            headers=admin_headers,
            data={"month": str(month), "year": "2026"},
            files={
                "file": (
                    f"Bao cao TBCH T{month} 2026.xlsx",
                    _xlsx_bytes(),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
        assert upload.status_code == 200, upload.text

    reports = client.get("/api/v1/okr/reports", headers=admin_headers).json()
    t4_report = next(report for report in reports if report["team"] == "TBCH" and report["report_month"] == 4)
    t3_report = next(report for report in reports if report["team"] == "TBCH" and report["report_month"] == 3)

    t4_warnings = client.get("/api/v1/okr/warnings?month=4&year=2026", headers=admin_headers)
    assert t4_warnings.status_code == 200, t4_warnings.text
    assert t4_warnings.json()
    assert {warning["team_report_id"] for warning in t4_warnings.json()} == {t4_report["id"]}

    all_warnings = client.get("/api/v1/okr/warnings", headers=admin_headers).json()
    assert {t3_report["id"], t4_report["id"]}.issubset({warning["team_report_id"] for warning in all_warnings})


def test_team_account_can_read_dashboard_but_not_management_endpoints(client):
    user_headers = _login(client, "TBCH", "tbch-pass")

    dashboard = client.get("/api/v1/okr/dashboard/4/2026", headers=user_headers)
    assert dashboard.status_code == 200, dashboard.text
    assert [team["team"] for team in dashboard.json()["teams"]]

    template = client.get("/api/v1/okr/reports/template", headers=user_headers)
    assert template.status_code == 200, template.text
    workbook = load_workbook(BytesIO(template.content), read_only=True)
    assert "Team_Report" in workbook.sheetnames

    assert client.get("/api/v1/okr/reports", headers=user_headers).status_code == 403
    assert client.get("/api/v1/okr/warnings", headers=user_headers).status_code == 403
    assert client.get("/api/v1/admin/kr-mapping", headers=user_headers).status_code == 403


def test_sk_image_lifecycle_is_separate_from_create_schema(client, admin_headers):
    user_headers = _create_user(client, admin_headers, "image-user")
    created = client.post(
        "/api/v1/fi/sk-ctkt",
        headers=user_headers,
        json={
            "author_name": "Image User",
            "team": "TBCH",
            "title": "SK with image",
            "content_description": "Content",
            "completion_plan": "T6/2026",
        },
    )
    assert created.status_code == 200, created.text
    record_id = created.json()["id"]
    assert "supporting_images" not in created.json()

    rejected_upload = client.post(
        f"/api/v1/fi/sk-ctkt/{record_id}/images",
        headers=user_headers,
        files={"file": ("not-image.txt", b"not an image", "text/plain")},
    )
    assert rejected_upload.status_code == 400

    uploaded = client.post(
        f"/api/v1/fi/sk-ctkt/{record_id}/images",
        headers=user_headers,
        files={"file": ("evidence.JPG", b"fake-jpg", "application/octet-stream")},
    )
    assert uploaded.status_code == 200, uploaded.text
    image_id = uploaded.json()["id"]
    assert uploaded.json()["content_type"] == "image/jpeg"

    listed = client.get(f"/api/v1/fi/sk-ctkt/{record_id}/images", headers=user_headers)
    assert listed.status_code == 200
    assert [image["id"] for image in listed.json()] == [image_id]

    detail = client.get(f"/api/v1/fi/sk-ctkt/{record_id}", headers=user_headers)
    assert detail.status_code == 200
    assert detail.json()["supporting_images"][0]["id"] == image_id
    json.dumps(detail.json(), ensure_ascii=False)

    deleted = client.delete(f"/api/v1/fi/sk-ctkt/{record_id}/images/{image_id}", headers=user_headers)
    assert deleted.status_code == 200
    assert deleted.json() == {"deleted": True}
    assert client.get(f"/api/v1/fi/sk-ctkt/{record_id}/images", headers=user_headers).json() == []


def test_team_uploaded_sk_image_is_visible_to_fi_after_submit(client):
    team_headers = _login(client, "TBCH", "tbch-pass")
    fi_headers = _login(client, "fi", "fi-pass")
    created = client.post(
        "/api/v1/fi/sk-ctkt",
        headers=team_headers,
        json={
            "author_name": "Đội TBCH",
            "team": "TBCH",
            "title": "FI xem được ảnh",
            "content_description": "Nội dung",
            "completion_plan": "T6/2026",
        },
    )
    assert created.status_code == 200, created.text
    record_id = created.json()["id"]
    image_bytes = b"image-bytes"

    uploaded = client.post(
        f"/api/v1/fi/sk-ctkt/{record_id}/images",
        headers=team_headers,
        files={"file": ("evidence.png", image_bytes, "image/png")},
    )
    assert uploaded.status_code == 200, uploaded.text
    image_id = uploaded.json()["id"]
    assert client.post(f"/api/v1/fi/sk-ctkt/{record_id}/submit", headers=team_headers, json={}).status_code == 200

    detail = client.get(f"/api/v1/fi/sk-ctkt/{record_id}", headers=fi_headers)
    assert detail.status_code == 200, detail.text
    assert [image["id"] for image in detail.json()["supporting_images"]] == [image_id]

    raw = client.get(f"/api/v1/fi/sk-ctkt/{record_id}/images/{image_id}/raw", headers=fi_headers)
    assert raw.status_code == 200, raw.text
    assert raw.content == image_bytes

    fi_upload = client.post(
        f"/api/v1/fi/sk-ctkt/{record_id}/images",
        headers=fi_headers,
        files={"file": ("fi-evidence.png", b"fi-image", "image/png")},
    )
    assert fi_upload.status_code == 403
    assert client.delete(f"/api/v1/fi/sk-ctkt/{record_id}/images/{image_id}", headers=fi_headers).status_code == 403
