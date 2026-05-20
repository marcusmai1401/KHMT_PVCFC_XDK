from __future__ import annotations

from dataclasses import asdict, dataclass, field
from datetime import datetime, timezone
import argparse
import hashlib
from pathlib import Path
import re
import threading
from typing import Any

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.core.config import settings
from app.db.session import create_session
from app.models.domain import HistoricalSnapshotModel, TeamMonthlySummaryModel, TeamReportModel
from app.services.okr.constants import TEAMS
from app.services.okr.historical_snapshot import (
    _dashboard_month_columns,
    find_dashboard_team_rows,
    import_historical_snapshot,
)
from app.services.okr.kr_mapping import KRMapping, mapping_by_code
from app.services.okr.rules import normalize_assessment
from app.services.okr.workbook import parse_team_report
from app.services.repositories import json_safe, make_id, warning_from_dict


SOURCE_FILE_PATTERN = re.compile(r"OKR tháng (\d{2})-(\d{4})", re.IGNORECASE)
HISTORICAL_YEAR = 2026
HISTORICAL_MONTHS = (1, 2, 3, 4)
TEMPLATE_FILES = {
    "TBHTĐK": "TBHTĐK.xlsx",
    "TBCH": "TBCH.xlsx",
    "TBĐL": "TBĐL.xlsx",
    "TCĐK": "TCĐK.xlsx",
}
T4_DISCIPLINE_OVERRIDES = {
    "TBĐL": {
        "discipline_status": "NOK",
        "discipline_description": "Một nhân sự Đội TBĐL không tuân thủ quy định giờ công",
        "monthly_assessment": "Không hoàn thành",
    },
    "TBCH": {
        "discipline_status": "NOK",
        "discipline_description": (
            "Một nhân sự Đội TBCH không tuân thủ đúng HDBD trong quá trình thực hiện công "
            "việc bảo dưỡng định kỳ thiết bị Quan trắc"
        ),
        "monthly_assessment": "Không hoàn thành",
    },
}


@dataclass(frozen=True)
class DiscoveredFile:
    path: Path
    month: int
    year: int
    file_name: str


@dataclass
class ImportIssue:
    source_file: str
    reason: str
    severity: str = "MEDIUM"
    sheet: str | None = None
    row: int | None = None
    column: str | None = None
    source_range: str | None = None
    error_type: str = "IMPORT_WARNING"

    def to_dict(self) -> dict[str, Any]:
        return asdict(self)


@dataclass
class FileImportResult:
    file_name: str
    month: int
    year: int
    teams_imported: list[str] = field(default_factory=list)
    records_per_team: dict[str, int] = field(default_factory=dict)
    table_counts: dict[str, dict[str, int]] = field(
        default_factory=lambda: {
            "team_reports": {"inserted": 0, "updated": 0, "skipped": 0, "failed": 0},
            "historical_snapshots": {"inserted": 0, "updated": 0, "skipped": 0, "failed": 0},
            "team_monthly_summaries": {"inserted": 0, "updated": 0, "skipped": 0, "failed": 0},
        }
    )
    rows_skipped: int = 0
    success: bool = True
    errors: list[ImportIssue] = field(default_factory=list)
    warnings: list[ImportIssue] = field(default_factory=list)

    def to_dict(self) -> dict[str, Any]:
        data = asdict(self)
        data["errors"] = [issue.to_dict() for issue in self.errors]
        data["warnings"] = [issue.to_dict() for issue in self.warnings]
        return json_safe(data)


@dataclass
class ImportSessionReport:
    file_results: list[FileImportResult] = field(default_factory=list)
    total_team_reports: int = 0
    total_snapshots: int = 0
    total_summaries_upserted: int = 0
    total_files_attempted: int = 0
    total_files_successful: int = 0
    warnings: list[ImportIssue] = field(default_factory=list)
    errors: list[ImportIssue] = field(default_factory=list)
    completeness: dict[str, Any] = field(default_factory=dict)
    complete: bool = False

    def to_dict(self) -> dict[str, Any]:
        return {
            "total_team_reports": self.total_team_reports,
            "total_snapshots": self.total_snapshots,
            "total_summaries_upserted": self.total_summaries_upserted,
            "total_files_attempted": self.total_files_attempted,
            "total_files_successful": self.total_files_successful,
            "complete": self.complete,
            "completeness": json_safe(self.completeness),
            "warnings": [issue.to_dict() for issue in self.warnings],
            "errors": [issue.to_dict() for issue in self.errors],
            "file_results": [result.to_dict() for result in self.file_results],
        }

    def render_text(self) -> str:
        lines = [
            "Historical import report",
            f"Files: {self.total_files_successful}/{self.total_files_attempted} successful",
            f"Team reports: {self.total_team_reports}",
            f"Historical snapshots: {self.total_snapshots}",
            f"Monthly summaries upserted: {self.total_summaries_upserted}",
            f"Complete: {'yes' if self.complete else 'no'}",
        ]
        for result in self.file_results:
            lines.append(
                f"- {result.file_name}: teams={','.join(result.teams_imported) or '-'} "
                f"reports={sum(result.table_counts['team_reports'].values())} "
                f"snapshots={result.table_counts['historical_snapshots']['inserted']} "
                f"warnings={len(result.warnings)} errors={len(result.errors)}"
            )
        if self.warnings:
            lines.append("Warnings:")
            for issue in self.warnings[:30]:
                location = issue.source_range or issue.sheet or issue.source_file
                lines.append(f"- [{issue.severity}] {location}: {issue.reason}")
            if len(self.warnings) > 30:
                lines.append(f"- ... {len(self.warnings) - 30} more warnings")
        if self.errors:
            lines.append("Errors:")
            for issue in self.errors:
                location = issue.source_range or issue.sheet or issue.source_file
                lines.append(f"- [{issue.severity}] {location}: {issue.reason}")
        return "\n".join(lines)


def extract_month_year_from_filename(filename: str) -> tuple[int, int]:
    match = SOURCE_FILE_PATTERN.search(Path(filename).name)
    if not match:
        raise ValueError(f"Malformed historical source filename: {filename}")
    month = int(match.group(1))
    year = int(match.group(2))
    if not 1 <= month <= 12:
        raise ValueError(f"Invalid month in historical source filename: {filename}")
    return month, year


def discover_source_files(directory: Path) -> list[DiscoveredFile]:
    discovered: list[DiscoveredFile] = []
    if not directory.exists():
        return discovered
    for path in sorted(directory.glob("*.xlsx")):
        try:
            month, year = extract_month_year_from_filename(path.name)
        except ValueError:
            continue
        if year == HISTORICAL_YEAR and month in HISTORICAL_MONTHS:
            discovered.append(DiscoveredFile(path=path, month=month, year=year, file_name=path.name))
    return discovered


def read_workbook_with_timeout(file_path: Path, timeout: int = 30):
    result: list[Any] = [None]
    error: list[BaseException | None] = [None]

    def _read() -> None:
        try:
            result[0] = load_workbook(file_path, read_only=True, data_only=True, keep_links=False)
        except BaseException as exc:  # pragma: no cover - defensive passthrough
            error[0] = exc

    thread = threading.Thread(target=_read, daemon=True)
    thread.start()
    thread.join(timeout=timeout)
    if thread.is_alive():
        # Python cannot safely terminate a worker thread here; it is daemonized so a hung
        # workbook read cannot keep the CLI process alive after this timeout path.
        raise TimeoutError(f"Reading {file_path.name} exceeded {timeout}s")
    if error[0]:
        raise error[0]
    return result[0]


def resolve_kr_mapping(workspace_dir: Path | None = None) -> dict[str, KRMapping]:
    root = workspace_dir or settings.workspace_dir
    candidates = [
        root / "template_xlsx" / "OKR_Workshop.xlsx",
        root / "KHMT_T1_T2_T3_T4" / "OKR tháng 04-2026 - X.ĐK.xlsx",
    ]
    for candidate in candidates:
        if not candidate.exists():
            continue
        return mapping_by_code(candidate, allow_fallback=False)
    raise FileNotFoundError("No canonical KR mapping source found")


def _source_hash(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def _issue_from_warning(source_file: str, warning: dict[str, Any]) -> ImportIssue:
    source_cell = warning.get("source_cell") or {}
    return ImportIssue(
        source_file=source_file,
        sheet=source_cell.get("sheet_name"),
        row=source_cell.get("row"),
        column=source_cell.get("column"),
        source_range=source_cell.get("source_range"),
        reason=str(warning.get("reason") or warning.get("warning_type") or "Import warning"),
        severity=str(warning.get("severity") or "MEDIUM"),
        error_type=str(warning.get("warning_type") or "IMPORT_WARNING"),
    )


def extract_dashboard_history_lookup(file_path: Path) -> dict[tuple[str, int], str]:
    workbook = load_workbook(file_path, read_only=True, data_only=True, keep_links=False)
    try:
        if "Dashboard" not in workbook.sheetnames:
            return {}
        sheet = workbook["Dashboard"]
        values: dict[tuple[str, int], str] = {}
        for row, team in find_dashboard_team_rows(sheet):
            for month, col in _dashboard_month_columns():
                assessment = str(sheet.cell(row, col).value or "").strip()
                if assessment:
                    values[(team, month)] = normalize_assessment(assessment) or assessment
        return values
    finally:
        workbook.close()


def parse_template_report(
    template_path: Path,
    fallback_workbook_path: Path,
    team: str,
    month: int,
    year: int,
    kr_mapping: dict[str, KRMapping],
) -> dict[str, Any]:
    source_path = template_path if template_path.exists() else fallback_workbook_path
    parsed = parse_team_report(source_path, team=team, month=month, year=year, kr_mapping=kr_mapping)
    parsed["team"] = team
    parsed["source_path"] = source_path
    return parsed


def apply_discipline_overrides(team_level: dict[str, Any], team: str, month: int) -> dict[str, Any]:
    updated = dict(team_level or {})
    if month == 4 and team in T4_DISCIPLINE_OVERRIDES:
        updated.update(T4_DISCIPLINE_OVERRIDES[team])
        updated.setdefault("source_cells", {})["discipline_status"] = {
            "sheet_name": "historical_import",
            "row": None,
            "column": None,
            "field_name": "Discipline_Override",
        }
    return updated


def _status_counts(assessments: list[dict[str, Any]]) -> dict[str, int]:
    counts: dict[str, int] = {}
    for assessment in assessments:
        status = str(assessment.get("dashboard_status") or "#N/A")
        counts[status] = counts.get(status, 0) + 1
    return counts


def upsert_team_report(
    db: Session,
    parsed_data: dict[str, Any],
    *,
    source_file: Path,
    source_hash: str,
    imported_by: str,
) -> tuple[TeamReportModel, str]:
    team = parsed_data.get("team")
    month = parsed_data.get("report_month")
    year = parsed_data.get("report_year")
    existing_records = db.execute(
        select(TeamReportModel).where(
            TeamReportModel.team == team,
            TeamReportModel.report_month == month,
            TeamReportModel.report_year == year,
            TeamReportModel.is_current_version.is_(True),
        )
    ).scalars().all()
    version = 1
    replaced_report_id = None
    action = "inserted"
    for existing in existing_records:
        existing.is_current_version = False
        version = max(version, existing.version + 1)
        replaced_report_id = existing.id
        action = "updated"

    report = TeamReportModel(
        id=make_id("report"),
        team=team,
        report_month=month,
        report_year=year,
        file_name=source_file.name,
        file_path=str(source_file),
        file_hash=source_hash,
        version=version,
        replaced_report_id=replaced_report_id,
        is_current_version=True,
        uploaded_by=imported_by,
        uploaded_at=datetime.now(timezone.utc),
        sheet_name=parsed_data.get("sheet_name"),
        validation_status="VALID",
        parsing_status="PARSED",
        team_month_assigned_manually=True,
        assessments=parsed_data.get("assessments", []),
        team_level=parsed_data.get("team_level", {}),
        source_cell_references=parsed_data.get("source_cell_references", []),
        source_type="historical_import",
        report_status="submitted",
        submitted_at=datetime.now(timezone.utc),
    )
    db.add(report)
    db.flush()
    return report, action


def upsert_team_monthly_summary(
    db: Session,
    parsed_data: dict[str, Any],
    *,
    source_file: Path,
    source_hash: str,
) -> tuple[TeamMonthlySummaryModel, str]:
    team = str(parsed_data.get("team"))
    month = int(parsed_data.get("report_month"))
    year = int(parsed_data.get("report_year"))
    team_level = parsed_data.get("team_level") or {}
    monthly_assessment = normalize_assessment(team_level.get("monthly_assessment")) or "Hoàn thành"
    discipline_status = team_level.get("discipline_status") or "OK"
    stats = {
        "assessment_count": len(parsed_data.get("assessments") or []),
        "dashboard_status_counts": _status_counts(parsed_data.get("assessments") or []),
        "source_file": source_file.name,
        "source_file_hash": source_hash,
        "source_cell_references": parsed_data.get("source_cell_references", []),
    }
    existing = db.execute(
        select(TeamMonthlySummaryModel).where(
            TeamMonthlySummaryModel.team == team,
            TeamMonthlySummaryModel.month == month,
            TeamMonthlySummaryModel.year == year,
        )
    ).scalar_one_or_none()
    if existing is None:
        record = TeamMonthlySummaryModel(
            id=make_id("summary"),
            team=team,
            month=month,
            year=year,
            discipline_status=discipline_status,
            discipline_description=team_level.get("discipline_description"),
            related_kr=team_level.get("related_kr"),
            monthly_assessment=monthly_assessment,
            stats=stats,
        )
        db.add(record)
        db.flush()
        return record, "inserted"
    existing.discipline_status = discipline_status
    existing.discipline_description = team_level.get("discipline_description")
    existing.related_kr = team_level.get("related_kr")
    existing.monthly_assessment = monthly_assessment
    existing.stats = stats
    db.flush()
    return existing, "updated"


def upsert_historical_snapshots(
    db: Session,
    file_path: Path,
    *,
    imported_by: str,
) -> dict[str, Any]:
    return import_historical_snapshot(
        db,
        file_path.read_bytes(),
        source_file_name=file_path.name,
        imported_by=imported_by,
    )


def _apply_dashboard_fallback(parsed: dict[str, Any], dashboard_history: dict[tuple[str, int], str]) -> None:
    team = parsed.get("team")
    month = parsed.get("report_month")
    if team not in TEAMS or not month:
        return
    team_level = parsed.setdefault("team_level", {})
    source_cells = team_level.get("source_cells") or {}
    monthly_source = source_cells.get("monthly_assessment") or {}
    if monthly_source.get("inferred_from") == "KR dashboard statuses" and (team, month) in dashboard_history:
        team_level["monthly_assessment"] = dashboard_history[(team, month)]
        source_cells["monthly_assessment"] = {
            "sheet_name": "Dashboard",
            "row": None,
            "column": None,
            "field_name": "Monthly_Assessment",
            "source": "Dashboard_History",
        }
        team_level["source_cells"] = source_cells


def _record_parsed_report(
    db: Session,
    result: FileImportResult,
    parsed: dict[str, Any],
    *,
    source_path: Path,
    imported_by: str,
) -> None:
    source_hash = _source_hash(source_path)
    parsed["team_level"] = apply_discipline_overrides(
        parsed.get("team_level") or {},
        str(parsed.get("team")),
        int(parsed.get("report_month") or 0),
    )
    try:
        report, report_action = upsert_team_report(
            db,
            parsed,
            source_file=source_path,
            source_hash=source_hash,
            imported_by=imported_by,
        )
        result.table_counts["team_reports"][report_action] += 1
        _summary, summary_action = upsert_team_monthly_summary(
            db,
            parsed,
            source_file=source_path,
            source_hash=source_hash,
        )
        result.table_counts["team_monthly_summaries"][summary_action] += 1
        result.teams_imported.append(str(parsed.get("team")))
        result.records_per_team[str(parsed.get("team"))] = len(parsed.get("assessments") or [])
        for warning in parsed.get("warnings", []):
            warning_from_dict(db, report.id, warning)
            result.warnings.append(_issue_from_warning(source_path.name, warning))
    except Exception as exc:
        result.success = False
        result.table_counts["team_reports"]["failed"] += 1
        result.errors.append(
            ImportIssue(
                source_file=source_path.name,
                severity="HIGH",
                reason=f"Failed to store team report {parsed.get('team')}: {exc}",
                error_type="STORAGE_FAILURE",
            )
        )


def _verify_completeness(db: Session, *, year: int = HISTORICAL_YEAR) -> dict[str, Any]:
    report_rows = db.execute(
        select(TeamReportModel.team, TeamReportModel.report_month).where(
            TeamReportModel.is_current_version.is_(True),
            TeamReportModel.report_status.in_(["submitted", "locked"]),
            TeamReportModel.report_year == year,
            TeamReportModel.report_month.in_(HISTORICAL_MONTHS),
        )
    ).all()
    summary_rows = db.execute(
        select(TeamMonthlySummaryModel.team, TeamMonthlySummaryModel.month).where(
            TeamMonthlySummaryModel.year == year,
            TeamMonthlySummaryModel.month.in_(HISTORICAL_MONTHS),
        )
    ).all()
    snapshot_rows = db.execute(
        select(HistoricalSnapshotModel.team, HistoricalSnapshotModel.month).where(
            HistoricalSnapshotModel.team.in_(TEAMS),
            HistoricalSnapshotModel.year == year,
            HistoricalSnapshotModel.month.in_(HISTORICAL_MONTHS),
        )
    ).all()
    expected = {(team, month) for team in TEAMS for month in HISTORICAL_MONTHS}
    report_set = {(team, int(month)) for team, month in report_rows if team in TEAMS and month}
    summary_set = {(team, int(month)) for team, month in summary_rows if team in TEAMS and month}
    snapshot_set = {(team, int(month)) for team, month in snapshot_rows if team in TEAMS and month}
    missing_reports = sorted(expected - report_set)
    missing_summaries = sorted(expected - summary_set)
    missing_snapshots = sorted(expected - snapshot_set)
    return {
        "team_reports_complete": not missing_reports,
        "team_monthly_summaries_complete": not missing_summaries,
        "dashboard_history_snapshots_complete": not missing_snapshots,
        "missing_team_reports": missing_reports,
        "missing_team_monthly_summaries": missing_summaries,
        "missing_dashboard_history_snapshots": missing_snapshots,
        "known_non_blocking_gaps": [
            "Unconfirmed data blocks are preserved as source references.",
            "Competency data is expected to be absent in T1/T2 and starts from T3.",
        ],
    }


def run_historical_import(
    source_directory: Path,
    imported_by: str = "historical_import",
    *,
    db: Session | None = None,
    workspace_dir: Path | None = None,
    commit: bool = True,
) -> ImportSessionReport:
    source_directory = Path(source_directory)
    root = workspace_dir or source_directory.parent
    own_session = db is None
    session = db or create_session()
    report = ImportSessionReport()
    try:
        discovered = discover_source_files(source_directory)
        by_month = {item.month: item for item in discovered}
        for required_month in HISTORICAL_MONTHS:
            if required_month not in by_month:
                report.warnings.append(
                    ImportIssue(
                        source_file=f"OKR tháng {required_month:02d}-{HISTORICAL_YEAR} - X.ĐK.xlsx",
                        severity="HIGH",
                        reason="Required historical source workbook is missing",
                        error_type="MISSING_SOURCE_WORKBOOK",
                    )
                )
        report.total_files_attempted = len(discovered)
        try:
            kr_mapping = resolve_kr_mapping(root)
        except Exception as exc:
            report.errors.append(
                ImportIssue(
                    source_file="OKR_Workshop.xlsx",
                    severity="HIGH",
                    reason=str(exc),
                    error_type="KR_MAPPING_MISSING",
                )
            )
            report.completeness = _verify_completeness(session)
            report.complete = False
            return report

        for item in discovered:
            file_result = FileImportResult(file_name=item.file_name, month=item.month, year=item.year)
            report.file_results.append(file_result)
            try:
                read_workbook_with_timeout(item.path)
            except Exception as exc:
                file_result.success = False
                file_result.errors.append(
                    ImportIssue(
                        source_file=item.file_name,
                        severity="HIGH",
                        reason=f"Cannot open workbook: {exc}",
                        error_type="WORKBOOK_OPEN_FAILED",
                    )
                )
                continue

            try:
                dashboard_history = extract_dashboard_history_lookup(item.path)
            except Exception as exc:
                dashboard_history = {}
                file_result.warnings.append(
                    ImportIssue(
                        source_file=item.file_name,
                        sheet="Dashboard",
                        severity="HIGH",
                        reason=f"Cannot parse dashboard history fallback: {exc}",
                        error_type="DASHBOARD_HISTORY_PARSE_FAILED",
                    )
                )

            try:
                snapshot_result = upsert_historical_snapshots(session, item.path, imported_by=imported_by)
                file_result.table_counts["historical_snapshots"]["inserted"] += int(
                    snapshot_result.get("imported_count") or 0
                )
                file_result.table_counts["historical_snapshots"]["updated"] += int(
                    snapshot_result.get("updated_count") or 0
                )
                file_result.table_counts["historical_snapshots"]["skipped"] += int(
                    snapshot_result.get("skipped_duplicates") or 0
                )
                for warning in snapshot_result.get("warnings", []):
                    file_result.warnings.append(_issue_from_warning(item.file_name, warning))
            except Exception as exc:
                file_result.success = False
                file_result.table_counts["historical_snapshots"]["failed"] += 1
                file_result.errors.append(
                    ImportIssue(
                        source_file=item.file_name,
                        severity="HIGH",
                        reason=f"Failed to import historical snapshots: {exc}",
                        error_type="SNAPSHOT_IMPORT_FAILED",
                    )
                )

            if item.month in {1, 2, 3}:
                for team in TEAMS:
                    try:
                        parsed = parse_team_report(
                            item.path,
                            team=team,
                            month=item.month,
                            year=item.year,
                            kr_mapping=kr_mapping,
                        )
                        parsed["source_path"] = item.path
                        _apply_dashboard_fallback(parsed, dashboard_history)
                        _record_parsed_report(
                            session,
                            file_result,
                            parsed,
                            source_path=item.path,
                            imported_by=imported_by,
                        )
                    except Exception as exc:
                        file_result.success = False
                        file_result.table_counts["team_reports"]["failed"] += 1
                        file_result.errors.append(
                            ImportIssue(
                                source_file=item.file_name,
                                sheet=team,
                                severity="HIGH",
                                reason=f"Failed to parse team {team}: {exc}",
                                error_type="TEAM_REPORT_PARSE_FAILED",
                            )
                        )
            else:
                template_dir = root / "template_xlsx"
                for team, template_name in TEMPLATE_FILES.items():
                    template_path = template_dir / template_name
                    if not template_path.exists():
                        file_result.warnings.append(
                            ImportIssue(
                                source_file=template_name,
                                severity="HIGH",
                                reason=f"Required Team_Template is missing; falling back to {item.file_name}",
                                error_type="MISSING_TEAM_TEMPLATE",
                            )
                        )
                    try:
                        parsed = parse_template_report(
                            template_path,
                            item.path,
                            team,
                            item.month,
                            item.year,
                            kr_mapping,
                        )
                        _apply_dashboard_fallback(parsed, dashboard_history)
                        _record_parsed_report(
                            session,
                            file_result,
                            parsed,
                            source_path=Path(parsed["source_path"]),
                            imported_by=imported_by,
                        )
                    except Exception as exc:
                        file_result.success = False
                        file_result.table_counts["team_reports"]["failed"] += 1
                        file_result.errors.append(
                            ImportIssue(
                                source_file=template_path.name,
                                sheet=team,
                                severity="HIGH",
                                reason=f"Failed to parse team template {team}: {exc}",
                                error_type="TEAM_TEMPLATE_PARSE_FAILED",
                            )
                        )

            if not file_result.errors:
                report.total_files_successful += 1
            report.warnings.extend(file_result.warnings)
            report.errors.extend(file_result.errors)

        if commit:
            session.commit()
        else:
            session.flush()
        report.total_team_reports = sum(
            result.table_counts["team_reports"]["inserted"]
            + result.table_counts["team_reports"]["updated"]
            for result in report.file_results
        )
        report.total_snapshots = sum(
            result.table_counts["historical_snapshots"]["inserted"]
            + result.table_counts["historical_snapshots"]["updated"]
            for result in report.file_results
        )
        report.total_summaries_upserted = sum(
            result.table_counts["team_monthly_summaries"]["inserted"]
            + result.table_counts["team_monthly_summaries"]["updated"]
            for result in report.file_results
        )
        report.completeness = _verify_completeness(session)
        report.complete = (
            not report.errors
            and report.completeness.get("team_reports_complete", False)
            and report.completeness.get("team_monthly_summaries_complete", False)
            and report.completeness.get("dashboard_history_snapshots_complete", False)
        )
        return report
    except Exception:
        if commit:
            session.rollback()
        raise
    finally:
        if own_session:
            session.close()


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Import historical KHMT/OKR data for T1-T4 2026")
    parser.add_argument(
        "source_directory",
        nargs="?",
        default=str(settings.workspace_dir / "KHMT_T1_T2_T3_T4"),
        help="Directory containing historical source workbooks",
    )
    parser.add_argument("--imported-by", default="historical_import")
    parser.add_argument("--no-commit", action="store_true", help="Run through parsing/storage without committing")
    args = parser.parse_args(argv)
    report = run_historical_import(
        Path(args.source_directory),
        imported_by=args.imported_by,
        commit=not args.no_commit,
    )
    print(report.render_text())
    return 0 if report.complete else 1


if __name__ == "__main__":
    raise SystemExit(main())
