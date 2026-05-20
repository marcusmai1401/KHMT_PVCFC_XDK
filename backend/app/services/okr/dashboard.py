from pathlib import Path
import logging
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import MergedCell

from app.core.config import settings
from app.services.okr.constants import DATA_SHEET_BLOCKS, DASHBOARD_COLUMNS, TEAM_DISPLAY_NAMES, TEAMS
from app.services.okr.chart_blocks import CHART_CONFIGS, build_chart_blocks, select_chart_snapshots_for_period
from app.services.okr.evaluation_rules import source_references as evaluation_rule_source_references
from app.services.okr.historical_snapshot import UNCONFIRMED_BLOCKS
from app.services.okr.kr_mapping import load_master_kr_mapping
from app.services.okr.kpi_rules import build_leader_kpi_allocations, summarize_leader_kpi_allocations
from app.services.okr.objective_sections import build_objective_sections
from app.services.okr.rules import calculate_skctkt
from app.services.okr.workbook_cleaner import strip_workbook_external_state
from app.services.repositories import now_utc


logger = logging.getLogger(__name__)


def build_empty_data_sheet() -> list[list[Any]]:
    rows = [[None for _ in range(16)] for _ in range(142)]
    rows[0][1] = "ĐK1.1"
    rows[1][1] = "ĐK1.1"
    for block in DATA_SHEET_BLOCKS:
        rows[block.start_row - 1][15] = block.name
    return rows


def _mapping_to_dict(mapping: Any) -> dict[str, Any]:
    if hasattr(mapping, "to_dict"):
        return mapping.to_dict()
    if isinstance(mapping, dict):
        return mapping
    return {
        "workshop_kr_code": mapping.workshop_kr_code,
        "kr_name": mapping.kr_name,
        "dashboard_column": mapping.dashboard_column,
        "measurement_type": mapping.measurement_type,
        "target_value": mapping.target_value,
        "source_row": getattr(mapping, "source_row", None),
    }


def _kr_sort_key(mapping: dict[str, Any]) -> tuple[int, int, str]:
    code = str(mapping.get("workshop_kr_code") or "")
    objective, _, kr = code.partition(".KR")
    if not objective.startswith("O") or not kr:
        return (10_000, 10_000, code)
    try:
        return (int(objective[1:]), int(kr), code)
    except ValueError:
        return (10_000, 10_000, code)


def _infer_report_period(team_reports: list[dict[str, Any]]) -> tuple[int, int] | None:
    for report in team_reports:
        month = report.get("report_month")
        year = report.get("report_year")
        if month and year:
            return int(month), int(year)
    return None


def build_dashboard_matrix(
    team_reports: list[dict[str, Any]],
    master: list[Any] | None = None,
    month: int | None = None,
    year: int | None = None,
    history_reports: list[dict[str, Any]] | None = None,
) -> dict[str, Any]:
    master_records = sorted(
        (_mapping_to_dict(record) for record in (master or load_master_kr_mapping())),
        key=_kr_sort_key,
    )
    period = (month, year) if month and year else _infer_report_period(team_reports)
    leader_allocations = build_leader_kpi_allocations(history_reports or team_reports, *period) if period else []
    leader_by_team = {row["team"]: row for row in leader_allocations}
    by_team = {team: {} for team in TEAMS}
    monthly = {team: "Hoàn thành" for team in TEAMS}
    discipline = {team: "OK" for team in TEAMS}

    for report in team_reports:
        team = report.get("team")
        if team not in by_team:
            continue
        for assessment in report.get("assessments", []):
            by_team[team][assessment["workshop_kr_code"]] = assessment["dashboard_status"]
        summary = report.get("team_level", {})
        monthly[team] = summary.get("monthly_assessment", monthly[team])
        discipline[team] = summary.get("discipline_status", discipline[team])

    rows = []
    for team in TEAMS:
        row = {
            "team": team,
            "team_name": TEAM_DISPLAY_NAMES[team],
            "discipline_status": discipline[team],
            "monthly_assessment": monthly[team],
            "leader_kpi_allocation": leader_by_team.get(
                team,
                {
                    "team": team,
                    "team_name": TEAM_DISPLAY_NAMES[team],
                    "current_assessment": monthly[team],
                    "good_or_better_streak_months": 0,
                    "a1": 0,
                    "a2": 0,
                    "triggered_rules": [],
                    "history": [],
                    "cap_note": "",
                },
            ),
            "kr_statuses": {},
        }
        for mapping in master_records:
            code = mapping["workshop_kr_code"]
            row["kr_statuses"][code] = by_team[team].get(code, "#N/A")
        rows.append(row)

    return {
        "columns": master_records,
        "dashboard_columns": DASHBOARD_COLUMNS,
        "teams": rows,
        "leader_kpi_allocations": leader_allocations,
        "kpi_allocation_summary": summarize_leader_kpi_allocations(leader_allocations),
        "workshop_staff_displayed": False,
    }


def _short_assessment(value: str | None) -> str | None:
    text = str(value or "").strip()
    if not text:
        return None
    lowered = text.lower()
    if "không" in lowered or "khong" in lowered:
        return "Không HT"
    if "tốt" in lowered or "tot" in lowered or "xuất sắc" in lowered or "xuat sac" in lowered:
        return "HT tốt"
    if text.upper() in {"HT", "HT TỐT", "KHÔNG HT"}:
        return text
    return "HT" if "hoàn thành" in lowered or "hoan thanh" in lowered else text


def _visible_teams(principal: dict[str, str] | None) -> list[str]:
    if principal and principal.get("role") == "Team_Account" and principal.get("user_id") in TEAMS:
        return [principal["user_id"]]
    return list(TEAMS)


def _filter_matrix(matrix: dict[str, Any], visible_teams: list[str]) -> dict[str, Any]:
    filtered = dict(matrix)
    filtered["teams"] = [row for row in matrix.get("teams", []) if row.get("team") in visible_teams]
    filtered["leader_kpi_allocations"] = [
        row for row in matrix.get("leader_kpi_allocations", []) if row.get("team") in visible_teams
    ]
    filtered["kpi_allocation_summary"] = summarize_leader_kpi_allocations(filtered["leader_kpi_allocations"])
    return filtered


def _apply_fi_ctkt_status(
    matrix: dict[str, Any],
    fi_counts_by_team: dict[str, int] | None,
    team_reports: list[dict[str, Any]],
) -> None:
    if fi_counts_by_team is None:
        return
    historical_teams = {
        report.get("team")
        for report in team_reports
        if report.get("source_type") == "historical_import"
    }
    for team_row in matrix.get("teams", []):
        team = team_row.get("team")
        if team in TEAMS and team not in historical_teams:
            team_row.setdefault("kr_statuses", {})["O5.KR13"] = calculate_skctkt(fi_counts_by_team.get(team, 0), 1)


def _metric_index(reports: list[dict[str, Any]]) -> dict[tuple[str, str], dict[str, Any]]:
    result: dict[tuple[str, str], dict[str, Any]] = {}
    for report in reports:
        team = report.get("team")
        if team not in TEAMS:
            continue
        for assessment in report.get("assessments", []):
            code = assessment.get("workshop_kr_code")
            metrics = assessment.get("metrics") or []
            if code and metrics:
                result[(team, code)] = metrics[0]
    return result


def _reports_through_period(
    reports: list[dict[str, Any]],
    *,
    month: int,
    year: int,
) -> list[dict[str, Any]]:
    return [
        report
        for report in reports
        if int(report.get("report_year") or 0) == year
        and 1 <= int(report.get("report_month") or 0) <= month
    ]


def _build_monthly_history(
    history_reports: list[dict[str, Any]],
    historical_snapshots: list[dict[str, Any]],
    *,
    year: int,
    visible_teams: list[str],
) -> list[dict[str, Any]]:
    db_values: dict[tuple[str, int], str | None] = {}
    for report in history_reports:
        team = report.get("team")
        month = report.get("report_month")
        if team in TEAMS and month and int(report.get("report_year") or 0) == year:
            db_values[(team, int(month))] = _short_assessment(report.get("team_level", {}).get("monthly_assessment"))

    snapshot_values: dict[tuple[str, int], str | None] = {}
    for snapshot in historical_snapshots:
        team = snapshot.get("team")
        month = snapshot.get("month")
        if team in TEAMS and month and int(snapshot.get("year") or 0) == year:
            snapshot_values[(team, int(month))] = _short_assessment(snapshot.get("monthly_assessment"))

    rows = []
    for team in visible_teams:
        months = []
        for month in range(1, 13):
            source = None
            assessment = None
            if (team, month) in db_values:
                assessment = db_values[(team, month)]
                source = "db"
            elif (team, month) in snapshot_values:
                assessment = snapshot_values[(team, month)]
                source = "snapshot"
            months.append({"month": month, "year": year, "assessment": assessment, "source": source})
        rows.append({"team": team, "team_name": TEAM_DISPLAY_NAMES[team], "months": months})
    return rows


def _build_minor_okr_summary(
    master_records: list[dict[str, Any]],
    matrix: dict[str, Any],
    team_reports: list[dict[str, Any]],
    *,
    visible_teams: list[str],
) -> list[dict[str, Any]]:
    metrics = _metric_index(team_reports)
    rows = []
    teams_by_code = {
        team_row["team"]: team_row.get("kr_statuses", {})
        for team_row in matrix.get("teams", [])
        if team_row.get("team") in visible_teams
    }
    for mapping in master_records:
        code = mapping["workshop_kr_code"]
        numeric_by_team = {
            team: metrics[(team, code)]
            for team in visible_teams
            if (team, code) in metrics
        }
        row = {
            "workshop_kr_code": code,
            "kr_name": mapping.get("kr_name") or code,
            "target_value": mapping.get("target_value"),
            "dashboard_column": mapping.get("dashboard_column"),
            "source_row": mapping.get("source_row"),
            "team_statuses": {
                team: teams_by_code.get(team, {}).get(code, "#N/A")
                for team in visible_teams
            },
            "numeric_metric": None,
        }
        if numeric_by_team:
            row["numeric_metric"] = {
                "teams": numeric_by_team,
                "actual": sum(float(metric.get("actual") or 0) for metric in numeric_by_team.values() if metric.get("actual") is not None),
                "target": next((metric.get("target") for metric in numeric_by_team.values() if metric.get("target") is not None), None),
            }
        rows.append(row)
    return rows


def _source_references() -> dict[str, Any]:
    return {
        "dashboard_history_range": "Dashboard!A20:AC25",
        "data_blocks": {key: config.source_reference for key, config in CHART_CONFIGS.items()},
        "evaluation_rules": evaluation_rule_source_references(),
        "unconfirmed_blocks": UNCONFIRMED_BLOCKS,
    }


def _collect_warnings(
    chart_blocks: dict[str, Any],
    historical_snapshots: list[dict[str, Any]],
    *,
    chart_reports: list[dict[str, Any]] | None = None,
) -> list[dict[str, Any]]:
    warnings: list[dict[str, Any]] = []
    for block in chart_blocks.values():
        warnings.extend(block.get("warnings") or [])
    for snapshot in historical_snapshots:
        warnings.extend(snapshot.get("warnings") or [])
    if not chart_reports:
        warnings.append(
            {
                "warning_type": "EMPTY_CHART_DATA",
                "severity": "LOW",
                "reason": "No submitted or locked reports are available for chart data in this period.",
            }
        )
    warnings.append(
        {
            "warning_type": "UNCONFIRMED_EXCEL_BLOCKS",
            "severity": "LOW",
            "reason": "Some Excel data blocks are exposed for review but are not counted into O2 KRs in the UI dashboard.",
            "items": UNCONFIRMED_BLOCKS,
        }
    )
    return warnings


def _period_data_state(
    month: int,
    team_reports: list[dict[str, Any]],
    historical_snapshots: list[dict[str, Any]],
    fi_counts_by_team: dict[str, int] | None = None,
) -> str:
    has_report_data = bool(team_reports)
    has_snapshot_data = any(int(snapshot.get("month") or 0) == month for snapshot in historical_snapshots)
    has_fi_data = any(int(value or 0) > 0 for value in (fi_counts_by_team or {}).values())
    if has_report_data and has_snapshot_data:
        return "partial"
    if has_report_data:
        return "ready"
    if has_snapshot_data or has_fi_data:
        return "partial"
    return "no_data"


def _latest_data_period(
    reports: list[dict[str, Any]],
    historical_snapshots: list[dict[str, Any]],
) -> dict[str, int] | None:
    periods: list[tuple[int, int]] = []
    for report in reports:
        month = report.get("report_month")
        year = report.get("report_year")
        if month and year:
            periods.append((int(month), int(year)))
    for snapshot in historical_snapshots:
        month = int(snapshot.get("month") or 0)
        year = snapshot.get("year")
        if 1 <= month <= 12 and year:
            periods.append((month, int(year)))
    if not periods:
        return None
    year, month = max((year, month) for month, year in periods)
    return {"month": month, "year": year}


def build_dashboard_view(
    month: int,
    year: int,
    team_reports: list[dict[str, Any]],
    master: list[Any] | None = None,
    *,
    history_reports: list[dict[str, Any]] | None = None,
    matrix_history_reports: list[dict[str, Any]] | None = None,
    historical_snapshots: list[dict[str, Any]] | None = None,
    headcounts: dict[str, dict[str, Any]] | None = None,
    fi_counts_by_team: dict[str, int] | None = None,
    principal: dict[str, str] | None = None,
) -> dict[str, Any]:
    master_records = sorted(
        (_mapping_to_dict(record) for record in (master or load_master_kr_mapping())),
        key=_kr_sort_key,
    )
    visible_teams = _visible_teams(principal)
    matrix = build_dashboard_matrix(
        team_reports,
        master_records,
        month=month,
        year=year,
        history_reports=matrix_history_reports or history_reports or team_reports,
    )
    _apply_fi_ctkt_status(matrix, fi_counts_by_team, team_reports)
    filtered_matrix = _filter_matrix(matrix, visible_teams)
    chart_reports = _reports_through_period(history_reports or team_reports, month=month, year=year)
    snapshots = historical_snapshots or []
    chart_snapshots = select_chart_snapshots_for_period(snapshots, month=month, year=year)
    chart_blocks = build_chart_blocks(
        chart_reports,
        month=month,
        year=year,
        visible_teams=visible_teams,
        historical_snapshots=chart_snapshots,
        headcounts=headcounts,
        fi_counts_by_team=fi_counts_by_team,
    )
    monthly_history = _build_monthly_history(
        history_reports or team_reports,
        snapshots,
        year=year,
        visible_teams=visible_teams,
    )
    minor_okr_summary = _build_minor_okr_summary(
        master_records,
        filtered_matrix,
        team_reports,
        visible_teams=visible_teams,
    )
    source_references = _source_references()
    warnings = _collect_warnings(chart_blocks, snapshots, chart_reports=chart_reports)
    technical_warnings = list(warnings)
    try:
        objective_sections = build_objective_sections(
            month=month,
            year=year,
            team_reports=team_reports,
            historical_snapshots=snapshots,
            chart_snapshots=chart_snapshots,
            headcounts=headcounts,
            fi_counts_by_team=fi_counts_by_team,
            chart_blocks=chart_blocks,
            matrix=filtered_matrix,
            minor_okr_summary=minor_okr_summary,
            history_reports=chart_reports,
        )
    except Exception as exc:
        logger.exception("Failed to build objective-first dashboard sections")
        objective_sections = []
        warning = {
            "warning_type": "OBJECTIVE_SECTIONS_BUILD_FAILED",
            "severity": "MEDIUM",
            "reason": str(exc),
        }
        technical_warnings.append(warning)
        warnings.append(warning)
    latest_data_period = _latest_data_period(history_reports or team_reports, snapshots)
    period = {
        "month": month,
        "year": year,
        "label": f"T{month}/{year}",
        "data_state": _period_data_state(month, team_reports, snapshots, fi_counts_by_team),
    }
    return {
        **filtered_matrix,
        "period": period,
        "matrix": filtered_matrix,
        "monthly_history": monthly_history,
        "chart_blocks": chart_blocks,
        "minor_okr_summary": minor_okr_summary,
        "source_references": source_references,
        "warnings": warnings,
        "objective_sections": objective_sections,
        "technical_metadata": {
            "warnings": technical_warnings,
            "source_references": source_references,
            "latest_data_period": latest_data_period,
        },
    }


def populate_data_sheet_from_reports(
    team_reports: list[dict[str, Any]],
    fi_counts_by_team: dict[str, int] | None = None,
) -> list[list[Any]]:
    rows = build_empty_data_sheet()
    month = next((r.get("report_month") for r in team_reports if r.get("report_month")), None)
    metric_by_month_team_code = {}
    for report in team_reports:
        team = report.get("team")
        if team not in TEAMS:
            continue
        report_month = report.get("report_month")
        for assessment in report.get("assessments", []):
            metrics = assessment.get("metrics", [])
            if metrics:
                metric_by_month_team_code[(report_month, team, assessment["workshop_kr_code"])] = metrics[0]

    def aggregate(months: list[int | None], teams: list[str], code: str) -> dict[str, Any]:
        actual = 0.0
        total = 0.0
        target = None
        has_actual = False
        has_total = False
        for report_month in months:
            for team in teams:
                metric = metric_by_month_team_code.get((report_month, team, code), {})
                if metric.get("actual") is not None:
                    actual += float(metric["actual"])
                    has_actual = True
                if metric.get("total") is not None:
                    total += float(metric["total"])
                    has_total = True
                if metric.get("target") is not None:
                    target = metric["target"]
        percentage = round(actual / total * 100, 2) if has_actual and has_total and total else None
        return {
            "actual": actual if has_actual else None,
            "total": total if has_total else None,
            "percentage": percentage,
            "target": target,
        }

    def write_metric(row_index: int, label: str, team: str, code: str, values: dict[str, Any], block: str) -> None:
        rows[row_index][0] = label
        rows[row_index][1] = team
        rows[row_index][2] = code
        rows[row_index][3] = values.get("actual")
        rows[row_index][4] = values.get("total")
        rows[row_index][5] = values.get("percentage")
        rows[row_index][6] = values.get("target")
        rows[row_index][15] = block

    selected_months = [month] if month else list(range(1, 13))

    # Rows 3-15: T1-T12 plus year-to-date SCĐX total across teams.
    for month_number in range(1, 13):
        values = aggregate([month_number], list(TEAMS), "O2.KR1")
        write_metric(month_number + 1, f"T{month_number}", "All teams", "O2.KR1", values, "scdx_monthly")
    cumulative_months = list(range(1, (month or 12) + 1))
    write_metric(14, "YTD", "All teams", "O2.KR1", aggregate(cumulative_months, list(TEAMS), "O2.KR1"), "scdx_monthly")

    # Rows 16-18: fixed SCĐX direct extraction for the three equipment teams.
    for offset, team in enumerate(["TBHTĐK", "TBCH", "TBĐL"]):
        write_metric(15 + offset, team, team, "O2.KR1", aggregate(selected_months, [team], "O2.KR1"), "scdx_by_team")

    # Rows 21-35: TCĐK-only monthly/shift structure, never rotating non-TCĐK teams into the block.
    for month_number in range(1, 13):
        write_metric(19 + month_number, f"T{month_number}", "TCĐK", "O2.KR1", aggregate([month_number], ["TCĐK"], "O2.KR1"), "scdx_tcdk_shift")
    for offset, label in enumerate(["YTD", "Shift summary 1", "Shift summary 2"], start=32):
        write_metric(offset, label, "TCĐK", "O2.KR1", aggregate(cumulative_months, ["TCĐK"], "O2.KR1"), "scdx_tcdk_shift")

    def write_team_month_matrix(start_index: int, row_count: int, code: str, block: str) -> None:
        for offset in range(row_count):
            team = TEAMS[offset % len(TEAMS)]
            matrix_month = offset // len(TEAMS) + 1
            write_metric(
                start_index + offset,
                f"T{matrix_month}",
                team,
                code,
                aggregate([matrix_month], [team], code),
                block,
            )

    write_team_month_matrix(42, 20, "O2.KR2", "bddk_npk")
    write_team_month_matrix(64, 20, "O3.KR2", "stop_cards")

    # Rows 86-89: VHDN/running participation, split from sports.
    for offset, team in enumerate(TEAMS):
        write_metric(85 + offset, "running", team, "O6.KR1", aggregate(selected_months, [team], "O6.KR1"), "vhdn_running")

    # Rows 91-94: Hội thao/chương trình chung participation.
    for offset, team in enumerate(TEAMS):
        write_metric(90 + offset, "sports", team, "O6.KR2", aggregate(selected_months, [team], "O6.KR2"), "vhdn_sports")

    # Rows 98-107: training actual/planned rows and summary slots, T1-T11 is handled in charts.
    row_index = 97
    for team in TEAMS:
        for label in ["actual", "planned"]:
            if row_index >= 107:
                break
            write_metric(row_index, label, team, "O5.KR3", aggregate(selected_months, [team], "O5.KR3"), "training_hours")
            row_index += 1
    while row_index < 107:
        write_metric(row_index, "total", "All teams", "O5.KR3", aggregate(selected_months, list(TEAMS), "O5.KR3"), "training_hours")
        row_index += 1

    # Rows 110-114: Sáng kiến only. CTKT is sourced separately from the FI module.
    for offset, team in enumerate(TEAMS):
        write_metric(109 + offset, team, team, "O5.KR12", aggregate(selected_months, [team], "O5.KR12"), "sk_initiatives")
    write_metric(113, "Total", "All teams", "O5.KR12", aggregate(selected_months, list(TEAMS), "O5.KR12"), "sk_initiatives")

    # Rows 117-127: TCĐK weekly SCĐX, fixed to W14-W22 plus summary rows.
    for offset, week in enumerate(range(14, 23)):
        write_metric(116 + offset, f"W{week}", "TCĐK", "O2.KR1", aggregate(selected_months, ["TCĐK"], "O2.KR1"), "weekly_scdx")
    for offset, label in enumerate(["YTD", "Open"], start=125):
        write_metric(offset, label, "TCĐK", "O2.KR1", aggregate(cumulative_months, ["TCĐK"], "O2.KR1"), "weekly_scdx")

    for offset in range(13):
        team = TEAMS[offset % len(TEAMS)] if offset < 12 else "All teams"
        write_metric(129 + offset, "Competency", team, "O5.KR1", aggregate(selected_months, [team] if team in TEAMS else list(TEAMS), "O5.KR1"), "competency")
    return rows


def _export_mapping_warnings() -> list[dict[str, Any]]:
    return [
        {
            "warning_type": "UNCONFIRMED_EXPORT_MAPPING",
            "severity": "LOW",
            "source_range": item["source_range"],
            "observed_label": item["observed_label"],
            "candidate_kr_codes": ", ".join(item["candidate_kr_codes"]),
            "mapping_status": item["mapping_status"],
            "reason": f"Legacy export preserves this block; UI dashboard treats it as unconfirmed. {item['reason']}",
        }
        for item in UNCONFIRMED_BLOCKS
    ]


def _write_export_warning_sheet(workbook: Workbook, warnings: list[dict[str, Any]]) -> None:
    if not warnings:
        return
    sheet_name = "OKR_Warnings"
    if sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        sheet.delete_rows(1, sheet.max_row)
    else:
        sheet = workbook.create_sheet(sheet_name)
    sheet.sheet_state = "hidden"
    headers = [
        "warning_type",
        "severity",
        "source_range",
        "observed_label",
        "candidate_kr_codes",
        "mapping_status",
        "reason",
    ]
    for col_index, header in enumerate(headers, start=1):
        sheet.cell(1, col_index).value = header
    for row_index, warning in enumerate(warnings, start=2):
        for col_index, header in enumerate(headers, start=1):
            sheet.cell(row_index, col_index).value = warning.get(header)


def _safe_write_cell(cell: Any, value: Any, warnings: list[dict[str, Any]], context: str) -> None:
    try:
        cell.value = value
    except Exception as exc:
        warning = {
            "warning_type": "EXPORT_CELL_WRITE_FAILED",
            "severity": "MEDIUM",
            "source_range": f"{cell.parent.title}!{cell.coordinate}",
            "observed_label": context,
            "candidate_kr_codes": "",
            "mapping_status": "write_failed",
            "reason": str(exc),
        }
        logger.warning(
            "Failed to write OKR export cell %s!%s: %s",
            cell.parent.title,
            cell.coordinate,
            exc,
        )
        warnings.append(warning)


def export_dashboard_workbook(
    team_reports: list[dict[str, Any]],
    output_path: Path | None = None,
    fi_counts_by_team: dict[str, int] | None = None,
) -> Path:
    settings.storage_dir.joinpath("exports").mkdir(parents=True, exist_ok=True)
    target = output_path or settings.storage_dir / "exports" / "okr-dashboard-export.xlsx"
    if settings.source_okr_workbook.exists():
        workbook = load_workbook(settings.source_okr_workbook, keep_links=False)
        strip_workbook_external_state(workbook)
        if "data" not in workbook.sheetnames:
            workbook.create_sheet("data")
        if "Dashboard" not in workbook.sheetnames:
            workbook.create_sheet("Dashboard")
    else:
        workbook = Workbook()
        workbook.active.title = "data"
        workbook.create_sheet("Dashboard")

    export_warnings = _export_mapping_warnings()
    for warning in export_warnings:
        logger.warning(
            "Unconfirmed OKR export mapping preserved: %s %s",
            warning["source_range"],
            warning["candidate_kr_codes"],
        )

    data_rows = populate_data_sheet_from_reports(team_reports, fi_counts_by_team=fi_counts_by_team)
    data_sheet = workbook["data"]
    for row_index, row in enumerate(data_rows, start=1):
        for col_index, value in enumerate(row, start=1):
            cell = data_sheet.cell(row_index, col_index)
            if isinstance(cell, MergedCell):
                continue
            if isinstance(cell.value, str) and cell.value.startswith("="):
                continue
            _safe_write_cell(cell, value, export_warnings, "data_sheet")

    period = _infer_report_period(team_reports)
    dashboard = build_dashboard_matrix(
        team_reports,
        month=period[0] if period else None,
        year=period[1] if period else None,
        history_reports=team_reports,
    )
    _apply_fi_ctkt_status(dashboard, fi_counts_by_team, team_reports)
    dash_sheet = workbook["Dashboard"]
    if not isinstance(dash_sheet["A1"], MergedCell):
        _safe_write_cell(dash_sheet["A1"], "BÁO CÁO KẾ HOẠCH MỤC TIÊU XƯỞNG ĐIỀU KHIỂN", export_warnings, "dashboard_title")
    if not isinstance(dash_sheet["I7"], MergedCell):
        _safe_write_cell(dash_sheet["I7"], "Kỷ luật", export_warnings, "dashboard_header")
    for coordinate, value in [("J7", "A2"), ("K7", "A1")]:
        if not isinstance(dash_sheet[coordinate], MergedCell):
            _safe_write_cell(dash_sheet[coordinate], value, export_warnings, "dashboard_header")
    for index, mapping in enumerate(dashboard["columns"], start=12):
        cell = dash_sheet.cell(7, index)
        if not isinstance(cell, MergedCell):
            _safe_write_cell(cell, mapping["workshop_kr_code"], export_warnings, "dashboard_kr_code")
    for row_offset, team in enumerate(dashboard["teams"], start=8):
        leader_allocation = team.get("leader_kpi_allocation") or {}
        for coordinate, value in [
            (1, team["team_name"]),
            (8, team["monthly_assessment"]),
            (9, team["discipline_status"]),
            (10, leader_allocation.get("a2") or None),
            (11, leader_allocation.get("a1") or None),
        ]:
            cell = dash_sheet.cell(row_offset, coordinate)
            if not isinstance(cell, MergedCell):
                _safe_write_cell(cell, value, export_warnings, "dashboard_team_row")
        for col_offset, mapping in enumerate(dashboard["columns"], start=12):
            cell = dash_sheet.cell(row_offset, col_offset)
            if not isinstance(cell, MergedCell):
                _safe_write_cell(
                    cell,
                    team["kr_statuses"][mapping["workshop_kr_code"]],
                    export_warnings,
                    f"dashboard_status_{mapping['workshop_kr_code']}",
                )

    workbook.properties.creator = "OKR Automation System"
    workbook.properties.lastModifiedBy = "OKR Automation System"
    workbook.properties.modified = now_utc().replace(tzinfo=None)
    workbook.properties.subject = f"Generated at {now_utc().isoformat()}"
    workbook.properties.keywords = "OKR export includes hidden OKR_Warnings sheet for mapping confidence"
    _write_export_warning_sheet(workbook, export_warnings)
    strip_workbook_external_state(workbook)
    workbook.save(target)
    return target
