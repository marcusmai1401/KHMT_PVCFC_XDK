from openpyxl import Workbook
from openpyxl import load_workbook
from io import BytesIO
from zipfile import ZipFile

from app.models.domain import HistoricalSnapshotModel
from app.services.okr.chart_blocks import build_chart_blocks
from app.services.okr.dashboard import build_dashboard_view, export_dashboard_workbook
from app.services.okr.evaluation_rules import classify_dashboard_assessment, source_references
from app.services.okr.historical_snapshot import _extract_dashboard_note_blocks, import_historical_snapshot
from app.services.okr.objective_sections import _apply_dashboard_narratives, resolve_indicator_value
from app.services.okr.period_resolver import resolve_default_period
from app.services.okr.team_normalizer import normalize_team_label


def _report(team: str, month: int, code: str, actual: int | None, total: int | None = None):
    return {
        "team": team,
        "report_month": month,
        "report_year": 2026,
        "team_level": {"monthly_assessment": "Hoàn thành tốt", "discipline_status": "OK"},
        "assessments": [
            {
                "workshop_kr_code": code,
                "dashboard_status": "OK",
                "metrics": [{"actual": actual, "total": total, "target": total}],
            }
        ],
    }


def _chart_snapshot(block_type: str, rows: list[dict], source_file_name: str, notes: list[str] | None = None):
    return {
        "source_file_name": source_file_name,
        "team": "__CHARTS__",
        "month": 0,
        "year": 2026,
        "chart_payload": {"block_type": block_type, "rows": rows, "notes": notes or []},
        "warnings": [],
    }


def _row(label, *values):
    return {"label": label, "values": [label, *values]}


def test_team_normalizer_covers_excel_aliases():
    assert normalize_team_label("Đội thiết bị Hệ thống điều khiển")[0] == "TBHTĐK"
    assert normalize_team_label("Đội thiết bị đo")[0] == "TBĐL"
    assert normalize_team_label("Đội thiết bị cơ cấu chấp hành")[0] == "TBCH"
    assert normalize_team_label("Tổ trực ca điều khiển")[0] == "TCĐK"


def test_evaluation_rules_match_excel_text_blocks():
    statuses = {"O1.KR1": "OK", "O5.KR13": "GOOD", "O6.KR1": "OK"}
    assert classify_dashboard_assessment(statuses, "OK") == "HT tốt"
    assert classify_dashboard_assessment({"O1.KR1": "OK", "O6.KR1": "GOOD"}, "OK") == "HT tốt"
    assert classify_dashboard_assessment({"O1.KR1": "OK", "O6.KR2": "GOOD"}, "OK") == "HT tốt"
    assert classify_dashboard_assessment({"O1.KR1": "OK", "O6.KR1": "GOOD"}, "NOK") == "Không HT"
    assert classify_dashboard_assessment({"O1.KR1": "NG", "O6.KR1": "GOOD"}, "OK") == "Không HT"
    assert classify_dashboard_assessment({"O1.KR1": "OK"}, "OK") == "HT"
    assert classify_dashboard_assessment({"O1.KR1": "NG"}, "OK") == "Không HT"
    assert "Dashboard!M15:P15" in source_references()["good"]
    assert "Dashboard!M16:P16" in source_references()["good"]


def test_chart_blocks_keep_null_distinct_from_zero():
    reports = [_report("TBCH", 4, "O6.KR1", 0, 10)]
    blocks = build_chart_blocks(reports, month=4, year=2026, visible_teams=["TBCH", "TBĐL"])
    items = blocks["vhdn_running"]["items"]
    assert items[0]["actual"] == 0
    assert items[0]["participation_rate"] == 0
    assert items[1]["actual"] is None
    assert blocks["training"]["labels"][-1] == "T11"


def test_o6_report_backfills_chart_only_kr_details():
    report = {
        "krs": [
            {"code": "O6.KR1", "title": "KR1 RÈN LUYỆN CHẠY BỘ", "lines": []},
            {"code": "O6.KR2", "title": "KR2 HỘI THAO", "lines": ["Giữ nguyên diễn giải"]},
        ],
        "notes": [],
    }
    visuals = [
        {
            "id": "o6_running",
            "payload": {
                "items": [
                    {
                        "team": "TBHTĐK",
                        "team_name": "Đội thiết bị hệ thống điều khiển",
                        "actual": 5,
                        "total": 9,
                        "participation_rate": 5 / 9,
                    }
                ],
            },
        },
        {"id": "o6_sports", "payload": {"items": []}},
    ]

    sections = [{"objective_code": "O6", "visuals": visuals}]
    _apply_dashboard_narratives(
        sections,
        {
            "report": {"O6": report},
            "o6_counts": {"running": {"actual": "6", "target": "20"}},
        },
    )
    enriched = sections[0]["report"]

    assert enriched["krs"][0]["lines"] == [
        "Đội thiết bị hệ thống điều khiển: 5/9 người tham gia, đạt 56%",
        "Số lần tổ chức (lũy kế): 6/20 (mục tiêu năm)",
    ]
    assert enriched["krs"][1]["lines"] == ["Giữ nguyên diễn giải"]
    assert report["krs"][0]["lines"] == []


def test_export_writes_hidden_mapping_warning_sheet(tmp_path):
    path = export_dashboard_workbook([], output_path=tmp_path / "export.xlsx")
    workbook = load_workbook(path, read_only=False, data_only=False, keep_links=False)

    assert "OKR_Warnings" in workbook.sheetnames
    sheet = workbook["OKR_Warnings"]
    assert sheet.sheet_state == "hidden"
    assert sheet["A2"].value == "UNCONFIRMED_EXPORT_MAPPING"
    assert sheet["C2"].value == "data!A117:D127"


def test_dashboard_view_keeps_backward_keys_and_exposes_full_workshop_to_team_account():
    reports = [_report("TBCH", 4, "O3.KR2", 12)]
    data = build_dashboard_view(
        4,
        2026,
        reports,
        history_reports=reports,
        fi_counts_by_team={"TBCH": 1},
        principal={"role": "Team_Account", "user_id": "TBCH"},
    )
    assert "columns" in data and "teams" in data
    assert [team["team"] for team in data["teams"]] == ["TBHTĐK", "TBCH", "TBĐL", "TCĐK"]
    assert data["teams"][0]["monthly_assessment"] == "N/A"
    assert data["teams"][0]["has_report"] is False
    tbch_history = next(row for row in data["monthly_history"] if row["team"] == "TBCH")
    assert len(tbch_history["months"]) == 12
    assert tbch_history["months"][3]["source"] == "db"
    assert data["source_references"]["unconfirmed_blocks"][0]["mapping_status"] == "needs_confirmation"
    assert data["period"] == {"month": 4, "year": 2026, "label": "T4/2026", "data_state": "ready"}
    assert [section["objective_code"] for section in data["objective_sections"]] == ["O1", "O2", "O3", "O4", "O5", "O6"]
    assert data["technical_metadata"]["source_references"]["unconfirmed_blocks"]


def test_dashboard_matrix_exposes_discipline_description():
    reports = [_report("TBCH", 4, "O3.KR2", 12)]
    reports[0]["team_level"] = {
        "monthly_assessment": "Không hoàn thành",
        "discipline_status": "NOK",
        "discipline_description": "Một nhân sự Đội TBCH không tuân thủ đúng HDBD",
    }

    data = build_dashboard_view(4, 2026, reports, history_reports=reports)
    team = next(row for row in data["teams"] if row["team"] == "TBCH")

    assert team["discipline_status"] == "NOK"
    assert team["discipline_description"] == "Một nhân sự Đội TBCH không tuân thủ đúng HDBD"
    assert team["kr_statuses"]["O1.KR1"] == "NG"


def test_dashboard_view_marks_empty_period_and_exposes_latest_data_period():
    snapshots = [
        {
            "team": "TCĐK",
            "month": 4,
            "year": 2026,
            "monthly_assessment": "HT tốt",
            "chart_payload": {},
            "warnings": [{"warning_type": "needs_confirmation", "severity": "LOW", "reason": "review"}],
        }
    ]
    data = build_dashboard_view(5, 2026, [], history_reports=[], historical_snapshots=snapshots)

    assert data["period"]["data_state"] == "no_data"
    assert data["technical_metadata"]["latest_data_period"] == {"month": 4, "year": 2026}
    assert len(data["objective_sections"]) == 6
    assert "needs_confirmation" not in str(data["objective_sections"])

    snapshot_period = build_dashboard_view(4, 2026, [], history_reports=[], historical_snapshots=snapshots)
    assert snapshot_period["period"]["data_state"] == "partial"
    assert snapshot_period["objective_sections"][3]["conclusion"]
    assert snapshot_period["objective_sections"][5]["conclusion"]


def test_period_resolver_priority_and_indicator_priority():
    resolved = resolve_default_period(
        last_selected=(5, 2026),
        latest_data=(4, 2026),
        workbook=(3, 2026),
        today=(6, 2026),
    )
    assert (resolved.month, resolved.year, resolved.source, resolved.label) == (5, 2026, "last_selected", "T5/2026")
    fallback = resolve_default_period(last_selected=(13, 2026), latest_data=None, workbook=None, today=(6, 2026))
    assert (fallback.month, fallback.year, fallback.source) == (6, 2026, "current")

    assert resolve_indicator_value(1, 2, 3, True) == (1, "db_locked", "ready")
    assert resolve_indicator_value(None, 2, 3, True) == (2, "normalized", "ready")
    assert resolve_indicator_value(None, None, 3, True) == (3, "dashboard_snapshot", "ready")
    assert resolve_indicator_value(None, None, None, False) == (None, None, "no_plan")


def test_dashboard_view_uses_fi_counts_as_normalized_o5_data():
    data = build_dashboard_view(4, 2026, [], history_reports=[], fi_counts_by_team={"TBCH": 2})
    o5 = next(section for section in data["objective_sections"] if section["objective_code"] == "O5")
    fi_visual = next(visual for visual in o5["visuals"] if visual["id"] == "o5_fi")

    assert data["period"]["data_state"] == "partial"
    assert fi_visual["data_state"] == "ready"
    assert fi_visual["source"] == "fi_module"
    assert fi_visual["payload"]["fi_counts_by_team"] == {"TBCH": 2}
    assert fi_visual["payload"]["target_basis"] == "monthly_per_team"
    assert fi_visual["payload"]["target_per_team"] == 1
    assert fi_visual["payload"]["target_team_count"] == 4
    assert fi_visual["payload"]["master_target"] == 4.0


def test_dashboard_view_prefers_selected_historical_snapshot_blocks_for_t1():
    history_reports = [
        _report("TBCH", 1, "O3.KR2", 999),
        _report("TBCH", 2, "O3.KR2", 59),
    ]
    t1_snapshots = [
        _chart_snapshot(
            "stop_by_month",
            [_row("T1", 51, None, None), _row("T2", None, None, None)],
            "OKR tháng 01-2026 - X.ĐK.xlsx",
        ),
        _chart_snapshot(
            "training",
            [
                _row("Kế hoạch ", 24, 47, 106, 166, 134, 51, 206, 0, 42, 257, 93),
                _row("thực hiện", 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0),
            ],
            "OKR tháng 01-2026 - X.ĐK.xlsx",
        ),
        _chart_snapshot(
            "vhdn_running",
            [
                _row("Đội thiết bị Hệ thống điều khiển", 4, 8, 0.5, 0.5),
                _row("Đội thiết bị Chấp hành", 5, 8, 0.625, 0.5),
                _row("Đội thiết bị Đo lường", 6, 8, 0.75, 0.5),
                _row("Tổ trực ca điều khiển", 5, 7, 0.7142857143, 0.5),
            ],
            "OKR tháng 01-2026 - X.ĐK.xlsx",
        ),
        _chart_snapshot(
            "sk_initiatives",
            [
                _row("Sáng kiến", None),
                _row("Đội thiết bị Hệ thống điều khiển", 1),
                _row("Đội thiết bị Chấp hành", 1),
                _row("Đội thiết bị Đo lường", 3),
                _row("Tổ trực ca điều khiển", 3),
            ],
            "OKR tháng 01-2026 - X.ĐK.xlsx",
        ),
        _chart_snapshot(
            "stop_by_month",
            [_row("T1", 777, None, None), _row("T2", 888, None, None)],
            "OKR tháng 02-2026 - X.ĐK.xlsx",
        ),
    ]

    data = build_dashboard_view(
        1,
        2026,
        history_reports[:1],
        history_reports=history_reports,
        historical_snapshots=t1_snapshots,
    )

    assert data["chart_blocks"]["stop_by_month"]["datasets"][0]["data"][:2] == [51, None]
    assert data["chart_blocks"]["training"]["datasets"][0]["data"][:4] == [24, 47, 106, 166]
    assert data["chart_blocks"]["training"]["datasets"][1]["data"][:4] == [0, 0, 0, 0]
    assert data["chart_blocks"]["vhdn_running"]["items"][0]["actual"] == 4
    assert data["chart_blocks"]["vhdn_running"]["items"][0]["total"] == 8
    assert data["chart_blocks"]["sk_initiatives"]["datasets"][0]["data"] == [1, 1, 3, 3]
    assert data["chart_blocks"]["sk_initiatives"]["total"] == 8.0


def test_dashboard_view_builds_o2_snapshot_visuals_and_o3_excel_style_visuals():
    snapshots = [
        _chart_snapshot(
            "o2_bddk",
            [
                _row("T1", 1302, 1495, 0.8709030100, 0.98),
                _row("Lũy kế ", 1302, 1495, 0.8709030100, 0.98),
                _row("Đội thiết bị Hệ thống điều khiển", 254, 254, 1, 0.98),
                _row("Đội thiết bị Chấp hành", 682, 801, 0.8514357054, 0.98),
                _row("Đội thiết bị Đo lường", 366, 440, 0.8318181818, 0.98),
            ],
            "OKR tháng 01-2026 - X.ĐK.xlsx",
        ),
        _chart_snapshot(
            "o2_scdx",
            [
                _row("Tổ trực ca điều khiển", 224, 230, 0.9739130435, 0.99),
                _row("T1", 224, 230, 0.9739130435, 0.99),
                _row("Lũy kế ", 224, 230, 0.9739130435, None),
            ],
            "OKR tháng 01-2026 - X.ĐK.xlsx",
        ),
        _chart_snapshot(
            "stop_by_team",
            [
                _row("Đội thiết bị Hệ thống điều khiển", 17, 10, 1.7, 0.5),
                _row("Đội thiết bị Chấp hành", 14, 14, 1, 0.5),
            ],
            "OKR tháng 01-2026 - X.ĐK.xlsx",
        ),
        _chart_snapshot(
            "stop_by_month",
            [_row("T1", 51, None, None), _row(None, 51, 200, 0.255)],
            "OKR tháng 01-2026 - X.ĐK.xlsx",
        ),
    ]

    data = build_dashboard_view(1, 2026, [], history_reports=[], historical_snapshots=snapshots)
    o2 = next(section for section in data["objective_sections"] if section["objective_code"] == "O2")
    o3 = next(section for section in data["objective_sections"] if section["objective_code"] == "O3")

    assert [visual["id"] for visual in o2["visuals"]] == [
        "o2_bddk_by_team",
        "o2_bddk_by_month",
        "o2_scdx_by_team",
        "o2_scdx_by_month",
    ]
    assert o2["visuals"][0]["kind"] == "metric_table"
    assert [row["actual"] for row in o2["visuals"][0]["payload"]["rows"]] == [254, 682, 366]
    assert o2["visuals"][0]["payload"]["summary_items"] == [
        {"label": "T1 thực hiện", "value": 1302},
        {"label": "T1 kế hoạch", "value": 1495},
        {"label": "T1 tỷ lệ", "value": 0.87090301, "format": "percent"},
    ]
    assert o2["visuals"][0]["payload"]["columns"][0]["label"] == "Đội/Tổ"
    assert o2["visuals"][1]["payload"]["summary_items"][2]["value"] == 0.87090301
    assert o2["visuals"][1]["payload"]["summary_items"][2]["label"] == "Lũy kế đến T1 tỷ lệ"
    assert o3["visuals"][0]["kind"] == "bar_line_chart"
    assert o3["visuals"][0]["payload"]["datasets"][2]["data"] == [1.7, 1]
    assert o3["visuals"][1]["payload"]["summary_items"] == [
        {"label": "Thực hiện", "value": 51},
        {"label": "Kế hoạch", "value": 200},
        {"label": "Tỷ lệ", "value": 0.255, "format": "percent"},
    ]


def test_dashboard_view_preserves_historical_o5_status_when_fi_counts_are_empty():
    historical_report = {
        **_report("TBCH", 1, "O5.KR13", 1),
        "source_type": "historical_import",
    }

    data = build_dashboard_view(
        1,
        2026,
        [historical_report],
        history_reports=[historical_report],
        fi_counts_by_team={"TBCH": 0},
    )

    tbch = next(team for team in data["teams"] if team["team"] == "TBCH")
    assert tbch["kr_statuses"]["O5.KR13"] == "OK"


def test_dashboard_view_adds_month_six_sap_compliance_visual_only_for_matching_period():
    sap_snapshot = {
        "source_file_name": "OKR tháng 06-2026 - X.ĐK.xlsx",
        "team": "__CHARTS__",
        "month": 0,
        "year": 2026,
        "chart_payload": {
            "block_type": "sap_compliance",
            "title": "Báo cáo tình hình thực hiện tuân thủ nghiệp vụ SAP",
            "period": {"month": 6, "year": 2026, "label": "T6/2026"},
            "backlog_total": 169,
            "totals": {"overdue_wo": 50, "unconfirmed_wo": 70, "violating_wo": 120},
            "rates": {"overdue_share": 0.296, "unconfirmed_share": 0.414, "violation_share": 0.71},
            "supervisors": [
                {
                    "name": "Trực ca điều khiển",
                    "overdue_wo": 27,
                    "unconfirmed_wo": 60,
                    "violating_wo": 87,
                }
            ],
        },
        "warnings": [],
    }
    period_snapshot = {
        "source_file_name": "OKR tháng 06-2026 - X.ĐK.xlsx",
        "team": "TCĐK",
        "month": 6,
        "year": 2026,
        "monthly_assessment": "HT tốt",
        "chart_payload": {},
        "warnings": [],
    }

    june = build_dashboard_view(6, 2026, [], history_reports=[], historical_snapshots=[period_snapshot, sap_snapshot])
    may = build_dashboard_view(5, 2026, [], history_reports=[], historical_snapshots=[period_snapshot, sap_snapshot])
    june_o6 = next(section for section in june["objective_sections"] if section["objective_code"] == "O6")
    may_o6 = next(section for section in may["objective_sections"] if section["objective_code"] == "O6")

    sap_visual = next(visual for visual in june_o6["visuals"] if visual["id"] == "o6_sap_compliance")
    assert sap_visual["kind"] == "sap_compliance"
    assert sap_visual["payload"]["totals"]["violating_wo"] == 120
    assert all(visual["id"] != "o6_sap_compliance" for visual in may_o6["visuals"])


def test_historical_snapshot_import_is_idempotent(db_session):
    workbook = Workbook()
    dashboard = workbook.active
    dashboard.title = "Dashboard"
    dashboard["A20"] = "LŨY KẾ NĂM 2026"
    dashboard["A22"] = "Tổ trực ca"
    dashboard["F22"] = "HT tốt"
    workbook.create_sheet("data")
    from io import BytesIO

    buffer = BytesIO()
    workbook.save(buffer)
    payload = buffer.getvalue()

    first = import_historical_snapshot(db_session, payload, source_file_name="snapshot.xlsx", imported_by="admin")
    second = import_historical_snapshot(db_session, payload, source_file_name="snapshot.xlsx", imported_by="admin")

    assert first["imported_count"] > 0
    assert second["skipped_duplicates"] == first["imported_count"]
    records = db_session.query(HistoricalSnapshotModel).filter(HistoricalSnapshotModel.team == "TCĐK").all()
    assert len(records) == 1
    assert records[0].monthly_assessment == "HT tốt"


def test_extract_dashboard_note_blocks_from_drawing_text():
    drawing_xml = """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
    <xdr:wsDr xmlns:xdr="http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"
              xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main">
      <xdr:twoCellAnchor>
        <xdr:sp><xdr:txBody>
          <a:bodyPr/><a:lstStyle/>
          <a:p><a:r><a:t>KR02 BÁO CÁO CÔNG TÁC BDĐK</a:t></a:r></a:p>
          <a:p><a:r><a:t>* Đội TBĐ</a:t></a:r></a:p>
          <a:p><a:r><a:t>25 mục không thực hiện do điều kiện khách quan</a:t></a:r></a:p>
        </xdr:txBody></xdr:sp>
      </xdr:twoCellAnchor>
      <xdr:twoCellAnchor>
        <xdr:sp><xdr:txBody>
          <a:bodyPr/><a:lstStyle/>
          <a:p><a:r><a:t>KR03 BÁO CÁO CÔNG TÁC SỬA CHỮA ĐỘT XUẤT</a:t></a:r></a:p>
          <a:p><a:r><a:t>Tồn đọng 7 hạng mục</a:t></a:r></a:p>
        </xdr:txBody></xdr:sp>
      </xdr:twoCellAnchor>
    </xdr:wsDr>""".encode()
    buffer = BytesIO()
    with ZipFile(buffer, "w") as archive:
        archive.writestr("xl/drawings/drawing1.xml", drawing_xml)

    notes = _extract_dashboard_note_blocks(buffer.getvalue())

    assert notes["o2_bddk"] == ["* Đội TBĐ", "25 mục không thực hiện do điều kiện khách quan"]
    assert notes["o2_scdx"] == ["Tồn đọng 7 hạng mục"]
