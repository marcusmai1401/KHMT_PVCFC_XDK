from dataclasses import asdict, dataclass
import logging
from pathlib import Path
import re

from openpyxl import load_workbook

from app.core.config import settings
from app.services.okr.constants import DASHBOARD_COLUMNS


logger = logging.getLogger(__name__)


@dataclass(frozen=True)
class KRMapping:
    workshop_kr_code: str
    kr_name: str
    dashboard_column: str
    measurement_type: str
    target_value: str
    source_row: int | None = None

    def to_dict(self) -> dict:
        return asdict(self)


FALLBACK_NAMES = {
    "O1": 3,
    "O2": 6,
    "O3": 3,
    "O4": 6,
    "O5": 15,
    "O6": 4,
}


def fallback_kr_mapping() -> list[KRMapping]:
    records: list[KRMapping] = []
    col_index = 0
    for objective, count in FALLBACK_NAMES.items():
        for i in range(1, count + 1):
            records.append(
                KRMapping(
                    workshop_kr_code=f"{objective}.KR{i}",
                    kr_name=f"{objective}.KR{i}",
                    dashboard_column=DASHBOARD_COLUMNS[col_index],
                    measurement_type="Unknown",
                    target_value="",
                )
            )
            col_index += 1
    return records


def extract_workshop_kr_code(raw_code: str) -> str | None:
    match = re.search(r"ĐK\.(O[1-6])\.KR(\d+)", raw_code or "")
    if match:
        return f"{match.group(1)}.KR{int(match.group(2))}"
    match = re.fullmatch(r"(O[1-6])\.KR(\d+)", raw_code or "")
    if match:
        return f"{match.group(1)}.KR{int(match.group(2))}"
    return None


def candidate_master_workbook_paths() -> list[Path]:
    return [
        settings.source_okr_workbook,
        settings.workspace_dir / "template_xlsx" / "OKR_Workshop.xlsx",
        settings.workspace_dir / "KHMT_T1_T2_T3_T4" / "OKR tháng 04-2026 - X.ĐK.xlsx",
    ]


def _resolve_master_workbook_path(workbook_path: Path | None = None) -> Path | None:
    if workbook_path is not None and workbook_path.exists():
        return workbook_path
    for candidate in candidate_master_workbook_paths():
        if candidate.exists():
            return candidate
    return None


def load_master_kr_mapping(workbook_path: Path | None = None, *, allow_fallback: bool = True) -> list[KRMapping]:
    path = _resolve_master_workbook_path(workbook_path)
    if path is None:
        if allow_fallback:
            logger.warning("Using generated fallback KR mapping because no canonical workbook was found")
            return fallback_kr_mapping()
        raise FileNotFoundError("No canonical KR mapping workbook found")
    workbook = load_workbook(path, read_only=True, data_only=False, keep_links=False)
    try:
        if "OKR X.ĐK 2026" not in workbook.sheetnames:
            if allow_fallback:
                logger.warning(
                    "Using generated fallback KR mapping because %s has no sheet OKR X.ĐK 2026",
                    path,
                )
                return fallback_kr_mapping()
            raise ValueError(f"Workbook {path} does not contain sheet OKR X.ĐK 2026")
        sheet = workbook["OKR X.ĐK 2026"]
        records: list[KRMapping] = []
        for row in range(5, 48):
            raw_code = str(sheet.cell(row, 5).value or "")
            workshop_code = extract_workshop_kr_code(raw_code)
            if not workshop_code:
                continue
            records.append(
                KRMapping(
                    workshop_kr_code=workshop_code,
                    kr_name=str(sheet.cell(row, 6).value or "").strip(),
                    dashboard_column=DASHBOARD_COLUMNS[len(records)],
                    measurement_type=str(sheet.cell(row, 7).value or "").strip(),
                    target_value=str(sheet.cell(row, 8).value or "").strip(),
                    source_row=row,
                )
            )
        if len(records) != 37:
            if not allow_fallback:
                raise ValueError(f"Expected 37 KR mapping rows from {path}, found {len(records)}")
            logger.warning(
                "Using generated fallback KR mapping because %s yielded %s rows instead of 37",
                path,
                len(records),
            )
            return fallback_kr_mapping()
        return records
    finally:
        workbook.close()


def mapping_by_code(workbook_path: Path | None = None, *, allow_fallback: bool = True) -> dict[str, KRMapping]:
    return {
        record.workshop_kr_code: record
        for record in load_master_kr_mapping(workbook_path, allow_fallback=allow_fallback)
    }
