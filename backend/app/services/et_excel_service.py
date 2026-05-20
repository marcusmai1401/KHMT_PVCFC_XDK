from __future__ import annotations

from datetime import timedelta
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy.orm import Session

from app.core.config import settings
from app.models.et_domain import CompetencyFramework, CompetencyItem, LearningPlan, Personnel
from app.schemas.et_schemas import FrameworkCreate, FrameworkItemCreate, PersonnelCreate
from app.services import et_service
from app.services.et_gap_calculator import EXCLUDED_CATEGORY
from app.services.et_service import ETValidationError


FRAMEWORK_TITLE = "KHUNG NĂNG LỰC CHUYÊN MÔN"
CATEGORIES = {"Cơ bản", "Trung cấp", "Nâng cao", "Nghiệp vụ hành chính"}


def import_frameworks_from_excel(db: Session, file_path: Path, actor: str) -> list[CompetencyFramework]:
    workbook = load_workbook(file_path, data_only=True)
    master_catalog = parse_master_catalog(workbook["Ma trận năng lực"]) if "Ma trận năng lực" in workbook.sheetnames else {}
    parsed = []
    errors = []
    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        if not sheet_name.startswith("KNL_ĐK_") or str(worksheet["A1"].value or "").strip() != FRAMEWORK_TITLE:
            continue
        try:
            parsed.append(parse_framework_from_sheet(worksheet, sheet_name, master_catalog))
        except ETValidationError as exc:
            errors.extend(exc.errors or [{"sheet": sheet_name, "message": exc.message}])
    if not parsed:
        errors.append({"message": "No KNL_ĐK_* framework sheets were found"})
    if errors:
        raise ETValidationError("Import validation failed", errors, status_code=422)

    created = []
    for framework_data in parsed:
        created.append(et_service.create_framework(db, FrameworkCreate(**framework_data), actor))
    return created


def parse_master_catalog(worksheet) -> dict[str, dict[str, Any]]:
    catalog: dict[str, dict[str, Any]] = {}
    for row in range(4, worksheet.max_row + 1):
        code = _clean(worksheet.cell(row, 4).value)
        if not code:
            continue
        catalog[code] = {
            "category": _clean(worksheet.cell(row, 1).value),
            "competency_name": _clean(worksheet.cell(row, 2).value),
            "competency_detail": _clean(worksheet.cell(row, 5).value),
        }
    return catalog


def parse_framework_from_sheet(worksheet, sheet_name: str, master_catalog: dict[str, dict[str, Any]]) -> dict[str, Any]:
    errors = []
    current_category = None
    seen_codes: set[str] = set()
    items = []
    for row in range(6, worksheet.max_row + 1):
        category = _clean(worksheet.cell(row, 1).value)
        if category:
            current_category = category
        code = _clean(worksheet.cell(row, 4).value)
        if not code:
            continue
        master = master_catalog.get(code, {})
        if code in seen_codes:
            errors.append({"sheet": sheet_name, "row": row, "column": "D", "message": f"Duplicate competency code: {code}"})
            continue
        seen_codes.add(code)
        name = _clean(worksheet.cell(row, 2).value) or master.get("competency_name")
        detail = _clean(worksheet.cell(row, 5).value) or master.get("competency_detail")
        category_value = current_category or master.get("category")
        if not category_value:
            errors.append({"sheet": sheet_name, "row": row, "column": "A", "message": "Missing category"})
        if not name:
            errors.append({"sheet": sheet_name, "row": row, "column": "B", "message": "Missing competency name"})
        level_requirements: dict[str, int] = {}
        for level in range(1, 9):
            value = worksheet.cell(row, 5 + level).value
            try:
                level_requirements[str(level)] = 0 if value in (None, "") else int(value)
            except (TypeError, ValueError):
                errors.append(
                    {
                        "sheet": sheet_name,
                        "row": row,
                        "column": chr(ord("F") + level - 1),
                        "message": "Invalid level requirement value",
                        "value": value,
                    }
                )
        try:
            stt = int(worksheet.cell(row, 3).value or len(items) + 1)
        except (TypeError, ValueError):
            stt = len(items) + 1
        items.append(
            FrameworkItemCreate(
                nlcm_code=code,
                competency_name=name or code,
                competency_detail=detail,
                category=category_value or "Cơ bản",
                stt=stt,
                level_requirements=level_requirements,
            ).model_dump()
        )
    if errors:
        raise ETValidationError("Framework sheet validation failed", errors, status_code=422)
    return {
        "code": sheet_name,
        "title": _clean(worksheet["A2"].value) or sheet_name,
        "is_active": True,
        "items": items,
    }


def import_personnel_from_excel(db: Session, file_path: Path, actor: str) -> list[Personnel]:
    workbook = load_workbook(file_path, data_only=True)
    worksheet = workbook.active
    headers = {_normalize_header(worksheet.cell(1, col).value): col for col in range(1, worksheet.max_column + 1)}
    required = {
        "employee_code": {"ma nhan vien", "employee code", "employee_code", "ma nv"},
        "full_name": {"ho ten", "full name", "full_name"},
        "position_code": {"vi tri chuc danh", "position code", "position_code", "ma khung nang luc"},
        "team": {"team", "to", "doi to"},
        "current_level": {"bac hien tai", "current level", "current_level"},
    }
    mapping: dict[str, int] = {}
    errors = []
    for field, aliases in required.items():
        for alias in aliases:
            if alias in headers:
                mapping[field] = headers[alias]
                break
        if field not in mapping:
            errors.append({"row": 1, "field": field, "message": "Missing required personnel import column"})
    if errors:
        raise ETValidationError("Personnel import validation failed", errors, 422)
    optional = {
        "hire_date": {"ngay bat dau lam viec", "hire date", "hire_date"},
        "status": {"trang thai", "status"},
        "user_id": {"user id", "user_id", "tai khoan"},
    }
    for field, aliases in optional.items():
        for alias in aliases:
            if alias in headers:
                mapping[field] = headers[alias]
                break
    rows = []
    for row in range(2, worksheet.max_row + 1):
        if not any(worksheet.cell(row, col).value for col in mapping.values()):
            continue
        data = {field: worksheet.cell(row, col).value for field, col in mapping.items()}
        data["employee_code"] = _clean(data.get("employee_code"))
        data["full_name"] = _clean(data.get("full_name"))
        data["position_code"] = _clean(data.get("position_code"))
        data["team"] = _clean(data.get("team"))
        data["status"] = _status_to_code(_clean(data.get("status")) or "active")
        try:
            data["current_level"] = int(data.get("current_level"))
            rows.append(PersonnelCreate(**data))
        except Exception as exc:
            errors.append({"row": row, "message": str(exc)})
    if errors:
        raise ETValidationError("Personnel import validation failed", errors, 422)
    created = []
    for row in rows:
        created.append(et_service.create_personnel(db, row, actor))
    return created


def export_framework_to_excel(framework: CompetencyFramework) -> Path:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = framework.code[:31]
    sheet["A1"] = FRAMEWORK_TITLE
    sheet["A2"] = framework.title
    headers = ["Phân nhóm", "Năng lực chuyên môn", "Stt", "Mã năng lực", "Chi tiết năng lực chuyên môn"]
    headers.extend([f"Bậc {level}" for level in range(1, 9)])
    sheet.append(headers)
    item_count = len(framework.items)
    data_end_row = max(6, 5 + item_count)
    sheet.append(["", "", "", "", "", *[f"=SUM({col}6:{col}{data_end_row})" for col in "FGHIJKLM"]])
    sheet.append(["", "", "", "", "", *range(1, 9)])
    for item in sorted(framework.items, key=lambda row: (row.category, row.stt)):
        sheet.append(
            [
                item.category,
                item.competency_name,
                item.stt,
                item.nlcm_code,
                item.competency_detail,
                *[item.level_requirements.get(str(level), 0) for level in range(1, 9)],
            ]
        )
    _style_sheet(sheet)
    return _save_workbook(workbook, f"{framework.code}-framework.xlsx")


def export_assessment_to_excel(assessment) -> Path:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Phiếu đánh giá"
    rows = [
        ["CN TỔNG CTY PHÂN BÓN DẦU KHÍ CÀ MAU - NHÀ MÁY ĐẠM CÀ MAU"],
        ["Xưởng Điều khiển"],
        [],
        ["PHIẾU ĐÁNH GIÁ NĂNG LỰC CHUYÊN MÔN"],
        [],
        ["I.", "Thông tin nhân sự được đánh giá"],
        ["Mã nhân viên:", assessment.personnel.employee_code, "", "Vị trí chức danh:", assessment.personnel.position_code],
        ["Họ tên:", assessment.personnel.full_name, "", "Bậc đánh giá:", assessment.personnel_level_at_assessment],
        [],
        ["II.", "Nội dung đánh giá"],
        ["Stt", "Mã NLCM", "Năng lực chuyên môn chi tiết", "Điểm chuẩn", "Điểm thực tế", "GAP", "Ghi chú"],
    ]
    for row in rows:
        sheet.append(row)
    for index, item in enumerate(sorted(assessment.items, key=lambda row: row.competency_item.stt), start=1):
        sheet.append(
            [
                index,
                item.competency_item.nlcm_code,
                item.competency_item.competency_detail or item.competency_item.competency_name,
                item.required_score,
                item.actual_score,
                item.gap,
                item.notes,
            ]
        )
    sheet.append([])
    sheet.append(["III.", "Kết quả đánh giá"])
    sheet.append(["Đạt", "X" if assessment.overall_result == "Đạt" else ""])
    sheet.append(["Không đạt", "X" if assessment.overall_result == "Không đạt" else ""])
    sheet.append(["Nhận xét/đánh giá:", assessment.notes or ""])
    sheet.append(["Nội dung đào tạo:", assessment.training_content or ""])
    _style_sheet(sheet)
    return _save_workbook(workbook, f"phieu-danh-gia-{assessment.personnel.employee_code}-{assessment.assessment_period}.xlsx")


def export_learning_plan_to_excel(plan: LearningPlan) -> Path:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Project Timeline"
    sheet["B1"] = "KẾ HOẠCH HỌC TẬP NHÂN SỰ MỚI"
    sheet["D3"] = f"{plan.personnel.full_name}\n{plan.start_date.strftime('%d/%m/%Y')}"
    sheet["B6"] = "PLAN WEEK"
    last_year = None
    last_quarter = None
    last_relative_month = None
    for week in range(1, plan.duration_months * 4 + 1):
        col = 4 + week
        target_date = plan.start_date + timedelta(days=(week - 1) * 7)
        relative_month = (target_date.year - plan.start_date.year) * 12 + target_date.month - plan.start_date.month + 1
        quarter = f"Q{((target_date.month - 1) // 3) + 1}"
        if target_date.year != last_year:
            sheet.cell(2, col).value = target_date.year
            last_year = target_date.year
        if quarter != last_quarter:
            sheet.cell(3, col).value = quarter
            last_quarter = quarter
        if relative_month != last_relative_month:
            sheet.cell(4, col).value = f"Tháng {relative_month}"
            last_relative_month = relative_month
        sheet.cell(5, col).value = target_date.day
        sheet.cell(6, col).value = week
    row_index = 7
    for category in ["Cơ bản", "Trung cấp", "Nâng cao", EXCLUDED_CATEGORY]:
        items = [item for item in plan.items if item.competency_item.category == category]
        if not items:
            continue
        sheet.cell(row_index, 2).value = category.upper()
        sheet.cell(row_index, 2).font = Font(bold=True)
        row_index += 1
        for number, item in enumerate(sorted(items, key=lambda row: row.competency_item.stt), start=1):
            sheet.cell(row_index, 1).value = number
            sheet.cell(row_index, 2).value = item.competency_item.competency_name
            sheet.cell(row_index, 3).value = item.competency_item.competency_detail
            if item.target_week:
                sheet.cell(row_index, 4 + item.target_week).value = item.target_level
            row_index += 1
    _style_sheet(sheet)
    return _save_workbook(workbook, f"ke-hoach-hoc-tap-{plan.personnel.employee_code}.xlsx")


def export_dashboard_summary_to_excel(data: dict[str, Any]) -> Path:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Dashboard"
    sheet.append(["Tổng nhân sự", data["aggregate"]["total_active_personnel"]])
    sheet.append(["Đạt", data["aggregate"]["pass_count"], data["aggregate"]["pass_percentage"]])
    sheet.append(["Không đạt", data["aggregate"]["fail_count"]])
    sheet.append(["Chưa đánh giá", data["aggregate"]["not_assessed_count"]])
    sheet.append(["Đang đánh giá", data["aggregate"]["draft_count"]])
    sheet.append([])
    sheet.append(["Mã NV", "Họ tên", "Team", "Vị trí", "Bậc", "Đạt", "GAP", "Tổng GAP", "Kết quả"])
    for row in data["rows"]:
        sheet.append(
            [
                row["employee_code"],
                row["full_name"],
                row["team"],
                row["position_code"],
                row["current_level"],
                row["achieved_count"],
                row["gap_count"],
                row["total_gap"],
                row["overall_result"],
            ]
        )
    _style_sheet(sheet)
    return _save_workbook(workbook, "et-dashboard-summary.xlsx")


def _save_workbook(workbook: Workbook, filename: str) -> Path:
    export_dir = settings.storage_dir / "exports"
    export_dir.mkdir(parents=True, exist_ok=True)
    path = export_dir / filename
    workbook.save(path)
    return path


def _style_sheet(sheet) -> None:
    header_fill = PatternFill("solid", fgColor="EDF2F7")
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if cell.row in {1, 3, 4, 11}:
                cell.font = Font(bold=True)
                cell.fill = header_fill
    for column_cells in sheet.columns:
        sheet.column_dimensions[column_cells[0].column_letter].width = min(max(len(str(cell.value or "")) for cell in column_cells) + 2, 42)


def _status_to_code(value: str) -> str:
    normalized = value.casefold()
    if normalized in {"nghỉ việc", "nghi viec", "inactive"}:
        return "inactive"
    if normalized in {"chuyển đi", "chuyen di", "transferred"}:
        return "transferred"
    return "active"


def _normalize_header(value: Any) -> str:
    text = _clean(value).casefold().replace("_", " ")
    replacements = {
        "ã": "a",
        "á": "a",
        "à": "a",
        "ả": "a",
        "ạ": "a",
        "ă": "a",
        "ắ": "a",
        "ằ": "a",
        "ẳ": "a",
        "ẵ": "a",
        "ặ": "a",
        "â": "a",
        "ấ": "a",
        "ầ": "a",
        "ẩ": "a",
        "ẫ": "a",
        "ậ": "a",
        "đ": "d",
        "é": "e",
        "è": "e",
        "ẻ": "e",
        "ẽ": "e",
        "ẹ": "e",
        "ê": "e",
        "ế": "e",
        "ề": "e",
        "ể": "e",
        "ễ": "e",
        "ệ": "e",
        "í": "i",
        "ì": "i",
        "ỉ": "i",
        "ĩ": "i",
        "ị": "i",
        "ó": "o",
        "ò": "o",
        "ỏ": "o",
        "õ": "o",
        "ọ": "o",
        "ô": "o",
        "ố": "o",
        "ồ": "o",
        "ổ": "o",
        "ỗ": "o",
        "ộ": "o",
        "ơ": "o",
        "ớ": "o",
        "ờ": "o",
        "ở": "o",
        "ỡ": "o",
        "ợ": "o",
        "ú": "u",
        "ù": "u",
        "ủ": "u",
        "ũ": "u",
        "ụ": "u",
        "ư": "u",
        "ứ": "u",
        "ừ": "u",
        "ử": "u",
        "ữ": "u",
        "ự": "u",
        "ý": "y",
        "ỳ": "y",
        "ỷ": "y",
        "ỹ": "y",
        "ỵ": "y",
    }
    for source, target in replacements.items():
        text = text.replace(source, target)
    return " ".join(text.split())


def _clean(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    return "" if text == "#N/A" else text
