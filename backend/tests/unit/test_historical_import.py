from datetime import datetime
from io import BytesIO

from openpyxl import Workbook
from sqlalchemy import select

from app.models.domain import HistoricalSnapshotModel, TeamMonthlySummaryModel, TeamReportModel
from app.services.okr.historical_import import (
    apply_dashboard_matrix_statuses,
    apply_discipline_overrides,
    discover_source_files,
    extract_dashboard_matrix_lookup,
    extract_month_year_from_filename,
    upsert_team_monthly_summary,
    upsert_team_report,
)
from app.services.okr.historical_snapshot import import_historical_snapshot
from app.services.okr.kr_mapping import KRMapping
from app.services.okr.workbook import get_report_columns_for_month, parse_team_report


def _mapping() -> dict[str, KRMapping]:
    return {
        "O1.KR1": KRMapping(
            workshop_kr_code="O1.KR1",
            kr_name="KR 1",
            dashboard_column="L",
            measurement_type="Text",
            target_value="1",
        )
    }


def test_extract_month_year_from_historical_filename():
    assert extract_month_year_from_filename("OKR tháng 04-2026 - X.ĐK.xlsx") == (4, 2026)


def test_discovery_filters_months_and_ignores_excel_lock_files(tmp_path):
    (tmp_path / "OKR tháng 05-2026 - X.ĐK.xlsx").write_bytes(b"source")
    (tmp_path / "OKR tháng 06-2026 - X.ĐK.xlsx").write_bytes(b"source")
    (tmp_path / "~$OKR tháng 05-2026 - X.ĐK.xlsx").write_bytes(b"lock")

    discovered = discover_source_files(tmp_path, months=(5,))

    assert [(item.month, item.file_name) for item in discovered] == [
        (5, "OKR tháng 05-2026 - X.ĐK.xlsx")
    ]


def test_parser_selects_requested_month_column_group_for_tbch(tmp_path):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "TBCH"
    sheet["Q2"] = "Tình hình thực hiện"
    sheet["R2"] = "Đánh giá"
    sheet["S2"] = "Ghi chú"
    sheet["T2"] = "Tình hình thực hiện"
    sheet["U2"] = "Đánh giá"
    sheet["V2"] = "Ghi chú"
    sheet["B4"] = "ĐK.O1.KR1.TBCH.O1.KR1"
    sheet["C4"] = "KR 1"
    sheet["Q4"] = "T1 implementation"
    sheet["R4"] = "Hoàn thành"
    sheet["T4"] = "T2 implementation"
    sheet["U4"] = "Hoàn thành tốt"
    path = tmp_path / "OKR tháng 02-2026 - X.ĐK.xlsx"
    workbook.save(path)

    parsed = parse_team_report(path, team="TBCH", month=2, year=2026, kr_mapping=_mapping())

    assert parsed["assessments"][0]["implementation_report"] == "T2 implementation"
    assert parsed["assessments"][0]["team_self_assessment"] == "Hoàn thành tốt"
    assert parsed["assessments"][0]["source_cells"]["implementation_report"]["column"] == "T"


def test_tbch_month_one_uses_explicit_p_q_r_group():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "TBCH"
    sheet["P2"] = "Tình hình thực hiện"
    sheet["Q2"] = "Đánh giá"
    sheet["R2"] = "Ghi chú"
    sheet["T2"] = "Tình hình thực hiện"
    sheet["U2"] = "Đánh giá"
    sheet["V2"] = "Ghi chú"

    assert get_report_columns_for_month(sheet, "TBCH", 1) == (16, 17, 18)


def test_multi_month_sheet_requires_report_month():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "TBHTĐK"
    sheet["M3"] = "Tình hình thực hiện"
    sheet["N3"] = "Đánh giá"
    sheet["O3"] = "Ghi chú"
    sheet["P3"] = "Tình hình thực hiện"
    sheet["Q3"] = "Đánh giá"
    sheet["R3"] = "Ghi chú"

    try:
        get_report_columns_for_month(sheet, "TBHTĐK", None)
    except ValueError as exc:
        assert "Reporting month is required" in str(exc)
    else:
        raise AssertionError("Expected ValueError for multi-month sheet without month")


def test_parser_accepts_compact_merged_month_header_and_preserves_planning_context(tmp_path):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "TCĐK"
    sheet.merge_cells("N2:P2")
    sheet["N2"] = "Báo cáo tình hình thực hiện tháng 5"
    sheet["A4"] = 2
    sheet["B4"] = "ĐK.O1.KR1.TCĐK.O1.KR1"
    sheet["C4"] = "KR 1"
    sheet["D4"] = "Số sự cố"
    sheet["E4"] = 0
    sheet["F4"] = "Quý"
    sheet["G4"] = 0.15
    sheet["H4"] = "Kế hoạch hành động"
    sheet["I4"] = datetime(2026, 1, 1)
    sheet["J4"] = "31/12/2026"
    sheet["L4"] = "Nguyễn Văn A"
    sheet["M4"] = "Trưởng đội"
    sheet["N4"] = "Không phát sinh sự cố"
    sheet["O4"] = "Hoàn thành"
    path = tmp_path / "OKR tháng 05-2026 - X.ĐK.xlsx"
    workbook.save(path)

    parsed = parse_team_report(path, team="TCĐK", month=5, year=2026, kr_mapping=_mapping())
    assessment = parsed["assessments"][0]

    assert assessment["implementation_report"] == "Không phát sinh sự cố"
    assert assessment["planning_context"] == {
        "sequence": 2,
        "raw_team_kr_code": "ĐK.O1.KR1.TCĐK.O1.KR1",
        "team_kr_name": "KR 1",
        "measurement": "Số sự cố",
        "purpose": 0,
        "measurement_frequency": "Quý",
        "objective_weight": 0.15,
        "action_plan": "Kế hoạch hành động",
        "start_date": "2026-01-01T00:00:00",
        "expected_completion_date": "31/12/2026",
        "owner": "Nguyễn Văn A",
        "evaluator": "Trưởng đội",
    }
    assert assessment["source_cells"]["owner"]["column"] == "L"


def test_missing_assessment_is_not_warned_when_kr_has_no_plan(tmp_path):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "TBCH"
    sheet["N2"] = "Tình hình thực hiện"
    sheet["O2"] = "Đánh giá"
    sheet["P2"] = "Ghi chú"
    sheet["B4"] = "ĐK.O1.KR1.TBCH.O1.KR1"
    sheet["C4"] = "KR 1"
    sheet["N4"] = "Trong tháng không có chương trình"
    path = tmp_path / "OKR tháng 05-2026 - X.ĐK.xlsx"
    workbook.save(path)

    parsed = parse_team_report(path, team="TBCH", month=5, year=2026, kr_mapping=_mapping())

    assert parsed["assessments"][0]["has_plan"] is False
    assert not any(
        warning["warning_type"] == "MISSING_REQUIRED_FIELD" for warning in parsed["warnings"]
    )


def test_dashboard_matrix_status_is_authoritative_and_audited(tmp_path):
    workbook = Workbook()
    dashboard = workbook.active
    dashboard.title = "Dashboard"
    dashboard["A8"] = "Tổ trực ca"
    dashboard["L8"] = "GOOD"
    path = tmp_path / "OKR tháng 05-2026 - X.ĐK.xlsx"
    workbook.save(path)
    mapping = _mapping()

    matrix = extract_dashboard_matrix_lookup(path, mapping)
    parsed = {
        "team": "TCĐK",
        "assessments": [
            {
                "workshop_kr_code": "O1.KR1",
                "dashboard_status": "OK",
                "source_cells": {},
            }
        ],
        "warnings": [],
        "source_cell_references": [],
    }
    apply_dashboard_matrix_statuses(parsed, matrix, mapping)

    assessment = parsed["assessments"][0]
    assert assessment["dashboard_status"] == "GOOD"
    assert assessment["dashboard_status_from_dashboard"] is True
    assert assessment["source_cells"]["dashboard_status"]["column"] == "L"
    assert parsed["warnings"][0]["warning_type"] == "DASHBOARD_STATUS_MISMATCH"


def test_historical_snapshot_import_detects_dashboard_history_row_dynamically(db_session):
    workbook = Workbook()
    dashboard = workbook.active
    dashboard.title = "Dashboard"
    dashboard["A1"] = "BÁO CÁO KẾ HOẠCH MỤC TIÊU THÁNG 04 NĂM 2026"
    dashboard["A7"] = "Đội/Tổ"
    dashboard["A8"] = "HTĐK"
    dashboard["F8"] = "HT"
    dashboard["H8"] = "HT tốt"
    workbook.create_sheet("data")
    buffer = BytesIO()
    workbook.save(buffer)

    result = import_historical_snapshot(
        db_session,
        buffer.getvalue(),
        source_file_name="dynamic-dashboard.xlsx",
        imported_by="test",
    )

    assert result["imported_count"] >= 2
    rows = db_session.execute(
        select(HistoricalSnapshotModel).where(HistoricalSnapshotModel.team == "TBHTĐK")
    ).scalars().all()
    assert {(row.month, row.monthly_assessment, row.source_range) for row in rows} >= {
        (1, "HT", "Dashboard!F8"),
        (2, "HT tốt", "Dashboard!H8"),
    }


def test_team_report_upsert_keeps_one_current_version_and_updates_summary(db_session, tmp_path):
    source = tmp_path / "source.xlsx"
    source.write_bytes(b"historical source")
    parsed = {
        "team": "TBCH",
        "report_month": 4,
        "report_year": 2026,
        "sheet_name": "TBCH",
        "assessments": [{"workshop_kr_code": "O1.KR1", "dashboard_status": "OK"}],
        "team_level": {"monthly_assessment": "Hoàn thành", "discipline_status": "OK"},
        "source_cell_references": [],
    }

    first, first_action = upsert_team_report(
        db_session,
        parsed,
        source_file=source,
        source_hash="hash-1",
        imported_by="test",
    )
    upsert_team_monthly_summary(db_session, parsed, source_file=source, source_hash="hash-1")
    parsed["team_level"] = {"monthly_assessment": "Không hoàn thành", "discipline_status": "NOK"}
    second, second_action = upsert_team_report(
        db_session,
        parsed,
        source_file=source,
        source_hash="hash-2",
        imported_by="test",
    )
    upsert_team_monthly_summary(db_session, parsed, source_file=source, source_hash="hash-2")
    db_session.flush()

    assert first_action == "inserted"
    assert second_action == "updated"
    assert first.id != second.id
    current = db_session.execute(
        select(TeamReportModel).where(
            TeamReportModel.team == "TBCH",
            TeamReportModel.report_month == 4,
            TeamReportModel.report_year == 2026,
            TeamReportModel.is_current_version.is_(True),
        )
    ).scalars().all()
    assert [row.id for row in current] == [second.id]
    summary = db_session.execute(
        select(TeamMonthlySummaryModel).where(
            TeamMonthlySummaryModel.team == "TBCH",
            TeamMonthlySummaryModel.month == 4,
            TeamMonthlySummaryModel.year == 2026,
        )
    ).scalar_one()
    assert summary.monthly_assessment == "Không hoàn thành"
    assert summary.discipline_status == "NOK"


def test_team_report_and_summary_upserts_skip_identical_source(db_session, tmp_path):
    source = tmp_path / "source.xlsx"
    source.write_bytes(b"historical source")
    parsed = {
        "team": "TBCH",
        "report_month": 5,
        "report_year": 2026,
        "sheet_name": "TBCH",
        "assessments": [{"workshop_kr_code": "O1.KR1", "dashboard_status": "OK"}],
        "team_level": {"monthly_assessment": "Hoàn thành", "discipline_status": "OK"},
        "source_cell_references": [],
    }

    first, first_action = upsert_team_report(
        db_session,
        parsed,
        source_file=source,
        source_hash="same-hash",
        imported_by="test",
    )
    _summary, first_summary_action = upsert_team_monthly_summary(
        db_session,
        parsed,
        source_file=source,
        source_hash="same-hash",
    )
    second, second_action = upsert_team_report(
        db_session,
        parsed,
        source_file=source,
        source_hash="same-hash",
        imported_by="test",
    )
    _summary, second_summary_action = upsert_team_monthly_summary(
        db_session,
        parsed,
        source_file=source,
        source_hash="same-hash",
    )

    assert first_action == "inserted"
    assert first_summary_action == "inserted"
    assert second_action == "skipped"
    assert second_summary_action == "skipped"
    assert second.id == first.id


def test_t4_discipline_override_lowers_monthly_assessment():
    updated = apply_discipline_overrides(
        {"monthly_assessment": "Hoàn thành", "discipline_status": "OK"},
        "TBĐL",
        4,
    )

    assert updated["discipline_status"] == "NOK"
    assert updated["monthly_assessment"] == "Không hoàn thành"
    assert "giờ công" in updated["discipline_description"]
