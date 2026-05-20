from openpyxl import Workbook, load_workbook

from app.services.okr.report_template import generate_standard_report_template
from app.services.okr.workbook import parse_team_report
from app.services.okr.workbook import _team_level


def test_team_level_parses_adjacent_and_inline_values_without_label_collision():
    workbook = Workbook()
    sheet = workbook.active
    sheet["B2"] = "Kỷ luật"
    sheet["C2"] = "NOK - vi phạm nội quy"
    sheet["B3"] = "Mô tả kỷ luật"
    sheet["C3"] = "Đi muộn"
    sheet["B4"] = "KR liên quan"
    sheet["C4"] = "O1.KR1"
    sheet["B5"] = "Đánh giá chung: Hoàn thành tốt"

    summary, warnings = _team_level(sheet, include_warnings=True)

    assert summary["discipline_status"] == "NOK"
    assert summary["discipline_description"] == "Đi muộn"
    assert summary["related_kr"] == "O1.KR1"
    assert summary["monthly_assessment"] == "Hoàn thành tốt"
    assert warnings == []


def test_team_level_missing_required_value_creates_warning():
    workbook = Workbook()
    sheet = workbook.active
    sheet["A1"] = "Discipline Status"
    sheet["A2"] = "Monthly Assessment"
    sheet["B2"] = "Hoàn thành"

    summary, warnings = _team_level(sheet, include_warnings=True)

    assert summary["monthly_assessment"] == "Hoàn thành"
    assert any(warning["warning_type"] == "MISSING_REQUIRED_FIELD" for warning in warnings)


def test_standard_report_template_is_parseable_with_metadata(tmp_path):
    path = generate_standard_report_template(output_path=tmp_path / "okr-team-report-template.xlsx")
    workbook = load_workbook(path)
    assert workbook.sheetnames == ["Team_Report"]
    sheet = workbook["Team_Report"]
    assert sheet["A1"].value == "TEMPLATE BÁO CÁO KẾ HOẠCH MỤC TIÊU XƯỞNG ĐIỀU KHIỂN"
    assert sheet.max_row == 44
    assert sheet.max_column == 16
    sheet["B2"] = "TBĐL"
    sheet["N7"] = "Không có kế hoạch trong tháng"
    sheet["O7"] = "N/A"
    workbook.save(path)

    parsed = parse_team_report(path)

    assert parsed["team"] == "TBĐL"
    assert parsed["report_month"] == 4
    assert parsed["report_year"] == 2026
    assert len(parsed["assessments"]) == 37
    assert parsed["assessments"][0]["workshop_kr_code"] == "O1.KR1"
