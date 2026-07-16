from __future__ import annotations

import json
import sys
from pathlib import Path

BACKEND_DIR = Path(__file__).resolve().parents[1]
if str(BACKEND_DIR) not in sys.path:
    sys.path.insert(0, str(BACKEND_DIR))

from openpyxl import load_workbook
from sqlalchemy import select

from app.core.config import settings
from app.db.session import create_session
from app.models.domain import (
    HistoricalSnapshotModel,
    TeamMonthlySummaryModel,
    TeamReportModel,
)
from app.services.okr.constants import TEAMS
from app.services.okr.rules import normalize_assessment
from app.services.okr.team_normalizer import normalize_team_label


MONTHS = [1, 2, 3, 4, 5]
YEAR = 2026
SOURCE_DIR = settings.workspace_dir / "KHMT_Monthly"
TEMPLATE_DIR = settings.workspace_dir / "template_xlsx"

SOURCE_FILES = {
    1: SOURCE_DIR / "OKR tháng 01-2026 - X.ĐK.xlsx",
    2: SOURCE_DIR / "OKR tháng 02-2026 - X.ĐK.xlsx",
    3: SOURCE_DIR / "OKR tháng 03-2026 - X.ĐK.xlsx",
    4: SOURCE_DIR / "OKR tháng 04-2026 - X.ĐK.xlsx",
    5: SOURCE_DIR / "OKR tháng 05-2026 - X.ĐK.xlsx",
}

TEMPLATE_FILES = {
    "TBHTĐK": TEMPLATE_DIR / "TBHTĐK.xlsx",
    "TBCH": TEMPLATE_DIR / "TBCH.xlsx",
    "TBĐL": TEMPLATE_DIR / "TBĐL.xlsx",
    "TCĐK": TEMPLATE_DIR / "TCĐK.xlsx",
}

TEAM_ROW_RANGES = {
    1: range(22, 28),
    2: range(22, 28),
    3: range(22, 28),
    4: range(20, 26),
    5: range(20, 27),
}

DASHBOARD_MONTH_COLS = {1: 6, 2: 8, 3: 10, 4: 12, 5: 14}


def read_dashboard_cell(file_path: Path, month: int) -> dict[str, str]:
    if not file_path.exists():
        return {}
    wb = load_workbook(file_path, read_only=True, data_only=True, keep_links=False)
    if "Dashboard" not in wb.sheetnames:
        wb.close()
        return {}
    sheet = wb["Dashboard"]
    col = DASHBOARD_MONTH_COLS[month]
    result: dict[str, str] = {}
    for row in TEAM_ROW_RANGES.get(month, range(22, 28)):
        label = str(sheet.cell(row, 1).value or "").strip()
        team, _ = normalize_team_label(label)
        if not team:
            continue
        val = str(sheet.cell(row, col).value or "").strip()
        if val:
            result[team] = val
    wb.close()
    return result


def read_team_level_from_xlsx(file_path: Path, team: str, month: int) -> dict[str, str | None]:
    if not file_path.exists():
        return {"error": "file not found"}
    wb = load_workbook(file_path, read_only=False, data_only=True, keep_links=False)
    target_sheet = None
    for name in wb.sheetnames:
        normalized, _ = normalize_team_label(name)
        if normalized == team:
            target_sheet = wb[name]
            break
    if target_sheet is None:
        for name in wb.sheetnames:
            nl = name.upper().strip()
            if team == "TBHTĐK" and ("HTĐK" in nl or "HTDK" in nl or "HỆ THỐNG" in nl):
                target_sheet = wb[name]
                break
            if team == "TBCH" and ("TBCH" in nl or "CHẤP HÀNH" in nl):
                target_sheet = wb[name]
                break
            if team == "TBĐL" and ("TBĐL" in nl or "TBDL" in nl or "TBĐ" in nl or "ĐO LƯỜNG" in nl):
                target_sheet = wb[name]
                break
            if team == "TCĐK" and ("TCĐK" in nl or "TCDK" in nl or "TRỰC CA" in nl):
                target_sheet = wb[name]
                break
    if target_sheet is None:
        wb.close()
        return {"error": f"sheet for {team} not found"}
    result: dict[str, str | None] = {
        "sheet_name": target_sheet.title,
        "monthly_assessment": None,
        "discipline_status": None,
    }
    for row in range(1, min(target_sheet.max_row, 60) + 1):
        for col in range(1, min(target_sheet.max_column, 22) + 1):
            raw = str(target_sheet.cell(row, col).value or "").strip().lower()
            if any(
                label in raw
                for label in (
                    "đánh giá chung",
                    "đánh giá tháng",
                    "danh gia chung",
                    "kết quả đánh giá",
                    "kết luận chung",
                )
            ):
                for c2 in range(col + 1, target_sheet.max_column + 1):
                    v = str(target_sheet.cell(row, c2).value or "").strip()
                    if v:
                        result["monthly_assessment"] = v
                        break
            if "kỷ luật" in raw or "ky luat" in raw or "discipline" in raw:
                for c2 in range(col + 1, min(target_sheet.max_column, col + 5) + 1):
                    v = str(target_sheet.cell(row, c2).value or "").strip()
                    if v:
                        result["discipline_status"] = v
                        break
    if team == "TBHTĐK":
        assessment_rows = [(39, 14), (39, 17), (39, 20), (39, 23)]
        idx = month - 1
        if idx < len(assessment_rows):
            r, c = assessment_rows[idx]
            v = str(target_sheet.cell(r, c).value or "").strip()
            if v:
                result["summary_cell"] = f"{target_sheet.cell(r, c).coordinate} = {v}"
    elif team == "TBCH":
        assessment_rows = [(44, 16), (44, 20), (44, 23), (44, 26)]
        idx = month - 1
        if idx < len(assessment_rows):
            r, c = assessment_rows[idx]
            v = str(target_sheet.cell(r, c).value or "").strip()
            if v:
                result["summary_cell"] = f"{target_sheet.cell(r, c).coordinate} = {v}"
    elif team == "TCĐK":
        v = str(target_sheet.cell(42, 15).value or "").strip()
        if v:
            result["summary_cell"] = f"O42 = {v}"
    elif team == "TBĐL":
        for row in range(1, min(target_sheet.max_row, 60) + 1):
            for col in range(1, min(target_sheet.max_column, 16) + 1):
                label = str(target_sheet.cell(row, col).value or "").strip().lower()
                if "kết quả đánh giá" in label:
                    for c2 in range(col + 1, min(target_sheet.max_column, col + 5) + 1):
                        v = str(target_sheet.cell(row, c2).value or "").strip()
                        if v:
                            result["summary_cell"] = f"{target_sheet.cell(row, c2).coordinate} = {v}"
                            break
    wb.close()
    return result


def query_db_team_reports(db, month: int) -> dict[str, dict]:
    rows = db.execute(
        select(TeamReportModel).where(
            TeamReportModel.report_month == month,
            TeamReportModel.report_year == YEAR,
            TeamReportModel.is_current_version.is_(True),
            TeamReportModel.source_type == "historical_import",
        )
    ).scalars().all()
    result = {}
    for r in rows:
        team_level = r.team_level or {}
        assessments = r.assessments or []
        status_counts: dict[str, int] = {}
        for a in assessments:
            s = a.get("dashboard_status", "?")
            status_counts[s] = status_counts.get(s, 0) + 1
        result[r.team] = {
            "id": r.id,
            "sheet_name": r.sheet_name,
            "file_name": r.file_name,
            "assessment_count": len(assessments),
            "dashboard_status_counts": status_counts,
            "monthly_assessment": team_level.get("monthly_assessment"),
            "discipline_status": team_level.get("discipline_status"),
            "discipline_description": team_level.get("discipline_description"),
            "version": r.version,
        }
    return result


def query_db_summaries(db, month: int) -> dict[str, dict]:
    rows = db.execute(
        select(TeamMonthlySummaryModel).where(
            TeamMonthlySummaryModel.month == month,
            TeamMonthlySummaryModel.year == YEAR,
        )
    ).scalars().all()
    result = {}
    for r in rows:
        result[r.team] = {
            "monthly_assessment": r.monthly_assessment,
            "discipline_status": r.discipline_status,
            "discipline_description": r.discipline_description,
            "stats": r.stats,
        }
    return result


def query_db_snapshots(db, month: int) -> dict[str, dict]:
    rows = db.execute(
        select(HistoricalSnapshotModel).where(
            HistoricalSnapshotModel.month == month,
            HistoricalSnapshotModel.year == YEAR,
            HistoricalSnapshotModel.team.in_(TEAMS),
        )
    ).scalars().all()
    result = {}
    for r in rows:
        result[r.team] = {
            "source_file": r.source_file_name,
            "source_range": r.source_range,
            "monthly_assessment": r.monthly_assessment,
        }
    return result


def print_separator(char: str = "=", width: int = 100):
    print(char * width)


def print_month_report(month: int, db):
    print_separator()
    print(f"  THÁNG {month} / {YEAR}")
    print_separator()

    source_file = SOURCE_FILES.get(month)
    file_exists = source_file.exists() if source_file else False
    print(f"\n  Source file: {source_file.name if source_file else 'N/A'}")
    print(f"  File exists: {'YES' if file_exists else 'NO'}")

    if month == 4:
        print(f"\n  Template files (T4 preferred):")
        for team, path in TEMPLATE_FILES.items():
            print(f"    {team}: {path.name} -> {'EXISTS' if path.exists() else 'MISSING'}")

    print(f"\n  --- XLSX Dashboard History (month {month} column) ---")
    if file_exists:
        xlsx_dashboard = read_dashboard_cell(source_file, month)
        for team in TEAMS:
            val = xlsx_dashboard.get(team, "(empty)")
            print(f"    {team}: {val}")
    else:
        print("    (file not found)")

    print(f"\n  --- XLSX Team-Level Summary (from source workbook) ---")
    if file_exists:
        for team in TEAMS:
            xlsx_info = read_team_level_from_xlsx(source_file, team, month)
            assessment = xlsx_info.get("monthly_assessment") or "(not found)"
            discipline = xlsx_info.get("discipline_status") or "(not found)"
            summary_cell = xlsx_info.get("summary_cell") or "(not found)"
            print(f"    {team}:")
            print(f"      monthly_assessment = {assessment}")
            print(f"      discipline_status  = {discipline}")
            print(f"      summary_cell       = {summary_cell}")

    if month == 4:
        print(f"\n  --- XLSX Team-Level Summary (from T4 templates) ---")
        for team, tmpl_path in TEMPLATE_FILES.items():
            if not tmpl_path.exists():
                print(f"    {team}: template missing")
                continue
            xlsx_info = read_team_level_from_xlsx(tmpl_path, team, month)
            assessment = xlsx_info.get("monthly_assessment") or "(not found)"
            discipline = xlsx_info.get("discipline_status") or "(not found)"
            summary_cell = xlsx_info.get("summary_cell") or "(not found)"
            print(f"    {team} (from {tmpl_path.name}):")
            print(f"      monthly_assessment = {assessment}")
            print(f"      discipline_status  = {discipline}")
            print(f"      summary_cell       = {summary_cell}")

    print(f"\n  --- DB team_reports (source_type=historical_import) ---")
    db_reports = query_db_team_reports(db, month)
    for team in TEAMS:
        if team in db_reports:
            r = db_reports[team]
            print(f"    {team}:")
            print(f"      assessments        = {r['assessment_count']} KRs")
            print(f"      status_counts      = {r['dashboard_status_counts']}")
            print(f"      monthly_assessment = {r['monthly_assessment']}")
            print(f"      discipline_status  = {r['discipline_status']}")
            print(f"      discipline_desc    = {r['discipline_description']}")
            print(f"      sheet_name         = {r['sheet_name']}")
            print(f"      file_name          = {r['file_name']}")
            print(f"      version            = {r['version']}")
        else:
            print(f"    {team}: NOT FOUND IN DB")

    print(f"\n  --- DB team_monthly_summaries ---")
    db_summaries = query_db_summaries(db, month)
    for team in TEAMS:
        if team in db_summaries:
            s = db_summaries[team]
            print(f"    {team}:")
            print(f"      monthly_assessment = {s['monthly_assessment']}")
            print(f"      discipline_status  = {s['discipline_status']}")
            print(f"      discipline_desc    = {s['discipline_description']}")
        else:
            print(f"    {team}: NOT FOUND IN DB")

    print(f"\n  --- DB historical_snapshots (Dashboard history) ---")
    db_snapshots = query_db_snapshots(db, month)
    for team in TEAMS:
        if team in db_snapshots:
            s = db_snapshots[team]
            print(f"    {team}:")
            print(f"      monthly_assessment = {s['monthly_assessment']}")
            print(f"      source_file        = {s['source_file']}")
            print(f"      source_range       = {s['source_range']}")
        else:
            print(f"    {team}: NOT FOUND IN DB")

    print(f"\n  --- CROSS CHECK: XLSX vs DB ---")
    if file_exists:
        xlsx_dashboard = read_dashboard_cell(source_file, month)
        for team in TEAMS:
            xlsx_val = xlsx_dashboard.get(team, "")
            db_val = db_snapshots.get(team, {}).get("monthly_assessment", "")
            match = "MATCH" if xlsx_val and db_val and xlsx_val == db_val else ("SKIP" if not xlsx_val else "MISMATCH")
            print(f"    {team} Dashboard snapshot: {match}  (xlsx='{xlsx_val}' vs db='{db_val}')")

        for team in TEAMS:
            xlsx_info = read_team_level_from_xlsx(source_file, team, month)
            xlsx_assessment = xlsx_info.get("monthly_assessment") or ""
            db_assessment = db_reports.get(team, {}).get("monthly_assessment") or ""
            normalized_xlsx = normalize_assessment(xlsx_assessment)
            normalized_db = normalize_assessment(db_assessment)
            match = (
                "MATCH"
                if normalized_xlsx and normalized_db and normalized_xlsx == normalized_db
                else ("SKIP" if not xlsx_assessment else "MISMATCH")
            )
            print(f"    {team} team_report monthly: {match}  (xlsx='{xlsx_assessment}' vs db='{db_assessment}')")

    print()


def main():
    print_separator("=")
    print("  HISTORICAL DATA IMPORT VERIFICATION")
    print(f"  Database: {settings.effective_database_url}")
    print(f"  Source dir: {SOURCE_DIR}")
    print(f"  Template dir: {TEMPLATE_DIR}")
    print_separator("=")

    db = create_session()
    try:
        for month in MONTHS:
            print_month_report(month, db)

        print_separator("=")
        print("  OVERALL SUMMARY")
        print_separator("=")
        for month in MONTHS:
            reports = query_db_team_reports(db, month)
            summaries = query_db_summaries(db, month)
            snapshots = query_db_snapshots(db, month)
            teams_with_reports = [t for t in TEAMS if t in reports]
            teams_with_summaries = [t for t in TEAMS if t in summaries]
            teams_with_snapshots = [t for t in TEAMS if t in snapshots]
            print(f"\n  T{month}:")
            print(f"    team_reports:           {len(teams_with_reports)}/4  {teams_with_reports}")
            print(f"    team_monthly_summaries: {len(teams_with_summaries)}/4  {teams_with_summaries}")
            print(f"    dashboard_snapshots:    {len(teams_with_snapshots)}/4  {teams_with_snapshots}")
            if len(teams_with_reports) < 4:
                missing = [t for t in TEAMS if t not in reports]
                print(f"    MISSING reports: {missing}")
            if len(teams_with_summaries) < 4:
                missing = [t for t in TEAMS if t not in summaries]
                print(f"    MISSING summaries: {missing}")
            if len(teams_with_snapshots) < 4:
                missing = [t for t in TEAMS if t not in snapshots]
                print(f"    MISSING snapshots: {missing}")

        print()
        print_separator("=")
        print("  T4 DISCIPLINE OVERRIDES CHECK")
        print_separator("=")
        t4_reports = query_db_team_reports(db, 4)
        t4_summaries = query_db_summaries(db, 4)
        for team in ["TBĐL", "TBCH"]:
            print(f"\n  {team}:")
            r = t4_reports.get(team, {})
            s = t4_summaries.get(team, {})
            print(f"    team_report  -> discipline_status={r.get('discipline_status')}, monthly={r.get('monthly_assessment')}")
            print(f"    summary      -> discipline_status={s.get('discipline_status')}, monthly={s.get('monthly_assessment')}")
            expected_desc = {
                "TBĐL": "Một nhân sự Đội TBĐL không tuân thủ quy định giờ công",
                "TBCH": "Một nhân sự Đội TBCH không tuân thủ đúng HDBD trong quá trình thực hiện công việc bảo dưỡng định kỳ thiết bị Quan trắc",
            }
            actual_desc = r.get("discipline_description") or s.get("discipline_description") or ""
            desc_match = "MATCH" if expected_desc.get(team) == actual_desc else "MISMATCH"
            print(f"    description  -> {desc_match}")
            print(f"      expected: {expected_desc.get(team)}")
            print(f"      actual:   {actual_desc}")

    finally:
        db.close()


if __name__ == "__main__":
    main()
