from pathlib import Path
import re
import unicodedata
from typing import Any

from openpyxl import load_workbook

from app.services.okr.constants import TEAMS
from app.services.okr.extraction import extract_metrics, warning_for_low_confidence, warnings_for_ambiguous_metrics
from app.services.okr.kr_mapping import extract_workshop_kr_code, mapping_by_code
from app.services.okr.rules import expected_status_for_kr, map_to_dashboard_status, normalize_assessment
from app.services.okr.team_normalizer import normalize_team_label


TEAM_ALIASES = {
    "TBHTĐK": ["TBHTĐK", "TBHTDK", "HỆ THỐNG"],
    "TBCH": ["TBCH", "CHẤP HÀNH"],
    "TBĐL": ["TBĐL", "TBDL", "TBĐ", "ĐO LƯỜNG"],
    "TCĐK": ["TCĐK", "TCDK", "TRỰC CA", "TC- ĐK"],
}


TBHTDK_MONTH_COLUMN_GROUPS = {
    1: (13, 14, 15),  # M:N:O
    2: (16, 17, 18),  # P:Q:R
    3: (19, 20, 21),  # S:T:U
    4: (22, 23, 24),  # V:W:X
}

TBCH_MONTH_COLUMN_GROUPS = {
    1: (16, 17, 18),  # P:Q:R
    2: (20, 21, 22),  # T:U:V
    3: (23, 24, 25),  # W:X:Y
    4: (26, 27, 28),  # Z:AA:AB
}


def identify_team(value: str) -> str | None:
    normalized_team, _ = normalize_team_label(value)
    if normalized_team:
        return normalized_team
    normalized = (value or "").upper()
    for team, aliases in TEAM_ALIASES.items():
        if any(alias.upper() in normalized for alias in aliases):
            return team
    return None


def identify_month(value: str) -> int | None:
    match = re.search(r"(?:THÁNG|T)[\s._-]*(1[0-2]|0?[1-9])", value.upper())
    if match:
        return int(match.group(1))
    match = re.search(r"(0?[1-9]|1[0-2])[-_/]20\d{2}", value)
    if match:
        return int(match.group(1))
    return None


def identify_year(value: str) -> int | None:
    match = re.search(r"(20\d{2})", value or "")
    return int(match.group(1)) if match else None


def _parse_int(value: str | None) -> int | None:
    text = str(value or "").strip()
    return int(text) if text.isdigit() else None


def _text(value: Any) -> str:
    return "" if value is None else str(value)


def detect_report_columns(sheet) -> tuple[int, int, int] | None:
    groups = detect_report_column_groups(sheet)
    return groups[0] if groups else None


def _is_assessment_header(value: str) -> bool:
    normalized = _strip_accents(value).lower()
    return "đánh giá" in value or "danh gia" in normalized or "assessment" in normalized


def _is_notes_header(value: str) -> bool:
    normalized = _strip_accents(value).lower()
    return "ghi chú" in value or "ghi chu" in normalized or "notes" in normalized


def detect_report_column_groups(sheet) -> list[tuple[int, int, int]]:
    groups: list[tuple[int, int, int]] = []
    seen: set[tuple[int, int, int]] = set()
    for row in range(1, min(sheet.max_row, 10) + 1):
        values = [str(sheet.cell(row, col).value or "").lower() for col in range(1, sheet.max_column + 1)]
        for idx, value in enumerate(values, start=1):
            if "tình hình thực hiện" not in value:
                continue
            assessment_header = values[idx] if idx < len(values) else ""
            notes_header = values[idx + 1] if idx + 1 < len(values) else ""
            if not _is_assessment_header(assessment_header) or not _is_notes_header(notes_header):
                continue
            group = (idx, idx + 1, idx + 2)
            if group not in seen:
                groups.append(group)
                seen.add(group)
    return groups


def _column_group_exists(sheet, group: tuple[int, int, int]) -> bool:
    report_col, assessment_col, notes_col = group
    if notes_col > sheet.max_column:
        return False
    report_header = ""
    assessment_header = ""
    notes_header = ""
    for row in range(1, min(sheet.max_row, 10) + 1):
        report_header = str(sheet.cell(row, report_col).value or "").lower()
        assessment_header = str(sheet.cell(row, assessment_col).value or "").lower()
        notes_header = str(sheet.cell(row, notes_col).value or "").lower()
        if "tình hình thực hiện" in report_header:
            return _is_assessment_header(assessment_header) and _is_notes_header(notes_header)
    return False


def get_report_columns_for_month(sheet, team: str | None, month: int | None) -> tuple[int, int, int]:
    groups = detect_report_column_groups(sheet)
    if not groups:
        raise ValueError("TEMPLATE_MISMATCH: Cannot identify implementation/assessment/notes columns")
    if month is not None and not 1 <= month <= 12:
        raise ValueError(f"TEMPLATE_MISMATCH: Invalid report month {month}")
    if len(groups) == 1:
        return groups[0]
    if month is None:
        raise ValueError("TEMPLATE_MISMATCH: Reporting month is required for multi-month sheets")

    if team == "TBHTĐK" and month in TBHTDK_MONTH_COLUMN_GROUPS:
        group = TBHTDK_MONTH_COLUMN_GROUPS[month]
        if _column_group_exists(sheet, group):
            return group
    if team == "TBCH":
        group = TBCH_MONTH_COLUMN_GROUPS.get(month)
        if group and _column_group_exists(sheet, group):
            return group

    if 1 <= month <= len(groups):
        return groups[month - 1]
    raise ValueError(
        f"TEMPLATE_MISMATCH: Cannot identify report columns for team={team or 'unknown'} month={month}"
    )


def _has_plan(report_text: str, assessment_text: str, notes: str) -> bool:
    text = f"{report_text} {assessment_text} {notes}".lower()
    no_plan_tokens = ["không có kế hoạch", "khong co ke hoach", "n/a", "không áp dụng", "khong ap dung"]
    return not any(token in text for token in no_plan_tokens)


TEAM_LEVEL_ALIASES = {
    "discipline_status": ["discipline status", "ky luat", "kỷ luật"],
    "discipline_description": ["discipline description", "mo ta ky luat", "mô tả kỷ luật", "mo ta", "mô tả"],
    "related_kr": ["related kr", "kr lien quan", "kr liên quan", "lien quan", "liên quan"],
    "objective_reasons": ["objective reasons", "ly do khach quan", "lý do khách quan"],
    "monthly_assessment": ["monthly assessment", "danh gia chung", "đánh giá chung", "danh gia thang", "đánh giá tháng"],
}

REPORT_METADATA_ALIASES = {
    "team": ["doi to", "đội/tổ", "doi/tổ", "team"],
    "report_month": ["thang bao cao", "tháng báo cáo", "report month"],
    "report_year": ["nam bao cao", "năm báo cáo", "report year"],
}


def _strip_accents(value: str) -> str:
    normalized = unicodedata.normalize("NFD", value)
    stripped = "".join(ch for ch in normalized if unicodedata.category(ch) != "Mn")
    return stripped.replace("đ", "d").replace("Đ", "D")


def _label_key(value: str) -> str | None:
    text = (value or "").strip().lower()
    if not text:
        return None
    label = text.split(":", 1)[0]
    compact = re.sub(r"[\s:._/-]+", " ", _strip_accents(label)).strip()
    matches: list[tuple[int, str]] = []
    for key, aliases in TEAM_LEVEL_ALIASES.items():
        for alias in aliases:
            normalized_alias = re.sub(r"[\s:._/-]+", " ", _strip_accents(alias)).strip()
            if compact == normalized_alias or compact.startswith(f"{normalized_alias} "):
                matches.append((len(alias), key))
    if not matches:
        return None
    return sorted(matches, reverse=True)[0][1]


def _inline_value(raw_value: str) -> str:
    if ":" not in raw_value:
        return ""
    return raw_value.split(":", 1)[1].strip()


def _metadata_key(value: str) -> str | None:
    text = (value or "").strip().lower()
    if not text:
        return None
    label = text.split(":", 1)[0]
    compact = re.sub(r"[\s:._/-]+", " ", _strip_accents(label)).strip()
    for key, aliases in REPORT_METADATA_ALIASES.items():
        for alias in aliases:
            normalized_alias = re.sub(r"[\s:._/-]+", " ", _strip_accents(alias)).strip()
            if compact == normalized_alias or compact.startswith(f"{normalized_alias} "):
                return key
    return None


def _metadata_values(sheet) -> dict[str, str]:
    values: dict[str, str] = {}
    for row in range(1, min(sheet.max_row, 20) + 1):
        for col in range(1, sheet.max_column + 1):
            raw_label = str(sheet.cell(row, col).value or "").strip()
            key = _metadata_key(raw_label)
            if not key or key in values:
                continue
            inline = _inline_value(raw_label)
            if inline:
                values[key] = inline
                continue
            for candidate_col in range(col + 1, min(sheet.max_column, col + 4) + 1):
                raw_value = str(sheet.cell(row, candidate_col).value or "").strip()
                if raw_value and not _metadata_key(raw_value):
                    values[key] = raw_value
                    break
    return values


def _next_value(sheet, row: int, col: int) -> tuple[str, dict | None]:
    inline = _inline_value(str(sheet.cell(row, col).value or ""))
    if inline:
        return inline, {
            "sheet_name": sheet.title,
            "row": row,
            "column": sheet.cell(row, col).column_letter,
            "field_name": "Team_Level",
        }
    for candidate_col in range(col + 1, min(sheet.max_column, col + 4) + 1):
        raw = str(sheet.cell(row, candidate_col).value or "").strip()
        if not raw:
            continue
        if _label_key(raw):
            return "", {
                "sheet_name": sheet.title,
                "row": row,
                "column": sheet.cell(row, candidate_col).column_letter,
                "field_name": "Team_Level",
            }
        return raw, {
            "sheet_name": sheet.title,
            "row": row,
            "column": sheet.cell(row, candidate_col).column_letter,
            "field_name": "Team_Level",
        }
    below = str(sheet.cell(row + 1, col).value or "").strip() if row < sheet.max_row else ""
    if below and not _label_key(below):
        return below, {
            "sheet_name": sheet.title,
            "row": row + 1,
            "column": sheet.cell(row + 1, col).column_letter,
            "field_name": "Team_Level",
        }
    return "", None


def _team_level(sheet, include_warnings: bool = False) -> dict[str, Any] | tuple[dict[str, Any], list[dict[str, Any]]]:
    summary: dict[str, Any] = {}
    source_cells: dict[str, dict[str, Any]] = {}
    warnings: list[dict[str, Any]] = []
    for row in range(1, sheet.max_row + 1):
        for col in range(1, sheet.max_column + 1):
            raw_label = str(sheet.cell(row, col).value or "").strip()
            key = _label_key(raw_label)
            if not key:
                continue
            value, source = _next_value(sheet, row, col)
            if not value and key in {"discipline_status", "monthly_assessment"}:
                warnings.append(
                    _warning(
                        "MISSING_REQUIRED_FIELD",
                        f"Missing team-level value for {key}",
                        "MEDIUM",
                        source
                        or {
                            "sheet_name": sheet.title,
                            "row": row,
                            "column": sheet.cell(row, col).column_letter,
                            "field_name": key,
                        },
                        {"field_name": key},
                    )
                )
            if key == "discipline_status":
                if "nok" in value.lower() or "vi phạm" in value.lower():
                    summary["discipline_status"] = "NOK"
                elif value:
                    summary["discipline_status"] = "OK"
                if key in summary and source:
                    source_cells[key] = source
            elif key == "monthly_assessment":
                if value:
                    summary["monthly_assessment"] = normalize_assessment(value) or value
                    if source:
                        source_cells[key] = source
            elif key == "discipline_description" and value and "discipline_description" not in summary:
                summary["discipline_description"] = value
                if source:
                    source_cells[key] = source
            elif key == "related_kr" and value:
                summary["related_kr"] = value
                if source:
                    source_cells[key] = source
            elif key == "objective_reasons" and value:
                summary["objective_reasons"] = value
                if source:
                    source_cells[key] = source
    if source_cells:
        summary["source_cells"] = source_cells
    if include_warnings:
        return summary, warnings
    return summary


def _normalize_lookup_text(value: str) -> str:
    text = _strip_accents(str(value or "")).lower()
    return re.sub(r"[^a-z0-9]+", " ", text).strip()


def _mapping_by_name(master: dict[str, Any]) -> dict[str, str]:
    return {_normalize_lookup_text(record.kr_name): code for code, record in master.items() if record.kr_name}


def _warning(warning_type: str, reason: str, severity: str = "MEDIUM", source_cell: dict | None = None, extracted_value: Any = None) -> dict:
    return {
        "warning_type": warning_type,
        "severity": severity,
        "source_cell": source_cell,
        "extracted_value": extracted_value,
        "reason": reason,
        "admin_action": "PENDING",
    }


def _cell_reference(sheet, row: int, col: int, field_name: str) -> dict[str, Any]:
    return {
        "sheet_name": sheet.title,
        "row": row,
        "column": sheet.cell(row, col).column_letter,
        "field_name": field_name,
    }


def _first_value_to_right(sheet, row: int, start_col: int) -> tuple[str, dict[str, Any] | None]:
    for col in range(start_col + 1, sheet.max_column + 1):
        value = str(sheet.cell(row, col).value or "").strip()
        if value:
            return value, _cell_reference(sheet, row, col, "Monthly_Assessment")
    return "", None


def _team_summary_from_known_layout(
    sheet,
    team: str | None,
    month: int | None,
    report_cols: tuple[int, int, int],
) -> tuple[str | None, dict[str, Any] | None]:
    report_col, assessment_col, _notes_col = report_cols
    candidates: list[tuple[int, int]] = []
    if team == "TBHTĐK":
        candidates.append((39, assessment_col))
    elif team == "TBCH":
        candidates.append((44, report_col))
        candidates.append((44, assessment_col))
    elif team == "TCĐK":
        candidates.append((42, assessment_col))
    elif team == "TBĐL":
        for row in range(1, min(sheet.max_row, 60) + 1):
            for col in range(1, min(sheet.max_column, 16) + 1):
                label = str(sheet.cell(row, col).value or "").strip().lower()
                if "kết quả đánh giá" in label or "ket qua danh gia" in _strip_accents(label):
                    value, source = _first_value_to_right(sheet, row, col)
                    if value:
                        return normalize_assessment(value) or value, source
        candidates.append((39, report_col))
        candidates.append((39, assessment_col))

    for row, col in candidates:
        if row <= sheet.max_row and col <= sheet.max_column:
            value = str(sheet.cell(row, col).value or "").strip()
            if value:
                return normalize_assessment(value) or value, _cell_reference(
                    sheet, row, col, "Monthly_Assessment"
                )
    return None, None


def _infer_monthly_assessment_from_krs(assessments: list[dict[str, Any]]) -> str | None:
    statuses = {item.get("dashboard_status") for item in assessments}
    if "NG" in statuses:
        return "Không hoàn thành"
    if "GOOD" in statuses:
        return "Hoàn thành tốt"
    if "OK" in statuses:
        return "Hoàn thành"
    return None


def parse_team_report(
    path: Path,
    team: str | None = None,
    month: int | None = None,
    year: int | None = None,
    kr_mapping: dict[str, Any] | None = None,
) -> dict[str, Any]:
    workbook = load_workbook(path, read_only=False, data_only=False, keep_links=False)
    selected_sheet = None
    selected_team = team
    for sheet_name in workbook.sheetnames:
        detected = identify_team(sheet_name)
        if team and detected == team:
            selected_sheet = workbook[sheet_name]
            selected_team = team
            break
        if not selected_sheet and detected in TEAMS:
            selected_sheet = workbook[sheet_name]
            selected_team = detected
    if selected_sheet is None:
        selected_sheet = workbook.active
        selected_team = team or selected_team
    metadata = _metadata_values(selected_sheet)
    selected_team = team or identify_team(metadata.get("team", "")) or selected_team or identify_team(path.name)
    report_month = month or identify_month(metadata.get("report_month", "")) or _parse_int(metadata.get("report_month")) or identify_month(path.name) or identify_month(str(selected_sheet.cell(1, 1).value or ""))
    report_year = year or identify_year(metadata.get("report_year", "")) or _parse_int(metadata.get("report_year")) or identify_year(path.name) or identify_year(str(selected_sheet.cell(1, 1).value or "")) or 2026

    report_cols = get_report_columns_for_month(selected_sheet, selected_team, report_month)
    report_col, assessment_col, notes_col = report_cols
    master = kr_mapping or mapping_by_code()
    master_by_name = _mapping_by_name(master)
    assessments: list[dict[str, Any]] = []
    warnings: list[dict[str, Any]] = []

    for row in range(1, selected_sheet.max_row + 1):
        raw_code = _text(selected_sheet.cell(row, 2).value or selected_sheet.cell(row, 5).value)
        name_code = master_by_name.get(_normalize_lookup_text(_text(selected_sheet.cell(row, 3).value)))
        workshop_code = extract_workshop_kr_code(raw_code)
        if name_code and (not workshop_code or workshop_code != name_code):
            workshop_code = name_code
        if not workshop_code or workshop_code not in master:
            continue
        report_text = _text(selected_sheet.cell(row, report_col).value)
        raw_assessment = _text(selected_sheet.cell(row, assessment_col).value)
        assessment = normalize_assessment(raw_assessment)
        notes = _text(selected_sheet.cell(row, notes_col).value)
        has_plan = _has_plan(report_text, raw_assessment, notes)
        metrics = extract_metrics(report_text, workshop_code)
        source_cells = {
            "implementation_report": _cell_reference(
                selected_sheet, row, report_col, "Implementation_Report"
            ),
            "team_self_assessment": _cell_reference(
                selected_sheet, row, assessment_col, "KR_Assessment"
            ),
            "notes": _cell_reference(selected_sheet, row, notes_col, "Notes"),
        }
        source_cell = source_cells["implementation_report"]
        for metric in metrics:
            warning = warning_for_low_confidence(metric, source_cell)
            if warning:
                warnings.append(warning)
        warnings.extend(warnings_for_ambiguous_metrics(metrics, source_cell))
        if not assessment:
            warnings.append(
                _warning(
                    "MISSING_REQUIRED_FIELD",
                    f"Missing KR_Assessment for {workshop_code}",
                    "HIGH",
                    {
                        "sheet_name": selected_sheet.title,
                        "row": row,
                        "column": selected_sheet.cell(row, assessment_col).column_letter,
                        "field_name": "KR_Assessment",
                    },
                    {"workshop_kr_code": workshop_code},
                )
            )
        metric_dicts = [metric.to_dict() for metric in metrics]
        dashboard_status = map_to_dashboard_status(assessment, has_plan=has_plan)
        expected = expected_status_for_kr(workshop_code, metric_dicts, notes)
        if expected and expected != "#N/A" and dashboard_status != expected:
            warnings.append(
                _warning(
                    "ASSESSMENT_MISMATCH",
                    f"Numerical data suggests {expected} but team self-assessment maps to {dashboard_status}",
                    "MEDIUM",
                    source_cell,
                    {
                        "workshop_kr_code": workshop_code,
                        "expected_status": expected,
                        "dashboard_status": dashboard_status,
                        "objective_reason": notes,
                    },
                )
            )
        assessments.append(
            {
                "workshop_kr_code": workshop_code,
                "kr_name": master[workshop_code].kr_name,
                "team_self_assessment": assessment,
                "dashboard_status": dashboard_status,
                "has_plan": has_plan,
                "implementation_report": report_text,
                "notes": notes,
                "source_cell": source_cell,
                "source_cells": source_cells,
                "metrics": metric_dicts,
            }
        )

    if selected_team is None:
        warnings.append(_warning("TEAM_MONTH_IDENTIFICATION_FAILED", "Cannot automatically identify team", extracted_value={"team": selected_team, "month": report_month}))
    if report_month is None:
        warnings.append(_warning("TEAM_MONTH_IDENTIFICATION_FAILED", "Cannot automatically identify reporting month", extracted_value={"team": selected_team, "month": report_month}))
    found_codes = {assessment["workshop_kr_code"] for assessment in assessments}
    missing_codes = sorted(set(master) - found_codes)
    if missing_codes:
        warnings.append(
            _warning(
                "TEMPLATE_VALIDATION_ERROR",
                f"Template contains {len(found_codes)} of 37 standard KR rows; missing: {', '.join(missing_codes[:12])}{'...' if len(missing_codes) > 12 else ''}",
                "HIGH",
                extracted_value={"missing_kr_codes": missing_codes},
            )
        )

    team_level, team_level_warnings = _team_level(selected_sheet, include_warnings=True)
    warnings.extend(team_level_warnings)
    layout_assessment, layout_source = _team_summary_from_known_layout(
        selected_sheet, selected_team, report_month, report_cols
    )
    if layout_assessment:
        team_level["monthly_assessment"] = layout_assessment
        team_level.setdefault("source_cells", {})["monthly_assessment"] = layout_source
    elif "monthly_assessment" not in team_level:
        inferred = _infer_monthly_assessment_from_krs(assessments)
        if inferred:
            team_level["monthly_assessment"] = inferred
            team_level.setdefault("source_cells", {})["monthly_assessment"] = {
                "sheet_name": selected_sheet.title,
                "row": None,
                "column": None,
                "field_name": "Monthly_Assessment",
                "inferred_from": "KR dashboard statuses",
            }

    result = {
        "team": selected_team,
        "report_month": report_month,
        "report_year": report_year,
        "sheet_name": selected_sheet.title,
        "assessments": assessments,
        "team_level": team_level,
        "warnings": warnings,
        "source_cell_references": [
            source_cell
            for assessment in assessments
            for source_cell in assessment.get("source_cells", {}).values()
        ]
        + list(team_level.get("source_cells", {}).values()),
    }
    workbook.close()
    return result
