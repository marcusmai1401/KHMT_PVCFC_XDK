from copy import copy
import hashlib
from pathlib import Path
from shutil import copyfile
from typing import Any

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.styles import Font
from openpyxl.styles import Alignment
from openpyxl.worksheet.datavalidation import DataValidation, DataValidationList

from app.core.config import settings
from app.models.domain import TeamReportModel
from app.services.okr.kr_mapping import extract_workshop_kr_code, load_master_kr_mapping
from app.services.okr.workbook_cleaner import strip_workbook_external_state


TEMPLATE_FILENAME = "okr-team-report-template.xlsx"
TEMPLATE_SHEET_NAME = "Team_Report"
BASE_SHEET_NAME = "TBĐL"
APPROVED_TEMPLATE_FILENAME = "TBĐL.xlsx"
TEAM_OPTIONS = ["TBHTĐK", "TBCH", "TBĐL", "TCĐK"]
ASSESSMENT_OPTIONS = ["Hoàn thành xuất sắc", "Hoàn thành tốt", "Hoàn thành", "Không hoàn thành", "N/A"]
MAIN_HEADERS = [
    "STT",
    "Mã mục tiêu/Kết quả then chốt Đội/Tổ",
    "Tên mục tiêu / Kết quả then chốt (Các kết quả then chốt (KR) phải đo lường được)",
    "Đo lường",
    "Mục đích",
    "Tần suất đo lường",
    "Tỷ trọng mục tiêu",
    "Kế hoạch hành động",
    "Ngày bắt đầu thực hiện",
    "Ngày hoàn thành dự kiến",
    "Ngân sách thực hiện (tỷ đồng)",
    "Người thực hiện/Đầu mối",
    "Kiểm tra đánh giá",
]
REPORT_HEADERS = ["Tình hình thực hiện", "Đánh giá", "Ghi chú"]


def _mapping_to_dict(mapping: Any) -> dict[str, Any]:
    if hasattr(mapping, "to_dict"):
        return mapping.to_dict()
    return {
        "workshop_kr_code": mapping.get("workshop_kr_code", ""),
        "kr_name": mapping.get("kr_name", ""),
        "measurement_type": mapping.get("measurement_type", ""),
        "target_value": mapping.get("target_value", ""),
    }


def _kr_sort_key(mapping: dict[str, Any]) -> tuple[int, int, str]:
    code = str(mapping.get("workshop_kr_code") or "")
    objective, _, kr = code.partition(".KR")
    if not objective.startswith("O") or not kr:
        return (10_000, 10_000, code)
    try:
        return (int(objective[1:]), int(kr), code)
    except ValueError:
        return (10_000, 10_000, code)


def _quoted_list(values: list[str]) -> str:
    return '"' + ",".join(values) + '"'


def _copy_style(source: Cell, target: Cell) -> None:
    if source.has_style:
        target.font = copy(source.font)
        target.fill = copy(source.fill)
        target.border = copy(source.border)
        target.alignment = copy(source.alignment)
        target.number_format = source.number_format
        target.protection = copy(source.protection)


def _style_range(sheet, cell_range: str, source: Cell) -> None:
    if ":" not in cell_range:
        _copy_style(source, sheet[cell_range])
        return
    for row in sheet[cell_range]:
        for cell in row:
            _copy_style(source, cell)


def _clear_sheet(sheet, max_row: int = 44, max_col: int = 16) -> None:
    for merged_range in list(sheet.merged_cells.ranges):
        sheet.unmerge_cells(str(merged_range))
    for row in sheet.iter_rows(min_row=1, max_row=max(max_row, sheet.max_row), min_col=1, max_col=max_col):
        for cell in row:
            cell.value = None
            cell.comment = None
            cell.hyperlink = None
    sheet.data_validations = DataValidationList()
    sheet.auto_filter.ref = None


def _add_validation(sheet, validation: DataValidation, ranges: list[str]) -> None:
    sheet.add_data_validation(validation)
    for cell_range in ranges:
        validation.add(cell_range)


def _set_title(sheet) -> None:
    title_style = sheet["A1"]
    sheet.merge_cells("A1:P1")
    sheet["A1"] = "TEMPLATE BÁO CÁO KẾ HOẠCH MỤC TIÊU XƯỞNG ĐIỀU KHIỂN"
    _style_range(sheet, "A1:P1", title_style)
    sheet["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _set_metadata(sheet) -> None:
    metadata = [
        ("A2", "Đội/Tổ", "B2:C2", ""),
        ("D2", "Tháng báo cáo", "E2", 4),
        ("F2", "Năm báo cáo", "G2", 2026),
        ("H2", "Kỷ luật", "I2", "OK"),
        ("J2", "Đánh giá chung", "K2:P2", "Hoàn thành"),
        ("A3", "Mô tả kỷ luật", "B3:F3", ""),
        ("G3", "KR liên quan", "H3:I3", ""),
        ("J3", "Lý do khách quan", "K3:P3", ""),
    ]
    label_style = sheet["A2"]
    input_style = sheet["A4"]
    for label_cell, label, input_range, value in metadata:
        sheet[label_cell] = label
        _copy_style(label_style, sheet[label_cell])
        sheet[label_cell].alignment = copy(label_style.alignment)
        if ":" in input_range:
            sheet.merge_cells(input_range)
            first_cell = input_range.split(":", 1)[0]
        else:
            first_cell = input_range
        sheet[first_cell] = value
        _style_range(sheet, input_range, input_style)
        sheet[first_cell].alignment = Alignment(vertical="center", wrap_text=True)


def _set_headers(sheet) -> None:
    main_header_style = sheet["A2"]
    month_header_style = sheet["N2"]
    report_header_style = sheet["N3"]
    for column_index, header in enumerate(MAIN_HEADERS, start=1):
        cell = sheet.cell(5, column_index, header)
        sheet.merge_cells(start_row=5, start_column=column_index, end_row=6, end_column=column_index)
        _style_range(sheet, f"{cell.coordinate}:{sheet.cell(6, column_index).coordinate}", main_header_style)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    sheet.merge_cells("N5:P5")
    sheet["N5"] = '="Báo cáo tình hình thực hiện tháng "&$E$2&"/"&$G$2'
    _style_range(sheet, "N5:P5", month_header_style)
    sheet["N5"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for offset, header in enumerate(REPORT_HEADERS, start=14):
        cell = sheet.cell(6, offset, header)
        _copy_style(report_header_style, cell)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _set_data_rows(sheet, records: list[dict[str, Any]]) -> None:
    data_style = sheet["A4"]
    report_style = sheet["N4"]
    first_data_row = 7
    for offset, record in enumerate(records):
        row = first_data_row + offset
        code = record["workshop_kr_code"]
        values = [
            offset + 1,
            f"ĐK.{code}",
            record["kr_name"],
            record["measurement_type"],
            record["target_value"],
            "Tháng",
            "",
            "",
            "01/01/2026",
            "31/12/2026",
            "",
            "",
            "",
            "",
            "",
            "",
        ]
        for column_index, value in enumerate(values, start=1):
            cell = sheet.cell(row, column_index, value)
            _copy_style(report_style if column_index >= 14 else data_style, cell)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
        source_row = min(4 + offset, 38)
        sheet.row_dimensions[row].height = sheet.row_dimensions[source_row].height or 80


def _set_summary_row(sheet, row: int = 44) -> None:
    result_label_style = sheet["B39"]
    result_value_style = sheet["N39"]
    sheet[f"A{row}"] = "TỔNG"
    _copy_style(sheet["A39"], sheet[f"A{row}"])
    sheet[f"G{row}"] = "=SUM(G7:G43)"
    _copy_style(sheet["G39"], sheet[f"G{row}"])
    sheet.merge_cells(start_row=row, start_column=2, end_row=row, end_column=13)
    sheet[f"B{row}"] = "KẾT QUẢ ĐÁNH GIÁ"
    _style_range(sheet, f"B{row}:M{row}", result_label_style)
    sheet.merge_cells(start_row=row, start_column=14, end_row=row, end_column=16)
    sheet[f"N{row}"] = "=$K$2"
    _style_range(sheet, f"N{row}:P{row}", result_value_style)
    sheet.row_dimensions[row].height = sheet.row_dimensions[39].height or 36


def _finish_sheet(sheet, last_data_row: int) -> None:
    sheet.freeze_panes = "B7"
    sheet.auto_filter.ref = None
    sheet.sheet_view.showGridLines = False
    sheet.data_validations = DataValidationList()
    _add_validation(sheet, DataValidation(type="list", formula1=_quoted_list(TEAM_OPTIONS), allow_blank=False), ["B2"])
    _add_validation(sheet, DataValidation(type="whole", operator="between", formula1="1", formula2="12", allow_blank=False), ["E2"])
    _add_validation(sheet, DataValidation(type="whole", operator="between", formula1="2026", formula2="2035", allow_blank=False), ["G2"])
    _add_validation(sheet, DataValidation(type="list", formula1='"OK,NOK"', allow_blank=False), ["I2"])
    _add_validation(
        sheet,
        DataValidation(type="list", formula1=_quoted_list(ASSESSMENT_OPTIONS), allow_blank=False),
        ["K2", f"O7:O{last_data_row}"],
    )
    sheet.protection.sheet = False


def _load_base_workbook(target: Path):
    source = next(
        (
            candidate
            for candidate in [
                settings.source_okr_workbook,
                settings.workspace_dir / "template_xlsx" / APPROVED_TEMPLATE_FILENAME,
                settings.workspace_dir / "KHMT_Monthly" / "OKR tháng 04-2026 - X.ĐK.xlsx",
            ]
            if candidate.exists()
        ),
        settings.source_okr_workbook,
    )
    if source.exists():
        workbook = load_workbook(source, keep_links=False)
        strip_workbook_external_state(workbook)
        if BASE_SHEET_NAME not in workbook.sheetnames:
            raise ValueError(f"Source workbook does not contain sheet {BASE_SHEET_NAME}")
        sheet = workbook[BASE_SHEET_NAME]
        for other_sheet in list(workbook.worksheets):
            if other_sheet is not sheet:
                workbook.remove(other_sheet)
        sheet.title = TEMPLATE_SHEET_NAME
        return workbook, sheet
    if target.exists():
        workbook = load_workbook(target, keep_links=False)
        strip_workbook_external_state(workbook)
        return workbook, workbook[TEMPLATE_SHEET_NAME]
    raise FileNotFoundError(f"Source workbook not found: {source}")


def standard_report_template_path() -> Path:
    return settings.storage_dir / "templates" / TEMPLATE_FILENAME


def approved_report_template_path() -> Path:
    return settings.workspace_dir / APPROVED_TEMPLATE_FILENAME


def generate_standard_report_template(master: list[Any] | None = None, output_path: Path | None = None) -> Path:
    approved_template = approved_report_template_path()
    target = output_path or standard_report_template_path()
    target.parent.mkdir(parents=True, exist_ok=True)
    if approved_template.exists() and approved_template.resolve() != target.resolve():
        copyfile(approved_template, target)
        return target

    records = sorted(
        (_mapping_to_dict(record) for record in (master or load_master_kr_mapping())),
        key=_kr_sort_key,
    )
    workbook, sheet = _load_base_workbook(target)
    _clear_sheet(sheet)
    _set_title(sheet)
    _set_metadata(sheet)
    _set_headers(sheet)
    _set_data_rows(sheet, records)
    last_data_row = 6 + len(records)
    _set_summary_row(sheet, row=last_data_row + 1)
    _finish_sheet(sheet, last_data_row)
    workbook.properties.creator = "OKR Automation System"
    workbook.properties.lastModifiedBy = "OKR Automation System"
    strip_workbook_external_state(workbook)
    workbook.save(target)
    return target


def _team_template_path(team: str | None) -> Path:
    if not team:
        raise ValueError("Team is required for web input Excel generation")
    return settings.workspace_dir / "template_xlsx" / f"{team}.xlsx"


def _detect_report_columns(sheet) -> tuple[int, int, int]:
    for row in range(1, min(sheet.max_row, 10) + 1):
        for col in range(1, sheet.max_column + 1):
            text = str(sheet.cell(row, col).value or "").lower()
            if "tình hình thực hiện" in text:
                return col, col + 1, col + 2
    return 14, 15, 16


def _template_kr_rows(sheet) -> dict[str, int]:
    rows: dict[str, int] = {}
    for row in range(1, sheet.max_row + 1):
        for col in range(1, min(sheet.max_column, 6) + 1):
            code = extract_workshop_kr_code(str(sheet.cell(row, col).value or ""))
            if code and code not in rows:
                rows[code] = row
                break
    return rows


def _copy_row_style(sheet, source_row: int, target_row: int, max_col: int) -> None:
    for col in range(1, max_col + 1):
        source = sheet.cell(source_row, col)
        target = sheet.cell(target_row, col)
        _copy_style(source, target)
        target.alignment = copy(source.alignment)
    sheet.row_dimensions[target_row].height = sheet.row_dimensions[source_row].height


def web_input_excel_path(report: TeamReportModel) -> Path:
    settings.storage_dir.joinpath("exports").mkdir(parents=True, exist_ok=True)
    suffix = "-draft" if report.report_status == "draft" else (f"-v{report.version}" if report.version > 1 else "")
    return (
        settings.storage_dir
        / "exports"
        / f"bao-cao-okr-{report.team}-T{report.report_month}-{report.report_year}{suffix}.xlsx"
    )


def generate_excel_from_web_input(report: TeamReportModel, output_path: Path | None = None) -> Path:
    template = _team_template_path(report.team)
    if not template.exists():
        raise FileNotFoundError(f"Team template not found: {template}")

    workbook = load_workbook(template, keep_links=False)
    strip_workbook_external_state(workbook)
    sheet = workbook[report.team] if report.team in workbook.sheetnames else workbook.active
    report_col, assessment_col, notes_col = _detect_report_columns(sheet)
    kr_rows = _template_kr_rows(sheet)

    title = str(sheet["A1"].value or "")
    if title:
        sheet["A1"] = f"{title.split('_THÁNG', 1)[0]}_THÁNG {report.report_month}/{report.report_year}"

    by_code = {item.get("workshop_kr_code"): item for item in report.assessments or []}
    for kr_code, row in kr_rows.items():
        item = by_code.get(kr_code) or {}
        sheet.cell(row, report_col).value = item.get("implementation_report") or ""
        sheet.cell(row, assessment_col).value = item.get("team_self_assessment") or ""
        sheet.cell(row, notes_col).value = item.get("notes") or ""

    last_kr_row = max(kr_rows.values()) if kr_rows else sheet.max_row
    section_row = last_kr_row + 2
    source_style_row = last_kr_row if last_kr_row else 1
    max_col = max(sheet.max_column, notes_col)
    for offset in range(0, 8 + len(report.arising_work or [])):
        _copy_row_style(sheet, source_style_row, section_row + offset, max_col)

    team_level = report.team_level or {}
    sheet.cell(section_row, 1).value = "THÔNG TIN BÁO CÁO"
    sheet.cell(section_row, 1).font = Font(bold=True)
    metadata_rows = [
        ("Đội/Tổ", report.team),
        ("Tháng báo cáo", report.report_month),
        ("Năm báo cáo", report.report_year),
        ("Kỷ luật", team_level.get("discipline_status") or "OK"),
        ("Đánh giá chung", team_level.get("monthly_assessment") or "Hoàn thành nhiệm vụ"),
        ("Mô tả kỷ luật", team_level.get("discipline_description") or ""),
        ("Mô tả chi tiết", team_level.get("detailed_description") or ""),
    ]
    for index, (label, value) in enumerate(metadata_rows, start=1):
        row = section_row + index
        sheet.cell(row, 1).value = label
        sheet.cell(row, 2).value = value

    arising_work = report.arising_work or []
    if arising_work:
        start = section_row + len(metadata_rows) + 2
        sheet.cell(start, 1).value = "CÔNG VIỆC PHÁT SINH"
        sheet.cell(start, 1).font = Font(bold=True)
        for index, item in enumerate(arising_work, start=1):
            row = start + index
            sheet.cell(row, 1).value = index
            sheet.cell(row, 2).value = item.get("content") or ""
            sheet.cell(row, assessment_col).value = item.get("status") or "Hoàn thành"

    target = output_path or web_input_excel_path(report)
    target.parent.mkdir(parents=True, exist_ok=True)
    workbook.properties.creator = "OKR Automation System"
    workbook.properties.lastModifiedBy = "OKR Automation System"
    strip_workbook_external_state(workbook)
    workbook.save(target)
    return target


def hash_file(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()
