from io import BytesIO

from openpyxl import Workbook, load_workbook
from sqlalchemy import func, select

from app.models.et_domain import CompetencyFramework, CompetencyItem
from app.services.pvcfc_knl_seed import seed_pvcfc_knl_frameworks


PVCFC_KNL_CODES = {"KNL_\u0110K_10", "KNL_\u0110K_12", "KNL_\u0110K_13", "KNL_\u0110K_14", "KNL_\u0110K_15"}
IMPORT_FRAMEWORK_CODES = ["KNL_\u0110K_10", "KNL_\u0110K_12", "KNL_\u0110K_13", "KNL_\u0110K_14", "KNL_\u0110K_15"]


def _login(client, user_id: str, password: str) -> dict[str, str]:
    response = client.post("/api/v1/auth/login", json={"user_id": user_id, "password": password})
    assert response.status_code == 200
    return {"Authorization": f"Bearer {response.json()['access_token']}"}


def _framework_workbook() -> BytesIO:
    workbook = Workbook()
    catalog = workbook.active
    catalog.title = "Ma trận năng lực"
    catalog.append(["Phân nhóm", "Tên năng lực", "STT", "Mã NLCM", "Chi tiết"])
    catalog.append([])
    catalog.append([])
    catalog.append(["Cơ bản", "HSE cương vị", 1, "ĐK_NLCM_001", "Kiến thức an toàn cương vị"])

    for code in IMPORT_FRAMEWORK_CODES:
        sheet = workbook.create_sheet(code)
        sheet["A1"] = "KHUNG NĂNG LỰC CHUYÊN MÔN"
        sheet["A2"] = f"Khung import test {code}"
        sheet.append([])
        sheet.append([])
        sheet.append(["Phân nhóm", "Tên năng lực", "STT", "Mã NLCM", "Chi tiết", 1, 2, 3, 4, 5, 6, 7, 8])
        sheet.append(["Cơ bản", "HSE cương vị", 1, "ĐK_NLCM_001", "Kiến thức an toàn cương vị", 1, 1, 1, 1, 1, 1, 1, 1])

    payload = BytesIO()
    workbook.save(payload)
    payload.seek(0)
    return payload


def _import_frameworks(client, headers):
    response = client.post(
        "/api/v1/et/frameworks/import",
        headers=headers,
        files={
            "file": (
                "knl.xlsx",
                _framework_workbook(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert response.status_code == 200, response.text
    assert len(response.json()["created"]) == 5
    return response.json()["created"]


def test_pvcfc_knl_seeded_framework_detail(client, admin_headers):
    response = client.get("/api/v1/et/frameworks", headers=admin_headers)
    assert response.status_code == 200
    frameworks = response.json()
    assert PVCFC_KNL_CODES.issubset({framework["code"] for framework in frameworks})

    framework = next(
        row for row in frameworks
        if row["code"] == "KNL_\u0110K_14" and row["is_active"] is True
    )
    detail_response = client.get(f"/api/v1/et/frameworks/{framework['id']}", headers=admin_headers)
    assert detail_response.status_code == 200
    detail = detail_response.json()
    assert len(detail["items"]) == 37
    assert [detail["level_sums"][str(level)] for level in range(1, 9)] == [40, 68, 75, 85, 91, 97, 102, 104]

    first_item = next(item for item in detail["items"] if item["nlcm_code"] == "\u0110K_NLCM_001")
    assert first_item["definition"]
    assert first_item["requirements_text"]


def test_pvcfc_knl_seed_is_idempotent(db_session):
    before_frameworks = (
        db_session.scalar(select(func.count()).select_from(CompetencyFramework).where(CompetencyFramework.code.in_(PVCFC_KNL_CODES)))
        or 0
    )
    before_items = (
        db_session.scalar(
            select(func.count())
            .select_from(CompetencyItem)
            .join(CompetencyFramework, CompetencyFramework.id == CompetencyItem.framework_id)
            .where(CompetencyFramework.code.in_(PVCFC_KNL_CODES))
        )
        or 0
    )

    first = seed_pvcfc_knl_frameworks(db_session, actor_id="admin")
    second = seed_pvcfc_knl_frameworks(db_session, actor_id="admin")
    db_session.flush()

    after_frameworks = (
        db_session.scalar(select(func.count()).select_from(CompetencyFramework).where(CompetencyFramework.code.in_(PVCFC_KNL_CODES)))
        or 0
    )
    after_items = (
        db_session.scalar(
            select(func.count())
            .select_from(CompetencyItem)
            .join(CompetencyFramework, CompetencyFramework.id == CompetencyItem.framework_id)
            .where(CompetencyFramework.code.in_(PVCFC_KNL_CODES))
        )
        or 0
    )
    assert first["skipped"] is True
    assert second["skipped"] is True
    assert after_frameworks == before_frameworks
    assert after_items == before_items


def test_framework_import_and_assessment_dashboard_flow(client, admin_headers):
    frameworks = _import_frameworks(client, admin_headers)
    framework_export = client.get(f"/api/v1/et/frameworks/{frameworks[0]['id']}/export", headers=admin_headers)
    assert framework_export.status_code == 200
    framework_workbook = load_workbook(BytesIO(framework_export.content), data_only=False)
    framework_sheet = framework_workbook.active
    assert framework_sheet["A1"].value == "KHUNG NĂNG LỰC CHUYÊN MÔN"
    assert framework_sheet["A2"].value == frameworks[0]["title"]
    assert framework_sheet["A3"].value == "Phân nhóm"
    assert framework_sheet["F4"].value.startswith("=SUM(F6:F")
    assert framework_sheet["F5"].value == 1
    assert framework_sheet["D6"].value == "ĐK_NLCM_001"
    personnel_response = client.post(
        "/api/v1/et/personnel",
        headers=admin_headers,
        json={
            "employee_code": "ET001",
            "full_name": "Nguyễn Văn ET",
            "position_code": "KNL_ĐK_14",
            "team": "TBHTĐK",
            "current_level": 2,
            "status": "active",
            "user_id": "TBHTĐK",
        },
    )
    assert personnel_response.status_code == 200, personnel_response.text
    personnel_id = personnel_response.json()["id"]

    assessment_response = client.post(
        "/api/v1/et/assessments",
        headers=admin_headers,
        json={"personnel_id": personnel_id, "assessment_period": "2026-Q2"},
    )
    assert assessment_response.status_code == 200, assessment_response.text
    assessment = assessment_response.json()
    assert assessment["status"] == "draft"
    assert assessment["personnel_level_at_assessment"] == 2
    assert assessment["items"]
    assert all(item["gap"] is None for item in assessment["items"])

    update_response = client.put(
        f"/api/v1/et/assessments/{assessment['id']}",
        headers=admin_headers,
        json={
            "items": [
                {"id": item["id"], "actual_score": item["required_score"]}
                for item in assessment["items"]
                if item["required_score"] <= 5
            ],
            "notes": "Đạt yêu cầu",
        },
    )
    assert update_response.status_code == 200, update_response.text
    assert update_response.json()["overall_result"] == "Đạt"

    submit_response = client.post(f"/api/v1/et/assessments/{assessment['id']}/submit", headers=admin_headers)
    assert submit_response.status_code == 200, submit_response.text
    submitted = submit_response.json()
    assert submitted["status"] == "submitted"
    assert submitted["is_latest"] is True

    dashboard_response = client.get("/api/v1/et/dashboard", headers=admin_headers)
    assert dashboard_response.status_code == 200
    dashboard = dashboard_response.json()
    assert dashboard["aggregate"]["total_active_personnel"] == 1
    assert dashboard["aggregate"]["pass_count"] == 1
    assert dashboard["rows"][0]["achieved_count"] > 0

    heatmap_response = client.get("/api/v1/et/dashboard/heatmap", headers=admin_headers)
    assert heatmap_response.status_code == 200
    assert heatmap_response.json()["rows"]

    export_response = client.get(f"/api/v1/et/assessments/{assessment['id']}/export", headers=admin_headers)
    assert export_response.status_code == 200
    assert export_response.content.startswith(b"PK")

    plan_response = client.post(
        "/api/v1/et/learning-plans",
        headers=admin_headers,
        json={
            "personnel_id": personnel_id,
            "title": "Kế hoạch test",
            "start_date": "2026-04-15",
            "duration_months": 2,
            "items": [{"item_id": assessment["items"][0]["item_id"], "target_week": 1, "target_level": 2}],
        },
    )
    assert plan_response.status_code == 200, plan_response.text
    plan_export = client.get(f"/api/v1/et/learning-plans/{plan_response.json()['id']}/export", headers=admin_headers)
    assert plan_export.status_code == 200
    plan_workbook = load_workbook(BytesIO(plan_export.content), data_only=True)
    plan_sheet = plan_workbook["Project Timeline"]
    assert plan_sheet["B1"].value == "KẾ HOẠCH HỌC TẬP NHÂN SỰ MỚI"
    assert plan_sheet["E2"].value == 2026
    assert plan_sheet["E3"].value == "Q2"
    assert plan_sheet["E4"].value == "Tháng 1"
    assert plan_sheet["E5"].value == 15
    assert plan_sheet["E6"].value == 1


def test_team_account_can_only_read_linked_personnel_assessment(client, admin_headers):
    _import_frameworks(client, admin_headers)
    own = client.post(
        "/api/v1/et/personnel",
        headers=admin_headers,
        json={
            "employee_code": "OWN",
            "full_name": "Own User",
            "position_code": "KNL_ĐK_13",
            "team": "TBĐL",
            "current_level": 1,
            "status": "active",
            "user_id": "TBĐL",
        },
    ).json()
    other = client.post(
        "/api/v1/et/personnel",
        headers=admin_headers,
        json={
            "employee_code": "OTHER",
            "full_name": "Other User",
            "position_code": "KNL_ĐK_14",
            "team": "TBHTĐK",
            "current_level": 1,
            "status": "active",
        },
    ).json()
    own_assessment = client.post(
        "/api/v1/et/assessments",
        headers=admin_headers,
        json={"personnel_id": own["id"], "assessment_period": "2026-Q2"},
    ).json()
    other_assessment = client.post(
        "/api/v1/et/assessments",
        headers=admin_headers,
        json={"personnel_id": other["id"], "assessment_period": "2026-Q2"},
    ).json()
    team_headers = _login(client, "TBĐL", "tbdl-pass")

    assert client.get(f"/api/v1/et/assessments/{own_assessment['id']}", headers=team_headers).status_code == 200
    assert client.get(f"/api/v1/et/assessments/{other_assessment['id']}", headers=team_headers).status_code == 403
    list_response = client.get("/api/v1/et/assessments", headers=team_headers)
    assert list_response.status_code == 200
    assert [row["id"] for row in list_response.json()] == [own_assessment["id"]]


def test_personnel_can_include_backend_users(client, admin_headers):
    response = client.get("/api/v1/et/personnel?include_users=true", headers=admin_headers)

    assert response.status_code == 200
    rows = response.json()
    admin_row = next(row for row in rows if row["user_id"] == "admin")
    assert admin_row["full_name"] == "Demo Admin"
    assert admin_row["role"] == "Admin"
    assert admin_row["team"] == ""
    assert admin_row["employee_code"] == ""
    assert admin_row["status"] == "active"
    assert admin_row["salary_grade"] == ""
    assert admin_row["source_type"] == "user"


def test_personnel_draft_search_and_hide_flow(client, admin_headers):
    draft_response = client.post(
        "/api/v1/et/personnel",
        headers=admin_headers,
        json={
            "full_name": "Draft Person",
            "role": "Technician",
            "team": "Draft Team",
            "salary_grade": "L3",
            "status": "active",
        },
    )
    assert draft_response.status_code == 200, draft_response.text
    draft = draft_response.json()
    assert draft["employee_code"] is None
    assert draft["position_code"] is None
    assert draft["current_level"] is None

    search_response = client.get("/api/v1/et/personnel?include_users=true&search=L3", headers=admin_headers)
    assert search_response.status_code == 200
    assert [row["id"] for row in search_response.json()] == [draft["id"]]

    assessment_response = client.post(
        "/api/v1/et/assessments",
        headers=admin_headers,
        json={"personnel_id": draft["id"], "assessment_period": "2026-Q3"},
    )
    assert assessment_response.status_code == 422

    hide_response = client.delete(f"/api/v1/et/personnel/visibility/personnel/{draft['id']}", headers=admin_headers)
    assert hide_response.status_code == 200
    list_response = client.get("/api/v1/et/personnel?include_users=true&search=Draft Person", headers=admin_headers)
    assert list_response.status_code == 200
    assert list_response.json() == []

    users_response = client.get("/api/v1/et/personnel?include_users=true", headers=admin_headers)
    admin_row = next(row for row in users_response.json() if row["user_id"] == "admin")
    hide_user = client.delete("/api/v1/et/personnel/visibility/user/admin", headers=admin_headers)
    assert hide_user.status_code == 200
    assert admin_row["source_type"] == "user"
    users_after_hide = client.get("/api/v1/et/personnel?include_users=true&search=admin", headers=admin_headers)
    assert users_after_hide.status_code == 200
    assert all(row["user_id"] != "admin" for row in users_after_hide.json())
