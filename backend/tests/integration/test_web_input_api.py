from io import BytesIO
from datetime import datetime, timezone

from openpyxl import load_workbook

from app.db.session import create_session
from app.models.domain import FIMonthlyAllocationModel, SKCTKTModel, TeamReportModel


def _login(client, user_id: str, password: str) -> dict[str, str]:
    response = client.post("/api/v1/auth/login", json={"user_id": user_id, "password": password})
    assert response.status_code == 200, response.text
    return {"Authorization": f"Bearer {response.json()['access_token']}"}


def _valid_payload(client, headers):
    mapping = client.get("/api/v1/okr/kr-mapping", headers=headers)
    assert mapping.status_code == 200, mapping.text
    return {
        "kr_assessments": [
            {
                "workshop_kr_code": item["workshop_kr_code"],
                "implementation_report": f"Đã thực hiện {item['workshop_kr_code']}",
                "team_self_assessment": "Hoàn thành",
                "notes": "",
            }
            for item in mapping.json()
        ],
        "arising_work": [{"content": "Hoàn thành xử lý công việc phát sinh", "status": "Hoàn thành"}],
        "monthly_conclusion": {
            "discipline_status": "OK",
            "discipline_description": "",
            "overall_assessment": "Hoàn thành nhiệm vụ",
            "detailed_description": "",
        },
        "objective_overrides": {"O1": "Hoàn thành tốt nhiệm vụ"},
    }


def _approved_fi(record_id: str, *, team: str, approved_at: datetime) -> SKCTKTModel:
    return SKCTKTModel(
        id=record_id,
        sk_code=f"FI/04/2026-{team}-{record_id[-2:]}",
        title=f"FI hợp lệ {record_id}",
        author_name="Nhân sự kiểm thử",
        author_user_id="staff-test",
        team=team,
        content_description="Nội dung FI kiểm thử",
        completion_plan="Kế hoạch hoàn thành",
        status="Approved",
        status_history=[],
        consider_for_khmt=False,
        is_public=True,
        is_counted_for_okr=False,
        is_historical_import=False,
        created_at=approved_at,
        submitted_at=approved_at,
        approved_at=approved_at,
    )


def test_web_input_draft_submit_lock_export_flow(client):
    headers = _login(client, "admin", "admin-pass")
    payload = _valid_payload(client, headers)

    draft = client.put(
        "/api/v1/okr/web-input/TBCH/4/2026/draft",
        headers=headers,
        json={"data": {**payload, "kr_assessments": payload["kr_assessments"][:2]}, "expected_version": None},
    )
    assert draft.status_code == 200, draft.text
    assert draft.json()["status"] == "Đang nhập"
    assert draft.json()["version"] == 1
    draft_report_id = draft.json()["report"]["id"]

    admin_headers = headers
    manager_reports = client.get("/api/v1/okr/reports", headers=admin_headers)
    assert manager_reports.status_code == 200, manager_reports.text
    assert manager_reports.json() == []
    draft_preview = client.get(f"/api/v1/okr/reports/{draft_report_id}/preview", headers=admin_headers)
    assert draft_preview.status_code == 403

    conflict = client.put(
        "/api/v1/okr/web-input/TBCH/4/2026/draft",
        headers=headers,
        json={"data": payload, "expected_version": 999},
    )
    assert conflict.status_code == 409

    submitted = client.post(
        "/api/v1/okr/web-input/TBCH/4/2026/submit",
        headers=headers,
        json={"data": payload},
    )
    assert submitted.status_code == 200, submitted.text
    assert submitted.json()["status"] == "Đã gửi"
    assert "Mục tiêu ĐK.O1.TBCH.O1: Hoàn thành tốt nhiệm vụ" in submitted.json()["email_text"]

    dashboard = client.get("/api/v1/okr/dashboard/4/2026", headers=headers)
    assert dashboard.status_code == 200, dashboard.text
    tbch = next(row for row in dashboard.json()["teams"] if row["team"] == "TBCH")
    assert tbch["monthly_assessment"] == "Hoàn thành nhiệm vụ"

    excel = client.get("/api/v1/okr/web-input/TBCH/4/2026/export/excel", headers=headers)
    assert excel.status_code == 200, excel.text
    workbook = load_workbook(BytesIO(excel.content), read_only=True)
    assert workbook.sheetnames
    assert any(
        cell == "Đã thực hiện O1.KR1"
        for sheet in workbook.worksheets
        for row in sheet.iter_rows(values_only=True)
        for cell in row
    )

    email = client.get("/api/v1/okr/web-input/TBCH/4/2026/export/email", headers=headers)
    assert email.status_code == 200, email.text
    assert email.json()["text"].count("Mục tiêu ĐK.") == 6

    fi_headers = _login(client, "fi", "fi-pass")
    fi_email = client.get("/api/v1/okr/web-input/TBCH/4/2026/export/email", headers=fi_headers)
    assert fi_email.status_code == 200, fi_email.text
    fi_excel = client.get("/api/v1/okr/web-input/TBCH/4/2026/export/excel", headers=fi_headers)
    assert fi_excel.status_code == 200, fi_excel.text

    manager_reports_after_submit = client.get("/api/v1/okr/reports", headers=admin_headers)
    assert manager_reports_after_submit.status_code == 200
    assert [report["report_status"] for report in manager_reports_after_submit.json()] == ["submitted"]

    with create_session() as db:
        db.add(
            _approved_fi(
                "sk-web-lock-01",
                team="TBCH",
                approved_at=datetime(2026, 4, 20, tzinfo=timezone.utc),
            )
        )
        db.commit()

    allocation_preview = client.get(
        "/api/v1/okr/web-input/TBCH/4/2026/fi-allocation-preview",
        headers=admin_headers,
    )
    assert allocation_preview.status_code == 200, allocation_preview.text
    assert allocation_preview.json()["required_count"] == 1
    assert allocation_preview.json()["selected_sk_ids"] == ["sk-web-lock-01"]

    locked = client.post(
        "/api/v1/okr/web-input/TBCH/4/2026/lock",
        headers=admin_headers,
        json={"reason": "Chốt tháng"},
    )
    assert locked.status_code == 200, locked.text
    assert locked.json()["status"] == "Đã chốt"
    assert locked.json()["fi_allocation"]["allocated_count"] == 1
    assert locked.json()["fi_allocation"]["selected_sk_ids"] == ["sk-web-lock-01"]

    blocked = client.put(
        "/api/v1/okr/web-input/TBCH/4/2026/draft",
        headers=headers,
        json={"data": payload, "expected_version": None},
    )
    assert blocked.status_code == 409

    unlocked = client.post(
        "/api/v1/okr/web-input/TBCH/4/2026/unlock",
        headers=admin_headers,
        json={"reason": "Mở để chỉnh sửa"},
    )
    assert unlocked.status_code == 200, unlocked.text
    assert unlocked.json()["status"] == "Đã gửi"

    with create_session() as db:
        plan = db.query(FIMonthlyAllocationModel).filter_by(team="TBCH", month=4, year=2026).one()
        selected = db.get(SKCTKTModel, "sk-web-lock-01")
        assert plan.status == "reopened"
        assert selected is not None and selected.consider_for_khmt is True


def test_fi_shortage_blocks_lock_and_preserves_submitted_report(client):
    headers = _login(client, "admin", "admin-pass")
    payload = _valid_payload(client, headers)
    payload["monthly_conclusion"]["overall_assessment"] = "Hoàn thành tốt nhiệm vụ"

    submitted = client.post(
        "/api/v1/okr/web-input/TBĐL/7/2026/submit",
        headers=headers,
        json={"data": payload},
    )
    assert submitted.status_code == 200, submitted.text

    with create_session() as db:
        db.add(
            _approved_fi(
                "sk-shortage-01",
                team="TBĐL",
                approved_at=datetime(2026, 6, 20, tzinfo=timezone.utc),
            )
        )
        db.commit()

    preview = client.get(
        "/api/v1/okr/web-input/TBĐL/7/2026/fi-allocation-preview",
        headers=headers,
    )
    assert preview.status_code == 200, preview.text
    assert preview.json()["required_count"] == 3
    assert preview.json()["available_count"] == 1
    assert preview.json()["shortage_count"] == 2
    assert preview.json()["can_finalize"] is False

    locked = client.post(
        "/api/v1/okr/web-input/TBĐL/7/2026/lock",
        headers=headers,
        json={"reason": "Không được khóa khi thiếu FI"},
    )
    assert locked.status_code == 409, locked.text
    assert locked.json()["detail"]["error_code"] == "FI_ALLOCATION_SHORTAGE"

    with create_session() as db:
        report = db.query(TeamReportModel).filter_by(
            team="TBĐL",
            report_month=7,
            report_year=2026,
            is_current_version=True,
        ).one()
        fi_record = db.get(SKCTKTModel, "sk-shortage-01")
        assert report.report_status == "submitted"
        assert fi_record is not None and fi_record.consider_for_khmt is False
        assert db.query(FIMonthlyAllocationModel).count() == 0


def test_team_account_cannot_access_other_team_web_input(client):
    headers = _login(client, "TBCH", "tbch-pass")
    response = client.get("/api/v1/okr/web-input/TBHTĐK/4/2026", headers=headers)
    assert response.status_code == 403


def test_team_account_cannot_write_web_input(client):
    headers = _login(client, "TBCH", "tbch-pass")
    payload = _valid_payload(client, headers)

    draft = client.put(
        "/api/v1/okr/web-input/TBCH/4/2026/draft",
        headers=headers,
        json={"data": payload, "expected_version": None},
    )
    assert draft.status_code == 403

    submitted = client.post(
        "/api/v1/okr/web-input/TBCH/4/2026/submit",
        headers=headers,
        json={"data": payload},
    )
    assert submitted.status_code == 403


def test_web_input_submit_returns_validation_errors(client):
    headers = _login(client, "admin", "admin-pass")
    response = client.post(
        "/api/v1/okr/web-input/TBCH/4/2026/submit",
        headers=headers,
        json={
            "data": {
                "kr_assessments": [
                    {
                        "workshop_kr_code": "O1.KR1",
                        "implementation_report": "",
                        "team_self_assessment": "Hoàn thành",
                        "notes": "",
                    }
                ],
                "arising_work": [],
                "monthly_conclusion": {
                    "discipline_status": "OK",
                    "discipline_description": "",
                    "overall_assessment": "Hoàn thành nhiệm vụ",
                    "detailed_description": "",
                },
                "objective_overrides": {},
            }
        },
    )
    assert response.status_code == 400
    detail = response.json()["detail"]
    assert detail["error_code"] == "VALIDATION_ERROR"
    assert any(item["kr_code"] == "O1.KR1" for item in detail["details"])
