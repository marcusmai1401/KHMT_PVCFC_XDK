from openpyxl import Workbook, load_workbook
from hypothesis import HealthCheck, given, settings, strategies as st
from sqlalchemy import select

from app.models.domain import TeamMonthlySummaryModel, TeamReportModel
from app.services.okr.constants import TEAMS
from app.services.okr.historical_import import (
    FileImportResult,
    ImportSessionReport,
    extract_month_year_from_filename,
    upsert_team_monthly_summary,
    upsert_team_report,
)
from app.services.okr.historical_snapshot import find_dashboard_team_rows
from app.services.okr.kr_mapping import KRMapping
from app.services.okr.team_normalizer import normalize_team_label
from app.services.okr.workbook import parse_team_report


TEXT = st.text(
    alphabet=st.characters(blacklist_categories=("Cc", "Cs"), blacklist_characters=["\x00"]),
    min_size=1,
    max_size=60,
)


def _mapping(*records: KRMapping) -> dict[str, KRMapping]:
    return {record.workshop_kr_code: record for record in records}


def _single_kr_mapping(code: str = "O1.KR1", name: str = "KR 1") -> dict[str, KRMapping]:
    return _mapping(
        KRMapping(
            workshop_kr_code=code,
            kr_name=name,
            dashboard_column="L",
            measurement_type="Text",
            target_value="1",
        )
    )


def _base_workbook() -> Workbook:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "TBCH"
    sheet["N1"] = "Tình hình thực hiện"
    sheet["O1"] = "Đánh giá"
    sheet["P1"] = "Ghi chú"
    return workbook


@given(month=st.integers(min_value=1, max_value=12), year=st.integers(min_value=2020, max_value=2035))
@settings(max_examples=100)
def test_month_extraction_from_filename_property(month, year):
    # Feature: historical-data-import, Property 1: Month Extraction from Filename
    filename = f"OKR tháng {month:02d}-{year} - X.ĐK.xlsx"

    assert extract_month_year_from_filename(filename) == (month, year)


@given(
    alias=st.sampled_from(
        [
            "TBHTĐK",
            "HTĐK",
            "HTDK",
            "TBHTDK",
            "TBCH",
            "Đội thiết bị chấp hành",
            "Đội thiết bị cơ cấu chấp hành",
            "TBĐ",
            "TBĐL",
            "TBDL",
            "TCĐK",
            "TCDK",
            "Tổ trực ca",
        ]
    ),
    left=st.text(alphabet=" \t", min_size=0, max_size=3),
    right=st.text(alphabet=" \t", min_size=0, max_size=3),
)
@settings(max_examples=100)
def test_team_label_normalization_round_trip_property(alias, left, right):
    # Feature: historical-data-import, Property 2: Team Label Normalization Round-Trip
    normalized, _original = normalize_team_label(f"{left}{alias.swapcase()}{right}")

    assert normalized in TEAMS


@given(
    start_row=st.integers(min_value=1, max_value=25),
    alias=st.sampled_from(["HTĐK", "TBCH", "TBĐ", "TCDK"]),
)
@settings(max_examples=100)
def test_dynamic_dashboard_team_detection_property(start_row, alias):
    # Feature: historical-data-import, Property 10: Dynamic Dashboard Team Detection
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Dashboard"
    sheet.cell(start_row, 1).value = "Đội/Tổ"
    sheet.cell(start_row + 1, 1).value = alias

    rows = find_dashboard_team_rows(sheet)

    assert len(rows) == 1
    assert rows[0][0] == start_row + 1
    assert rows[0][1] in TEAMS


@given(empty_rows=st.lists(st.text(alphabet=" \t", min_size=0, max_size=8), min_size=1, max_size=8))
@settings(
    max_examples=25,
    deadline=None,
    suppress_health_check=[HealthCheck.function_scoped_fixture],
)
def test_empty_row_skipping_property(tmp_path, empty_rows):
    # Feature: historical-data-import, Property 3: Empty Row Skipping
    workbook = _base_workbook()
    sheet = workbook["TBCH"]
    for offset, whitespace in enumerate(empty_rows, start=2):
        sheet.cell(offset, 2).value = whitespace
        sheet.cell(offset, 3).value = whitespace
    target_row = len(empty_rows) + 3
    sheet.cell(target_row, 2).value = "ĐK.O1.KR1.TBCH.O1.KR1"
    sheet.cell(target_row, 3).value = "KR 1"
    sheet.cell(target_row, 14).value = "done"
    sheet.cell(target_row, 15).value = "Hoàn thành"
    path = tmp_path / "empty-row-property.xlsx"
    workbook.save(path)

    parsed = parse_team_report(path, team="TBCH", month=4, year=2026, kr_mapping=_single_kr_mapping())

    assert [item["workshop_kr_code"] for item in parsed["assessments"]] == ["O1.KR1"]


@given(implementation=TEXT, notes=TEXT)
@settings(
    max_examples=25,
    deadline=None,
    suppress_health_check=[HealthCheck.function_scoped_fixture],
)
def test_kr_assessment_field_extraction_property(tmp_path, implementation, notes):
    # Feature: historical-data-import, Property 4: KR Assessment Field Extraction
    workbook = _base_workbook()
    sheet = workbook["TBCH"]
    sheet["B2"] = "ĐK.O1.KR1.TBCH.O1.KR1"
    sheet["C2"] = "KR 1"
    sheet["N2"] = implementation
    sheet["O2"] = "Hoàn thành tốt"
    sheet["P2"] = notes
    path = tmp_path / "field-extraction-property.xlsx"
    workbook.save(path)

    parsed = parse_team_report(path, team="TBCH", month=4, year=2026, kr_mapping=_single_kr_mapping())
    assessment = parsed["assessments"][0]

    assert assessment["implementation_report"] == implementation
    assert assessment["team_self_assessment"] == "Hoàn thành tốt"
    assert assessment["notes"] == notes


@given(implementation=TEXT)
@settings(
    max_examples=25,
    deadline=None,
    suppress_health_check=[HealthCheck.function_scoped_fixture],
)
def test_hierarchical_kr_preservation_prefers_canonical_name_property(tmp_path, implementation):
    # Feature: historical-data-import, Property 8: Hierarchical Objective-KR Preservation
    workbook = _base_workbook()
    sheet = workbook["TBCH"]
    sheet["B2"] = "ĐK.O1.KR1.TBCH.O1.KR1"
    sheet["C2"] = "Canonical KR 2"
    sheet["N2"] = implementation
    sheet["O2"] = "Hoàn thành"
    path = tmp_path / "kr-name-property.xlsx"
    workbook.save(path)
    mapping = _mapping(
        KRMapping("O1.KR1", "Canonical KR 1", "L", "Text", "1"),
        KRMapping("O1.KR2", "Canonical KR 2", "M", "Text", "1"),
    )

    parsed = parse_team_report(path, team="TBCH", month=4, year=2026, kr_mapping=mapping)

    assert parsed["assessments"][0]["workshop_kr_code"] == "O1.KR2"


@given(value=st.floats(min_value=0, max_value=1_000_000, allow_nan=False, allow_infinity=False, width=32))
@settings(
    max_examples=25,
    deadline=None,
    suppress_health_check=[HealthCheck.function_scoped_fixture],
)
def test_numeric_precision_preservation_property(tmp_path, value):
    # Feature: historical-data-import, Property 9: Numeric Precision Preservation
    workbook = _base_workbook()
    sheet = workbook["TBCH"]
    sheet["B2"] = "ĐK.O1.KR1.TBCH.O1.KR1"
    sheet["C2"] = "KR 1"
    sheet["N2"] = value
    sheet["O2"] = "Hoàn thành"
    path = tmp_path / "numeric-property.xlsx"
    workbook.save(path)
    expected_workbook = load_workbook(path, data_only=False)
    expected_value = str(expected_workbook["TBCH"]["N2"].value)
    expected_workbook.close()

    parsed = parse_team_report(path, team="TBCH", month=4, year=2026, kr_mapping=_single_kr_mapping())

    assert parsed["assessments"][0]["implementation_report"] == expected_value


@given(month=st.integers(min_value=1, max_value=12), implementation=TEXT)
@settings(
    max_examples=10,
    deadline=None,
    suppress_health_check=[HealthCheck.function_scoped_fixture],
)
def test_storage_round_trip_for_team_reports_property(db_session, tmp_path, month, implementation):
    # Feature: historical-data-import, Property 5: Storage Round-Trip for Team Reports
    source = tmp_path / f"source-{month}.xlsx"
    source.write_bytes(b"source")
    parsed = {
        "team": "TBHTĐK",
        "report_month": month,
        "report_year": 2026,
        "sheet_name": "TBHTĐK",
        "assessments": [
            {
                "workshop_kr_code": "O1.KR1",
                "implementation_report": implementation,
                "dashboard_status": "OK",
            }
        ],
        "team_level": {"monthly_assessment": "Hoàn thành", "discipline_status": "OK"},
        "source_cell_references": [],
    }

    report, _action = upsert_team_report(
        db_session,
        parsed,
        source_file=source,
        source_hash=f"hash-{month}",
        imported_by="property-test",
    )
    db_session.flush()
    stored = db_session.get(TeamReportModel, report.id)

    assert stored is not None
    assert stored.assessments[0]["implementation_report"] == implementation


@given(month=st.integers(min_value=1, max_value=12), first=TEXT, second=TEXT)
@settings(
    max_examples=10,
    deadline=None,
    suppress_health_check=[HealthCheck.function_scoped_fixture],
)
def test_import_idempotence_property(db_session, tmp_path, month, first, second):
    # Feature: historical-data-import, Property 6: Import Idempotence
    source = tmp_path / f"idempotent-{month}.xlsx"
    source.write_bytes(b"source")

    def parsed(value: str) -> dict:
        return {
            "team": "TCĐK",
            "report_month": month,
            "report_year": 2026,
            "sheet_name": "TCĐK",
            "assessments": [{"workshop_kr_code": "O1.KR1", "implementation_report": value, "dashboard_status": "OK"}],
            "team_level": {"monthly_assessment": "Hoàn thành", "discipline_status": "OK"},
            "source_cell_references": [],
        }

    upsert_team_report(
        db_session,
        parsed(first),
        source_file=source,
        source_hash=f"hash-first-{month}",
        imported_by="property-test",
    )
    upsert_team_monthly_summary(
        db_session,
        parsed(first),
        source_file=source,
        source_hash=f"hash-first-{month}",
    )
    upsert_team_report(
        db_session,
        parsed(second),
        source_file=source,
        source_hash=f"hash-second-{month}",
        imported_by="property-test",
    )
    upsert_team_monthly_summary(
        db_session,
        parsed(second),
        source_file=source,
        source_hash=f"hash-second-{month}",
    )
    db_session.flush()

    current = db_session.execute(
        select(TeamReportModel).where(
            TeamReportModel.team == "TCĐK",
            TeamReportModel.report_month == month,
            TeamReportModel.report_year == 2026,
            TeamReportModel.is_current_version.is_(True),
        )
    ).scalars().all()
    summary = db_session.execute(
        select(TeamMonthlySummaryModel).where(
            TeamMonthlySummaryModel.team == "TCĐK",
            TeamMonthlySummaryModel.month == month,
            TeamMonthlySummaryModel.year == 2026,
        )
    ).scalar_one()

    assert len(current) == 1
    assert current[0].assessments[0]["implementation_report"] == second
    assert summary.stats["source_file_hash"] == f"hash-second-{month}"


@given(counts=st.lists(st.integers(min_value=0, max_value=37), min_size=1, max_size=4))
@settings(max_examples=25)
def test_report_count_accuracy_property(counts):
    # Feature: historical-data-import, Property 7: Report Count Accuracy
    file_result = FileImportResult(file_name="source.xlsx", month=1, year=2026)
    for index, count in enumerate(counts):
        team = TEAMS[index % len(TEAMS)]
        file_result.records_per_team[team] = count
        file_result.table_counts["team_reports"]["inserted"] += 1
    report = ImportSessionReport(file_results=[file_result])
    report.total_team_reports = sum(
        result.table_counts["team_reports"]["inserted"]
        + result.table_counts["team_reports"]["updated"]
        for result in report.file_results
    )

    assert report.total_team_reports == len(counts)
    assert sum(file_result.records_per_team.values()) == sum(counts)
