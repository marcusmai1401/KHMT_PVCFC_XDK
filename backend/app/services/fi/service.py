from collections import Counter
from datetime import datetime, timezone
from pathlib import Path
from typing import Any
import unicodedata

from fastapi import HTTPException
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.core.config import settings
from app.core.security import Role
from app.models.domain import SKCTKTModel, SKCodeSequenceModel, SKImageModel
from app.services.fi.completion import (
    completion_date_to_datetime,
    completion_plan_completed_at,
    completion_plan_indicates_done,
)
from app.services.fi.workflow import FIAction, SKStatus, is_public_status, next_status
from app.services.repositories import audit, make_id, model_to_dict, notify

NON_DRAFT_STATUSES = {status.value for status in SKStatus if status != SKStatus.DRAFT}
OWNER_DELETABLE_STATUSES = {SKStatus.DRAFT.value, SKStatus.SUBMITTED.value, SKStatus.NEED_MORE_INFO.value}
REVIEWABLE_STATUSES = {
    SKStatus.SUBMITTED.value,
    SKStatus.REVIEWED.value,
    SKStatus.DEFERRED.value,
    SKStatus.APPROVED.value,
    SKStatus.REJECTED.value,
}
REVIEW_DECISION_ACTIONS = {FIAction.APPROVE.value, FIAction.DEFER.value, FIAction.REJECT.value}
HISTORICAL_REVIEW_ACTIONS = REVIEW_DECISION_ACTIONS
SHARED_CONTENT_EDIT_FIELDS = {
    "content_description",
    "completion_date",
    "completion_done",
    "completion_plan",
    "title",
}
# Người dùng được phép đứng tên đăng ký SK-CTKT.
AUTHOR_ROLES = {Role.TEAM_ACCOUNT.value, Role.STAFF.value, Role.FI_COORDINATOR.value}
# Sau khi đã gửi duyệt, tác giả vẫn được sửa nội dung nhưng phải gửi noti
# cho người xét duyệt biết để xem lại.
EDITABLE_AFTER_SUBMIT_STATUSES = {
    SKStatus.SUBMITTED.value,
    SKStatus.NEED_MORE_INFO.value,
    SKStatus.REVIEWED.value,
    SKStatus.APPROVED.value,
    SKStatus.REJECTED.value,
    SKStatus.DEFERRED.value,
}
AUTHOR_CONTENT_EDITABLE_STATUSES = {SKStatus.DRAFT.value, *EDITABLE_AFTER_SUBMIT_STATUSES}


def _utc_now() -> datetime:
    return datetime.now(timezone.utc)


def generate_sk_code(db: Session, team: str, year: int, month: int) -> str:
    """Sinh mã SK-CTKT theo format FI/MM/YYYY-TEAM-NN.

    Sequence được reset mỗi tháng cho từng team — ví dụ:
      * TBĐL tháng 5/2026: FI/05/2026-TBĐL-01, -02, -03 ...
      * TBĐL tháng 6/2026: lại quay về -01
      * TBCH tháng 5/2026: -01 (độc lập với TBĐL)
    """
    prefix = f"FI/{month:02d}/{year}-{team}"
    sequence = db.get(SKCodeSequenceModel, prefix)
    if sequence is None:
        sequence = SKCodeSequenceModel(prefix=prefix, next_value=1)
        db.add(sequence)
        db.flush()
    value = sequence.next_value
    sequence.next_value += 1
    return f"{prefix}-{value:02d}"


def _int_or_default(value: Any, default: int) -> int:
    try:
        return int(value)
    except (TypeError, ValueError):
        return default


def _completed_at_from_payload(
    payload: dict[str, Any],
    *,
    fallback: datetime,
    existing: datetime | None = None,
) -> datetime | None:
    completion_done = payload.get("completion_done")
    if completion_done is not None:
        if not completion_done:
            return None
        return (
            completion_date_to_datetime(payload.get("completion_date"))
            or completion_plan_completed_at(payload.get("completion_plan"), fallback=fallback)
            or fallback
        )
    if "completion_plan" in payload:
        inferred = completion_plan_completed_at(payload.get("completion_plan"), fallback=fallback)
        if inferred is not None:
            return inferred
    return existing


def submitter_user_id(record: SKCTKTModel) -> str | None:
    history = record.status_history if isinstance(record.status_history, list) else []
    if not history:
        return None
    first = history[0] if isinstance(history[0], dict) else {}
    comments = first.get("comments") if isinstance(first.get("comments"), dict) else {}
    submitted_by = comments.get("submitted_by") or comments.get("created_by")
    return str(submitted_by or first.get("changed_by") or "") or None


def is_author_or_submitter(record: SKCTKTModel, actor: str) -> bool:
    return record.author_user_id == actor or submitter_user_id(record) == actor


def create_sk_ctkt(
    db: Session,
    payload: dict[str, Any],
    actor: str,
    *,
    submit_immediately: bool = False,
) -> SKCTKTModel:
    now = _utc_now()
    team = payload["team"]
    year = _int_or_default(payload.get("year") or payload.get("registration_year"), now.year)
    registration_month = _int_or_default(payload.get("registration_month"), now.month)
    if registration_month < 1 or registration_month > 12:
        registration_month = now.month
    registration_year = _int_or_default(payload.get("registration_year"), year)
    completed_at = _completed_at_from_payload(payload, fallback=now)
    initial_status = SKStatus.SUBMITTED if submit_immediately else SKStatus.DRAFT
    record = SKCTKTModel(
        id=make_id("sk"),
        sk_code=generate_sk_code(db, team, registration_year, registration_month),
        title=payload["title"],
        author_name=payload["author_name"],
        author_user_id=payload.get("author_user_id") or actor,
        team=team,
        content_description=payload["content_description"],
        completion_plan=payload["completion_plan"],
        status=initial_status.value,
        status_history=[
            {
                "from_status": None,
                "to_status": initial_status.value,
                "changed_by": actor,
                "changed_at": now.isoformat(),
                "reason": "web_registration",
                "comments": {
                    "registration_month": registration_month,
                    "registration_year": registration_year,
                    "source": "web",
                    "submitted_by": actor,
                },
            }
        ],
        consider_for_khmt=False,
        is_public=False,
        is_counted_for_okr=False,
        is_historical_import=False,
        created_at=now,
        updated_at=now,
        submitted_at=now if submit_immediately else None,
        completed_at=completed_at,
    )
    db.add(record)
    audit(db, actor, "SK_CTKT", record.id, "create", model_to_dict(record))
    if submit_immediately:
        _notify_transition(db, record, record.id)
    db.commit()
    db.refresh(record)
    return record


def can_view_sk(record: SKCTKTModel, principal: dict[str, str]) -> bool:
    """Quyền xem SK trong danh sách private /fi/sk-ctkt.

    Mọi role trong module FI được xem toàn bộ SK đã gửi/legacy để tra cứu liên đội.
    Bản nháp vẫn chỉ hiện cho chính tác giả, vì đó chưa phải thông tin FI đã công bố.
    """
    role = principal["role"]
    user_id = principal["user_id"]
    allowed_roles = {
        Role.ADMIN.value,
        Role.FI_COORDINATOR.value,
        Role.WORKSHOP_LEADER.value,
        Role.TEAM_ACCOUNT.value,
        Role.STAFF.value,
    }
    if role not in allowed_roles:
        return False
    if record.status == SKStatus.DRAFT.value:
        return is_author_or_submitter(record, user_id)
    return True


def require_visible(record: SKCTKTModel, principal: dict[str, str]) -> None:
    if not can_view_sk(record, principal):
        raise HTTPException(status_code=403, detail="Not allowed")


def update_sk_ctkt(
    db: Session,
    record_id: str,
    payload: dict[str, Any],
    actor: str,
    role: str,
) -> SKCTKTModel:
    record = db.get(SKCTKTModel, record_id)
    if record is None:
        raise KeyError("SK-CTKT not found")
    requested_fields = set(payload)
    shared_content_edit = bool(requested_fields) and requested_fields <= SHARED_CONTENT_EDIT_FIELDS
    if record.is_historical_import:
        raise PermissionError("FI legacy chỉ dùng để tra cứu lịch sử, không chỉnh sửa nội dung")
    if not is_author_or_submitter(record, actor):
        raise PermissionError("Chỉ tác giả hoặc người gửi hộ mới được chỉnh sửa")
    if record.status in {SKStatus.CANCELLED.value, SKStatus.COMPLETED.value}:
        raise PermissionError("SK đã hoàn tất hoặc đã hủy, không chỉnh sửa được nội dung")
    if record.status not in AUTHOR_CONTENT_EDITABLE_STATUSES:
        raise PermissionError("Trạng thái hiện tại không cho phép chỉnh sửa nội dung")
    if not shared_content_edit:
        raise PermissionError("Tác giả chỉ được cập nhật tiêu đề, nội dung đăng ký và kế hoạch thực hiện")
    before = model_to_dict(record)
    if "team" in payload and payload["team"] != record.team:
        raise PermissionError("Tài khoản không được đổi đội/tổ của SK")
    previous_status = record.status
    for field in ["title", "content_description", "completion_plan"]:
        if field in payload:
            setattr(record, field, payload[field])
    if {"completion_plan", "completion_done", "completion_date"} & requested_fields:
        record.completed_at = _completed_at_from_payload(
            payload,
            fallback=record.created_at or _utc_now(),
            existing=record.completed_at,
        )
    record.updated_at = _utc_now()
    audit(db, actor, "SK_CTKT", record_id, "update", {"before": before, "after": model_to_dict(record)})
    # Nếu tác giả chỉnh sửa SK đã gửi duyệt thì gửi noti cho FI_Coordinator +
    # Admin để xét duyệt lại.
    if previous_status in EDITABLE_AFTER_SUBMIT_STATUSES:
        _notify_content_edit(db, record, actor)
    db.commit()
    db.refresh(record)
    return record


def _notify_content_edit(db: Session, record: SKCTKTModel, actor: str) -> None:
    payload = {
        "id": record.id,
        "sk_code": record.sk_code,
        "status": record.status,
        "team": record.team,
        "edited_by": actor,
    }
    notify(db, "SK_CONTENT_EDITED", payload, recipient_role=Role.FI_COORDINATOR.value)
    notify(db, "SK_CONTENT_EDITED", payload, recipient_role=Role.ADMIN.value)


def _validate_actor_for_transition(record: SKCTKTModel, action: str, actor: str, role: str) -> None:
    if action in {FIAction.SUBMIT.value, FIAction.CANCEL.value} and role in AUTHOR_ROLES:
        if not is_author_or_submitter(record, actor):
            raise PermissionError("Chỉ tài khoản tạo SK mới được thực hiện thao tác này")
        if action == FIAction.SUBMIT.value and record.status not in {
            SKStatus.DRAFT.value,
            SKStatus.NEED_MORE_INFO.value,
            SKStatus.SUBMITTED.value,
        }:
            raise PermissionError("Chỉ gửi duyệt được SK ở trạng thái Nháp hoặc Cần bổ sung")
        if action == FIAction.CANCEL.value and record.status not in {
            SKStatus.DRAFT.value,
            SKStatus.NEED_MORE_INFO.value,
        }:
            raise PermissionError("Chỉ hủy được SK ở trạng thái Nháp hoặc Cần bổ sung")
    if role == Role.FI_COORDINATOR.value and action in REVIEW_DECISION_ACTIONS:
        if record.status not in REVIEWABLE_STATUSES:
            raise PermissionError(
                "Chỉ đánh giá được SK ở trạng thái Chờ xét duyệt, Đã xem xét, Xem xét sau, Đồng ý hoặc Không đồng ý"
            )
        # Tránh xung đột lợi ích: Đầu mối FI không được xét duyệt SK do chính mình đăng ký.
        # Admin sẽ là người xét duyệt cho các SK của FI_Coordinator.
        if record.author_user_id == actor or submitter_user_id(record) == actor:
            raise PermissionError(
                "Đầu mối FI không được xét duyệt SK do chính mình đứng tên hoặc gửi hộ — vui lòng để Admin xét duyệt"
            )


def transition_sk_ctkt(
    db: Session,
    record_id: str,
    action: str,
    actor: str,
    role: str,
    note: str | None = None,
    comments: str | None = None,
) -> SKCTKTModel:
    record = db.get(SKCTKTModel, record_id)
    if record is None:
        raise KeyError("SK-CTKT not found")
    if action == FIAction.SUBMIT.value and record.status == SKStatus.SUBMITTED.value:
        _validate_actor_for_transition(record, action, actor, role)
        return record
    if record.is_historical_import:
        if action not in HISTORICAL_REVIEW_ACTIONS or record.status not in REVIEWABLE_STATUSES:
            raise PermissionError("FI legacy chỉ cho đánh giá các mục Chờ xét duyệt, Xem xét sau, Đồng ý hoặc Không đồng ý")
    _validate_actor_for_transition(record, action, actor, role)
    result = next_status(record.status, action, role, note)
    before = record.status
    now = _utc_now()
    record.status = result.to_status.value
    record.is_public = is_public_status(record.status)
    if comments and action == FIAction.REVIEW.value:
        record.fi_coordinator_comments = comments
    if action in REVIEW_DECISION_ACTIONS:
        record.decision_note = note.strip() if note and note.strip() else None
    elif note:
        record.decision_note = note
    if record.is_historical_import:
        record.consider_for_khmt = (
            record.status in {SKStatus.APPROVED.value, SKStatus.COMPLETED.value}
            and record.khmt_month is not None
            and record.khmt_year is not None
        )
        record.is_counted_for_okr = record.consider_for_khmt
    if record.status == SKStatus.SUBMITTED.value:
        record.submitted_at = now
    elif record.status == SKStatus.REVIEWED.value:
        record.reviewed_at = now
    elif record.status == SKStatus.APPROVED.value:
        record.approved_at = now
    elif record.status == SKStatus.COMPLETED.value:
        record.completed_at = now
    if before == SKStatus.APPROVED.value and record.status not in {SKStatus.APPROVED.value, SKStatus.COMPLETED.value}:
        record.approved_at = None
    history = {
        "from_status": before,
        "to_status": record.status,
        "changed_by": actor,
        "changed_at": now.isoformat(),
        "reason": note,
        "comments": comments,
    }
    record.status_history = [*record.status_history, history]
    audit(db, actor, "SK_CTKT", record_id, f"transition:{action}", history)
    _notify_transition(db, record, record_id)
    db.commit()
    db.refresh(record)
    return record


def _notify_transition(db: Session, record: SKCTKTModel, record_id: str) -> None:
    payload = {"id": record_id, "status": record.status, "sk_code": record.sk_code}
    if record.status == SKStatus.SUBMITTED.value:
        notify(db, "SK_SUBMITTED", payload, recipient_role=Role.FI_COORDINATOR.value)
        notify(db, "SK_SUBMITTED", payload, recipient_role=Role.ADMIN.value)
    elif record.status == SKStatus.NEED_MORE_INFO.value:
        notify(db, "SK_NEED_MORE_INFO", payload, recipient_user_id=record.author_user_id)
    elif record.status == SKStatus.REVIEWED.value:
        notify(db, "SK_REVIEWED", payload, recipient_role=Role.FI_COORDINATOR.value)
    elif record.status == SKStatus.APPROVED.value:
        notify(db, "SK_APPROVED", payload, recipient_user_id=record.author_user_id)
        notify(db, "SK_APPROVED", payload, recipient_role=Role.WORKSHOP_LEADER.value)
    elif record.status in {SKStatus.REJECTED.value, SKStatus.DEFERRED.value, SKStatus.CANCELLED.value, SKStatus.COMPLETED.value}:
        notify(db, f"SK_{record.status.upper()}", payload, recipient_user_id=record.author_user_id)
    else:
        notify(db, "SK_STATUS_CHANGED", payload, recipient_user_id=record.author_user_id)


def assign_khmt(
    db: Session,
    record_id: str,
    month: int,
    year: int,
    actor: str,
    role: str = Role.ADMIN.value,
    principal_team: str | None = None,
) -> SKCTKTModel:
    record = db.get(SKCTKTModel, record_id)
    if record is None:
        raise KeyError("SK-CTKT not found")
    if month < 1 or month > 12:
        raise ValueError("Tháng KHMT phải nằm trong khoảng 1-12")
    if year < 2020 or year > 2100:
        raise ValueError("Năm KHMT không hợp lệ")
    _ensure_khmt_editable(record, actor, role, principal_team)
    record.khmt_month = month
    record.khmt_year = year
    record.consider_for_khmt = True
    record.is_counted_for_okr = True
    now = _utc_now()
    history = {
        "from_status": record.status,
        "to_status": record.status,
        "changed_by": actor,
        "changed_at": now.isoformat(),
        "reason": "khmt_assignment",
        "comments": {
            "khmt_month": month,
            "khmt_year": year,
            "source": "workflow",
        },
    }
    record.status_history = [*record.status_history, history]
    record.updated_at = now
    audit(db, actor, "SK_CTKT", record_id, "assign_khmt", {"month": month, "year": year})
    db.commit()
    db.refresh(record)
    return record


def clear_khmt(
    db: Session,
    record_id: str,
    actor: str,
    role: str = Role.ADMIN.value,
    principal_team: str | None = None,
) -> SKCTKTModel:
    record = db.get(SKCTKTModel, record_id)
    if record is None:
        raise KeyError("SK-CTKT not found")
    _ensure_khmt_editable(record, actor, role, principal_team)
    previous_month = record.khmt_month
    previous_year = record.khmt_year
    record.khmt_month = None
    record.khmt_year = None
    record.consider_for_khmt = False
    record.is_counted_for_okr = False
    now = _utc_now()
    history = {
        "from_status": record.status,
        "to_status": record.status,
        "changed_by": actor,
        "changed_at": now.isoformat(),
        "reason": "khmt_unassignment",
        "comments": {
            "previous_khmt_month": previous_month,
            "previous_khmt_year": previous_year,
            "source": "workflow",
        },
    }
    record.status_history = [*record.status_history, history]
    record.updated_at = now
    audit(
        db,
        actor,
        "SK_CTKT",
        record_id,
        "clear_khmt",
        {"previous_month": previous_month, "previous_year": previous_year},
    )
    db.commit()
    db.refresh(record)
    return record


def _ensure_khmt_editable(record: SKCTKTModel, actor: str, role: str, principal_team: str | None) -> None:
    if record.status not in {SKStatus.APPROVED.value, SKStatus.COMPLETED.value}:
        raise ValueError("Only Approved or Completed SK-CTKT can be assigned to KHMT")
    if role == Role.ADMIN.value:
        return
    if role != Role.TEAM_ACCOUNT.value:
        raise PermissionError("Chỉ tài khoản đội/tổ được ghi nhận tháng KHMT")
    team_code = principal_team or actor
    if record.team != team_code:
        raise PermissionError("Tài khoản đội/tổ chỉ được ghi nhận KHMT cho đội/tổ của mình")


FI_DASHBOARD_TEAMS = ["TBCH", "TBĐL", "TBHTĐK", "TCĐK"]
FI_DASHBOARD_STATUSES = [
    SKStatus.DRAFT.value,
    SKStatus.SUBMITTED.value,
    SKStatus.NEED_MORE_INFO.value,
    SKStatus.REVIEWED.value,
    SKStatus.APPROVED.value,
    SKStatus.REJECTED.value,
    SKStatus.DEFERRED.value,
    SKStatus.CANCELLED.value,
    SKStatus.COMPLETED.value,
]
PENDING_STATUSES = {
    SKStatus.SUBMITTED.value,
    SKStatus.NEED_MORE_INFO.value,
    SKStatus.REVIEWED.value,
}
FI_REPORTABLE_STATUSES = {status for status in FI_DASHBOARD_STATUSES if status != SKStatus.DRAFT.value}
FI_EXPORT_DECISIONS = {"approved", "rejected", "deferred", "pending"}
FI_EXPORT_KHMT_FILTERS = {"in", "out"}
FI_EXPORT_COMPLETION_FILTERS = {"done", "pending"}
FI_STATUS_LABELS = {
    SKStatus.DRAFT.value: "Bản nháp",
    SKStatus.SUBMITTED.value: "Chờ xét duyệt",
    SKStatus.NEED_MORE_INFO.value: "Cần bổ sung",
    SKStatus.REVIEWED.value: "Đã xem xét",
    SKStatus.APPROVED.value: "Đã phê duyệt",
    SKStatus.REJECTED.value: "Từ chối",
    SKStatus.DEFERRED.value: "Xem xét sau",
    SKStatus.CANCELLED.value: "Đã hủy",
    SKStatus.COMPLETED.value: "Hoàn tất",
}
FI_DECISION_LABELS = {
    "approved": "Đồng ý",
    "rejected": "Không đồng ý",
    "deferred": "Xem xét sau",
    "pending": "Chưa duyệt",
}
FI_KHMT_LABELS = {"in": "Đã vào KHMT", "out": "Chưa vào KHMT"}
FI_COMPLETION_LABELS = {"done": "Đã hoàn thành", "pending": "Chưa hoàn thành"}
FI_EXPORT_DATA_HEADERS = [
    "STT",
    "Mã SK",
    "Tên SK-CTKT",
    "Tác giả",
    "Tài khoản tác giả",
    "Đội/tổ",
    "Tháng đăng ký",
    "Trạng thái",
    "Kết luận LĐX",
    "KHMT",
    "Tháng KHMT",
    "Hoàn thành",
    "Kế hoạch hoàn thành",
    "Nội dung",
    "Nhận xét FI/BM01",
    "Ghi chú quyết định",
]
FI_EXPORT_DATA_HEADER_ROW = 4

# --- FI export visual theme (PVCFC brand: navy #2C398E + green #169045) ---
FI_NAVY = "2C398E"
FI_GREEN = "169045"
FI_NAVY_TINT = "EAEDFA"
FI_GREEN_TINT = "E6F3EC"
FI_INK = "0F172A"
FI_INK_SOFT = "475569"
FI_STRIPE = "F5F8FC"
FI_SURFACE = "FFFFFF"
FI_GRID = "D5DDEC"
FI_ON_DARK = "FFFFFF"
FI_BASE_FONT = "Calibri"


def is_fi_reportable(record: SKCTKTModel) -> bool:
    """FI dashboards/reports only count records that have been sent to FI."""
    return record.status != SKStatus.DRAFT.value


def _empty_fi_dashboard_bucket(team: str | None = None) -> dict[str, Any]:
    data: dict[str, Any] = {
        "total": 0,
        "approved": 0,
        "completed": 0,
        "deferred": 0,
        "pending": 0,
        "review_passed": 0,
        "review_failed": 0,
        "rejected": 0,
        "cancelled": 0,
        "draft": 0,
        "khmt_considered": 0,
        "khmt_not_considered": 0,
        "completed_count": 0,
        "not_completed": 0,
        "historical": 0,
        "current": 0,
        "status_counts": {status: 0 for status in FI_DASHBOARD_STATUSES},
    }
    if team is not None:
        data["team"] = team
    return data


def _is_khmt_considered(record: SKCTKTModel) -> bool:
    return bool(record.consider_for_khmt)


def _is_completion_done(record: SKCTKTModel) -> bool:
    return bool(
        record.status == SKStatus.COMPLETED.value
        or record.completed_at is not None
        or completion_plan_indicates_done(record.completion_plan)
    )


def _add_record_to_bucket(bucket: dict[str, Any], record: SKCTKTModel) -> None:
    status = record.status
    bucket["total"] += 1
    bucket["status_counts"][status] = bucket["status_counts"].get(status, 0) + 1
    if status in {SKStatus.APPROVED.value, SKStatus.COMPLETED.value}:
        bucket["approved"] += 1
        bucket["review_passed"] += 1
    if _is_completion_done(record):
        bucket["completed"] += 1
        bucket["completed_count"] += 1
    else:
        bucket["not_completed"] += 1
    if status == SKStatus.DEFERRED.value:
        bucket["deferred"] += 1
    if status in PENDING_STATUSES:
        bucket["pending"] += 1
    if status == SKStatus.REJECTED.value:
        bucket["rejected"] += 1
        bucket["review_failed"] += 1
    if status == SKStatus.CANCELLED.value:
        bucket["cancelled"] += 1
    if status == SKStatus.DRAFT.value:
        bucket["draft"] += 1
    if _is_khmt_considered(record):
        bucket["khmt_considered"] += 1
    elif status in {SKStatus.APPROVED.value, SKStatus.COMPLETED.value}:
        bucket["khmt_not_considered"] += 1
    if record.is_historical_import:
        bucket["historical"] += 1
    else:
        bucket["current"] += 1


def fi_dashboard(db: Session, principal: dict[str, str]) -> dict[str, Any]:
    # Dashboard FI là cái nhìn toàn xưởng (tổng hợp theo team/tháng), không
    # phụ thuộc người dùng nên không lọc theo can_view_sk. Mọi role có quyền
    # vào dashboard đều thấy số liệu giống nhau.
    _ = principal  # principal được giữ lại cho audit, không dùng để filter dữ liệu.
    visible_records = db.execute(
        select(SKCTKTModel).where(SKCTKTModel.status.in_(FI_REPORTABLE_STATUSES))
    ).scalars().all()
    by_team = {team: _empty_fi_dashboard_bucket(team) for team in FI_DASHBOARD_TEAMS}
    totals = _empty_fi_dashboard_bucket()
    khmt_by_month: dict[tuple[int, int], int] = {}
    for record in visible_records:
        team_bucket = by_team.setdefault(record.team, _empty_fi_dashboard_bucket(record.team))
        _add_record_to_bucket(team_bucket, record)
        _add_record_to_bucket(totals, record)
        if _is_khmt_considered(record) and record.khmt_month is not None and record.khmt_year is not None:
            key = (record.khmt_year, record.khmt_month)
            khmt_by_month[key] = khmt_by_month.get(key, 0) + 1
    return {
        "generated_at": _utc_now().isoformat(),
        "teams": list(by_team.values()),
        "totals": totals,
        "khmt_by_month": [
            {"year": year, "month": month, "count": count}
            for (year, month), count in sorted(khmt_by_month.items())
        ],
        "status_order": FI_DASHBOARD_STATUSES,
    }


def count_for_okr(db: Session, month: int, year: int) -> dict[str, int]:
    counts = {"TBHTĐK": 0, "TBCH": 0, "TBĐL": 0, "TCĐK": 0}
    records = db.execute(
        select(SKCTKTModel).where(
            SKCTKTModel.khmt_month == month,
            SKCTKTModel.khmt_year == year,
            SKCTKTModel.status.in_([SKStatus.APPROVED.value, SKStatus.COMPLETED.value]),
            SKCTKTModel.consider_for_khmt.is_(True),
            SKCTKTModel.is_counted_for_okr.is_(True),
        )
    ).scalars()
    for record in records:
        if record.team in counts:
            counts[record.team] += 1
    return counts


def build_fi_report_export_filters(
    *,
    teams: str | None = None,
    registration_months: str | None = None,
    decisions: str | None = None,
    khmt: str | None = None,
    completion: str | None = None,
    authors: str | None = None,
    submitters: str | None = None,
) -> dict[str, set[Any]]:
    return {
        "teams": _parse_csv_values(teams),
        "registration_months": _parse_month_values(registration_months),
        "decisions": _parse_allowed_values(decisions, FI_EXPORT_DECISIONS, "decisions"),
        "khmt": _parse_allowed_values(khmt, FI_EXPORT_KHMT_FILTERS, "khmt"),
        "completion": _parse_allowed_values(completion, FI_EXPORT_COMPLETION_FILTERS, "completion"),
        "authors": _parse_csv_values(authors),
        "submitters": _parse_csv_values(submitters),
    }


def export_fi_reports_to_excel(db: Session, filters: dict[str, set[Any]]) -> Path:
    records = _filtered_fi_export_records(db, filters)
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Tong hop"
    data_sheet = workbook.create_sheet("Du lieu FI")
    generated_at = _utc_now()

    _write_fi_export_summary(summary_sheet, records, filters, generated_at)
    _write_fi_export_rows(data_sheet, records, generated_at)
    _style_fi_export_workbook(workbook)

    export_dir = settings.storage_dir / "exports"
    export_dir.mkdir(parents=True, exist_ok=True)
    filename = f"fi-reports-export-{generated_at.strftime('%Y%m%d-%H%M%S')}.xlsx"
    path = export_dir / filename
    workbook.save(path)
    return path


def _parse_csv_values(value: str | None) -> set[str]:
    if not value:
        return set()
    return {item.strip() for item in value.split(",") if item.strip()}


def _parse_month_values(value: str | None) -> set[int]:
    months: set[int] = set()
    for item in _parse_csv_values(value):
        try:
            month = int(item)
        except ValueError as exc:
            raise ValueError("registration_months chỉ nhận giá trị tháng 1-12") from exc
        if month < 1 or month > 12:
            raise ValueError("registration_months chỉ nhận giá trị tháng 1-12")
        months.add(month)
    return months


def _parse_allowed_values(value: str | None, allowed: set[str], field: str) -> set[str]:
    values = _parse_csv_values(value)
    invalid = sorted(values - allowed)
    if invalid:
        raise ValueError(f"{field} không hợp lệ: {', '.join(invalid)}")
    return values


def _filtered_fi_export_records(db: Session, filters: dict[str, set[Any]]) -> list[SKCTKTModel]:
    records = db.execute(
        select(SKCTKTModel).where(SKCTKTModel.status != SKStatus.DRAFT.value)
    ).scalars().all()
    filtered = [record for record in records if _fi_export_record_matches(record, filters)]
    return sorted(filtered, key=_fi_export_sort_key)


def _fi_export_record_matches(record: SKCTKTModel, filters: dict[str, set[Any]]) -> bool:
    registration_month, _ = _fi_registration_period(record)
    teams = filters.get("teams") or set()
    months = filters.get("registration_months") or set()
    decisions = filters.get("decisions") or set()
    khmt = filters.get("khmt") or set()
    completion = filters.get("completion") or set()
    authors = filters.get("authors") or set()
    submitters = filters.get("submitters") or set()
    if teams and record.team not in teams:
        return False
    if months and registration_month not in months:
        return False
    if decisions and _fi_decision_filter(record) not in decisions:
        return False
    if khmt and _fi_khmt_filter(record) not in khmt:
        return False
    if completion and _fi_completion_filter(record) not in completion:
        return False
    if authors and not _fi_person_filter_matches(record.author_user_id, record.author_name, authors):
        return False
    submitter_id = submitter_user_id(record)
    if submitters and not _fi_person_filter_matches(submitter_id, None, submitters):
        return False
    return True


def _normalize_person_filter_value(value: str | None) -> str:
    text = str(value or "").strip().lower().replace("đ", "d")
    return "".join(
        char for char in unicodedata.normalize("NFD", text)
        if unicodedata.category(char) != "Mn"
    )


def _fi_person_filter_matches(user_id: str | None, display_name: str | None, selected_keys: set[Any]) -> bool:
    keys: set[str] = set()
    if user_id and user_id != "historical-import":
        keys.add(f"id:{user_id}")
    if user_id == "historical-import":
        keys.add("name:he thong")
    name_key = _normalize_person_filter_value(display_name)
    if name_key:
        keys.add(f"name:{name_key}")
    return bool(keys & {str(key) for key in selected_keys})


def _fi_export_sort_key(record: SKCTKTModel) -> tuple[int, int, int, float]:
    registration_month, registration_year = _fi_registration_period(record)
    source_row = record.bm01_source_row or 0
    created_at = _datetime_for_sort(record.created_at)
    return (
        -(registration_year or 0),
        -(registration_month or 0),
        source_row,
        -created_at.timestamp() if created_at else 0,
    )


def _fi_registration_period(record: SKCTKTModel) -> tuple[int | None, int]:
    history = record.status_history if isinstance(record.status_history, list) else []
    first = history[0] if history and isinstance(history[0], dict) else {}
    comments = first.get("comments") if isinstance(first.get("comments"), dict) else {}
    month = _int_or_none(comments.get("registration_month"))
    year = _int_or_none(comments.get("registration_year"))
    if month is not None and 1 <= month <= 12:
        return month, year or _fallback_registration_year(record)
    created_at = _datetime_for_sort(record.created_at)
    if created_at is not None:
        return created_at.month, created_at.year
    return None, year or _utc_now().year


def _fallback_registration_year(record: SKCTKTModel) -> int:
    created_at = _datetime_for_sort(record.created_at)
    if created_at is not None:
        return created_at.year
    if record.khmt_year:
        return int(record.khmt_year)
    return _utc_now().year


def _int_or_none(value: Any) -> int | None:
    try:
        return int(value)
    except (TypeError, ValueError):
        return None


def _datetime_for_sort(value: datetime | None) -> datetime | None:
    if value is None:
        return None
    if value.tzinfo is None:
        return value.replace(tzinfo=timezone.utc)
    return value.astimezone(timezone.utc)


def _excel_datetime(value: datetime | None) -> datetime | str:
    normalized = _datetime_for_sort(value)
    return normalized.replace(tzinfo=None) if normalized else ""


def _fi_decision_filter(record: SKCTKTModel) -> str:
    if record.status in {SKStatus.APPROVED.value, SKStatus.COMPLETED.value}:
        return "approved"
    if record.status == SKStatus.REJECTED.value:
        return "rejected"
    if record.status == SKStatus.DEFERRED.value:
        return "deferred"
    return "pending"


def _fi_khmt_filter(record: SKCTKTModel) -> str:
    return "in" if record.consider_for_khmt else "out"


def _fi_completion_filter(record: SKCTKTModel) -> str:
    if record.status == SKStatus.COMPLETED.value:
        return "done"
    if record.completed_at is not None:
        return "done"
    return "done" if completion_plan_indicates_done(record.completion_plan) else "pending"


def _fi_registration_label(record: SKCTKTModel) -> str:
    month, year = _fi_registration_period(record)
    return f"T{month}/{year}" if month else "Chưa rõ"


def _fi_khmt_label(record: SKCTKTModel) -> str:
    if record.consider_for_khmt:
        return f"KHMT T{record.khmt_month}/{record.khmt_year}" if record.khmt_month and record.khmt_year else "Đã vào KHMT"
    return "Chưa vào KHMT"


def _fi_review_note(record: SKCTKTModel) -> str:
    return record.fi_coordinator_comments or record.bm01_raw_conclusion or ""


def _write_fi_export_summary(
    sheet,
    records: list[SKCTKTModel],
    filters: dict[str, set[Any]],
    generated_at: datetime,
) -> None:
    approved_count = sum(1 for record in records if _fi_decision_filter(record) == "approved")
    khmt_count = sum(1 for record in records if _fi_khmt_filter(record) == "in")
    done_count = sum(1 for record in records if _fi_completion_filter(record) == "done")

    sheet.merge_cells("A1:H1")
    sheet["A1"] = "BÁO CÁO FI/SK-CTKT"
    sheet.merge_cells("A2:H2")
    sheet["A2"] = "Tổng hợp sáng kiến FI/SK-CTKT theo bộ lọc đang áp dụng (không bao gồm bản nháp)"
    sheet["A4"] = "Thời điểm xuất"
    sheet["B4"] = _excel_datetime(generated_at)
    sheet["A5"] = "Phạm vi dữ liệu"
    sheet["B5"] = "Toàn bộ sáng kiến đã gửi FI"
    sheet["A6"] = "Số dòng dữ liệu"
    sheet["B6"] = len(records)

    navy = (FI_NAVY, FI_NAVY_TINT)
    green = (FI_GREEN, FI_GREEN_TINT)
    _write_summary_card(sheet, "D", "Tổng SK", len(records), "sáng kiến", navy)
    _write_summary_card(sheet, "E", "Đồng ý", approved_count, "đạt xét duyệt", green)
    _write_summary_card(sheet, "F", "Đã vào KHMT", khmt_count, "được ghi nhận", navy)
    _write_summary_card(sheet, "G", "Hoàn thành", done_count, "đã hoàn thành", green)

    filter_rows = [
        ("Đội/tổ", _filter_value_text(filters.get("teams"), {})),
        ("Tháng đăng ký", _filter_value_text(filters.get("registration_months"), {}, prefix="T")),
        ("Kết luận LĐX", _filter_value_text(filters.get("decisions"), FI_DECISION_LABELS)),
        ("KHMT", _filter_value_text(filters.get("khmt"), FI_KHMT_LABELS)),
        ("Hoàn thành", _filter_value_text(filters.get("completion"), FI_COMPLETION_LABELS)),
    ]
    sheet.merge_cells("A9:B9")
    sheet["A9"] = "Bộ lọc đang áp dụng"
    sheet["A10"] = "Bộ lọc"
    sheet["B10"] = "Giá trị"
    for row_index, (label, value) in enumerate(filter_rows, start=11):
        sheet.cell(row_index, 1).value = label
        sheet.cell(row_index, 2).value = value

    first_stats_row = 18
    first_group_rows = [
        _write_counter_table(
            sheet,
            first_stats_row,
            1,
            "Theo đội/tổ",
            "Đội/tổ",
            Counter(record.team for record in records),
        ),
        _write_counter_table(
            sheet,
            first_stats_row,
            4,
            "Theo tháng đăng ký",
            "Tháng",
            Counter(_fi_registration_label(record) for record in records),
        ),
        _write_counter_table(
            sheet,
            first_stats_row,
            7,
            "Theo kết luận LĐX",
            "Kết luận",
            Counter(FI_DECISION_LABELS[_fi_decision_filter(record)] for record in records),
        ),
    ]
    second_stats_row = max(first_group_rows) + 3
    _write_counter_table(
        sheet,
        second_stats_row,
        1,
        "Theo KHMT",
        "KHMT",
        Counter(FI_KHMT_LABELS[_fi_khmt_filter(record)] for record in records),
    )
    _write_counter_table(
        sheet,
        second_stats_row,
        4,
        "Theo hoàn thành",
        "Hoàn thành",
        Counter(FI_COMPLETION_LABELS[_fi_completion_filter(record)] for record in records),
    )


def _filter_value_text(values: set[Any] | None, labels: dict[str, str], *, prefix: str = "") -> str:
    if not values:
        return "Tất cả"
    rendered = []
    for value in sorted(values):
        key = str(value)
        rendered.append(labels.get(key, f"{prefix}{key}"))
    return ", ".join(rendered)


def _write_summary_card(
    sheet,
    column: str,
    title: str,
    value: int,
    helper: str,
    accent: tuple[str, str],
) -> None:
    """A KPI tile spanning rows 4-6 of one column: label / big number / helper."""
    accent_color, tint = accent
    border = _thin_border()

    label = sheet[f"{column}4"]
    label.value = title
    label.fill = PatternFill("solid", fgColor=accent_color)
    label.font = Font(name=FI_BASE_FONT, bold=True, size=10, color=FI_ON_DARK)
    label.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    number = sheet[f"{column}5"]
    number.value = value
    number.fill = PatternFill("solid", fgColor=tint)
    number.font = Font(name=FI_BASE_FONT, bold=True, size=22, color=accent_color)
    number.alignment = Alignment(horizontal="center", vertical="center")
    number.number_format = "#,##0"

    note = sheet[f"{column}6"]
    note.value = helper
    note.fill = PatternFill("solid", fgColor=tint)
    note.font = Font(name=FI_BASE_FONT, italic=True, size=9, color=FI_INK_SOFT)
    note.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row in (4, 5, 6):
        sheet[f"{column}{row}"].border = border


def _write_counter_table(sheet, start_row: int, start_col: int, title: str, label: str, counts: Counter) -> int:
    """A small breakdown table (title banner / header / striped rows). Styled inline."""
    border = _thin_border()
    end_col = start_col + 1

    sheet.merge_cells(start_row=start_row, start_column=start_col, end_row=start_row, end_column=end_col)
    for col in (start_col, end_col):
        cell = sheet.cell(start_row, col)
        cell.fill = PatternFill("solid", fgColor=FI_NAVY)
        cell.font = Font(name=FI_BASE_FONT, bold=True, size=11, color=FI_ON_DARK)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    sheet.cell(start_row, start_col).value = title

    header_row = start_row + 1
    label_cell = sheet.cell(header_row, start_col)
    count_cell = sheet.cell(header_row, end_col)
    label_cell.value = label
    count_cell.value = "Số lượng"
    for cell in (label_cell, count_cell):
        cell.fill = PatternFill("solid", fgColor=FI_NAVY_TINT)
        cell.font = Font(name=FI_BASE_FONT, bold=True, color=FI_INK)
        cell.border = border
    label_cell.alignment = Alignment(horizontal="left", vertical="center")
    count_cell.alignment = Alignment(horizontal="center", vertical="center")

    items = sorted(counts.items(), key=lambda item: str(item[0])) if counts else [("Không có dữ liệu", 0)]
    row_index = header_row + 1
    for offset, (key, count) in enumerate(items):
        stripe = FI_STRIPE if offset % 2 else FI_SURFACE
        key_cell = sheet.cell(row_index, start_col)
        value_cell = sheet.cell(row_index, end_col)
        key_cell.value = key
        value_cell.value = count
        for cell in (key_cell, value_cell):
            cell.fill = PatternFill("solid", fgColor=stripe)
            cell.border = border
        key_cell.font = Font(name=FI_BASE_FONT, color=FI_INK)
        key_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        value_cell.font = Font(name=FI_BASE_FONT, bold=True, color=FI_INK)
        value_cell.alignment = Alignment(horizontal="center", vertical="center")
        value_cell.number_format = "#,##0"
        row_index += 1
    return row_index - 1


def _write_fi_export_rows(sheet, records: list[SKCTKTModel], generated_at: datetime) -> None:
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(FI_EXPORT_DATA_HEADERS))
    sheet["A1"] = "DANH SÁCH FI/SK-CTKT"
    sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(FI_EXPORT_DATA_HEADERS))
    generated_label = _excel_datetime(generated_at).strftime("%d/%m/%Y %H:%M")
    sheet["A2"] = f"Xuất lúc {generated_label}  •  Tổng {len(records)} sáng kiến  •  Không bao gồm bản nháp"
    for col_index, header in enumerate(FI_EXPORT_DATA_HEADERS, start=1):
        sheet.cell(FI_EXPORT_DATA_HEADER_ROW, col_index).value = header
    for index, record in enumerate(records, start=1):
        row_index = FI_EXPORT_DATA_HEADER_ROW + index
        row_values = [
            index,
            record.sk_code,
            record.title,
            record.author_name,
            record.author_user_id,
            record.team,
            _fi_registration_label(record),
            FI_STATUS_LABELS.get(record.status, record.status),
            FI_DECISION_LABELS[_fi_decision_filter(record)],
            _fi_khmt_label(record),
            f"T{record.khmt_month}/{record.khmt_year}" if record.khmt_month and record.khmt_year else "",
            FI_COMPLETION_LABELS[_fi_completion_filter(record)],
            record.completion_plan,
            record.content_description,
            _fi_review_note(record),
            record.decision_note or "",
        ]
        for col_index, value in enumerate(row_values, start=1):
            sheet.cell(row_index, col_index).value = value


def _style_fi_export_workbook(workbook: Workbook) -> None:
    _style_fi_summary_sheet(workbook["Tong hop"])
    _style_fi_data_sheet(workbook["Du lieu FI"])


def _style_fi_summary_sheet(sheet) -> None:
    sheet.sheet_view.showGridLines = False
    sheet.sheet_properties.tabColor = FI_NAVY
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.print_options.horizontalCentered = True
    sheet.page_margins.left = 0.3
    sheet.page_margins.right = 0.3
    sheet.page_margins.top = 0.5
    sheet.page_margins.bottom = 0.5
    border = _thin_border()

    # Title banner + subtitle.
    title = sheet["A1"]
    title.font = Font(name=FI_BASE_FONT, bold=True, size=18, color=FI_ON_DARK)
    title.fill = PatternFill("solid", fgColor=FI_NAVY)
    title.alignment = Alignment(horizontal="center", vertical="center")
    subtitle = sheet["A2"]
    subtitle.font = Font(name=FI_BASE_FONT, italic=True, size=10, color=FI_INK_SOFT)
    subtitle.fill = PatternFill("solid", fgColor=FI_NAVY_TINT)
    subtitle.alignment = Alignment(horizontal="center", vertical="center")
    sheet.row_dimensions[1].height = 32
    sheet.row_dimensions[2].height = 20

    # Export metadata block (A4:B6). KPI tiles (D4:G6) are styled where written.
    for row in range(4, 7):
        label = sheet.cell(row, 1)
        value = sheet.cell(row, 2)
        label.fill = PatternFill("solid", fgColor=FI_NAVY_TINT)
        label.font = Font(name=FI_BASE_FONT, bold=True, color=FI_INK)
        label.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        label.border = border
        value.fill = PatternFill("solid", fgColor=FI_SURFACE)
        value.font = Font(name=FI_BASE_FONT, color=FI_INK)
        value.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        value.border = border
    sheet["B4"].number_format = "dd/mm/yyyy hh:mm"
    sheet["B6"].font = Font(name=FI_BASE_FONT, bold=True, size=13, color=FI_NAVY)
    sheet.row_dimensions[4].height = 22
    sheet.row_dimensions[5].height = 34
    sheet.row_dimensions[6].height = 20

    # Filter recap block: banner (A9:B9), header (A10:B10), values (A11:B15).
    banner = sheet["A9"]
    banner.font = Font(name=FI_BASE_FONT, bold=True, color=FI_ON_DARK)
    banner.fill = PatternFill("solid", fgColor=FI_NAVY)
    banner.alignment = Alignment(horizontal="left", vertical="center")
    sheet.cell(9, 2).fill = PatternFill("solid", fgColor=FI_NAVY)
    sheet.row_dimensions[9].height = 20
    for col in (1, 2):
        head = sheet.cell(10, col)
        head.font = Font(name=FI_BASE_FONT, bold=True, color=FI_INK)
        head.fill = PatternFill("solid", fgColor=FI_NAVY_TINT)
        head.alignment = Alignment(horizontal="left", vertical="center")
    for offset, row in enumerate(range(11, 16)):
        stripe = FI_STRIPE if offset % 2 else FI_SURFACE
        for col in (1, 2):
            cell = sheet.cell(row, col)
            cell.fill = PatternFill("solid", fgColor=stripe)
            cell.font = Font(name=FI_BASE_FONT, color=FI_INK)
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    for row in range(9, 16):
        for col in (1, 2):
            sheet.cell(row, col).border = border

    widths = {"A": 24, "B": 32, "C": 3, "D": 17, "E": 17, "F": 17, "G": 22, "H": 14}
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width


def _style_fi_data_sheet(sheet) -> None:
    sheet.sheet_view.showGridLines = False
    sheet.sheet_properties.tabColor = FI_GREEN
    sheet.freeze_panes = f"A{FI_EXPORT_DATA_HEADER_ROW + 1}"
    last_col = get_column_letter(len(FI_EXPORT_DATA_HEADERS))
    sheet.auto_filter.ref = f"A{FI_EXPORT_DATA_HEADER_ROW}:{last_col}{sheet.max_row}"
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_margins.left = 0.2
    sheet.page_margins.right = 0.2
    sheet.page_margins.top = 0.4
    sheet.page_margins.bottom = 0.4

    stripe_fill = PatternFill("solid", fgColor=FI_STRIPE)
    border = _thin_border()
    grid_side = Side(style="thin", color=FI_GRID)
    header_border = Border(
        left=grid_side,
        right=grid_side,
        top=grid_side,
        bottom=Side(style="medium", color=FI_GREEN),
    )

    title = sheet["A1"]
    title.font = Font(name=FI_BASE_FONT, bold=True, size=18, color=FI_ON_DARK)
    title.fill = PatternFill("solid", fgColor=FI_NAVY)
    title.alignment = Alignment(horizontal="center", vertical="center")
    subtitle = sheet["A2"]
    subtitle.font = Font(name=FI_BASE_FONT, italic=True, size=10, color=FI_INK_SOFT)
    subtitle.fill = PatternFill("solid", fgColor=FI_NAVY_TINT)
    subtitle.alignment = Alignment(horizontal="center", vertical="center")
    sheet.row_dimensions[1].height = 32
    sheet.row_dimensions[2].height = 20

    for col_index in range(1, len(FI_EXPORT_DATA_HEADERS) + 1):
        cell = sheet.cell(FI_EXPORT_DATA_HEADER_ROW, col_index)
        cell.fill = PatternFill("solid", fgColor=FI_NAVY)
        cell.font = Font(name=FI_BASE_FONT, bold=True, color=FI_ON_DARK)
        cell.border = header_border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sheet.row_dimensions[FI_EXPORT_DATA_HEADER_ROW].height = 36

    centered_cols = {1, 6, 7, 8, 9, 10, 11, 12}
    tone_cols = (8, 9, 10, 12)
    for row_index in range(FI_EXPORT_DATA_HEADER_ROW + 1, sheet.max_row + 1):
        striped = row_index % 2 == 0
        for col_index in range(1, len(FI_EXPORT_DATA_HEADERS) + 1):
            cell = sheet.cell(row_index, col_index)
            if striped:
                cell.fill = stripe_fill
            cell.border = border
            cell.font = Font(name=FI_BASE_FONT, color=FI_INK)
            if col_index in centered_cols:
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            else:
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        for col_index in tone_cols:
            cell = sheet.cell(row_index, col_index)
            _apply_fi_export_tone(cell, str(cell.value or ""))
        # No explicit row height: let Excel auto-fit wrapped content so long
        # descriptions are never clipped.

    widths = [
        6,    # STT
        20,   # Mã SK
        40,   # Tên SK-CTKT
        20,   # Tác giả
        18,   # Tài khoản tác giả
        10,   # Đội/tổ
        12,   # Tháng đăng ký
        15,   # Trạng thái
        14,   # Kết luận LĐX
        16,   # KHMT
        12,   # Tháng KHMT
        15,   # Hoàn thành
        24,   # Kế hoạch hoàn thành
        50,   # Nội dung
        34,   # Nhận xét FI/BM01
        30,   # Ghi chú quyết định
    ]
    for col_index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(col_index)].width = width


def _thin_border() -> Border:
    side = Side(style="thin", color=FI_GRID)
    return Border(left=side, right=side, top=side, bottom=side)


def _apply_fi_export_tone(cell, value: str) -> None:
    tone = _fi_export_tone(value)
    if tone is None:
        return
    fill_color, font_color = tone
    cell.fill = PatternFill("solid", fgColor=fill_color)
    cell.font = Font(name=FI_BASE_FONT, bold=True, color=font_color)


def _fi_export_tone(value: str) -> tuple[str, str] | None:
    if value in {
        "Đồng ý",
        "Đã phê duyệt",
        "Hoàn tất",
        "Đã vào KHMT",
        "Đã hoàn thành",
    } or value.startswith("KHMT "):
        return "DCFCE7", "166534"
    if value in {"Không đồng ý", "Từ chối", "Đã hủy"}:
        return "FEE2E2", "991B1B"
    if value in {"Xem xét sau", "Cần bổ sung", "Chưa hoàn thành"}:
        return "FEF3C7", "92400E"
    if value in {"Chờ xét duyệt", "Đã xem xét", "Chưa duyệt"}:
        return "DBEAFE", "1E3A8A"
    if value == "Chưa vào KHMT":
        return "F1F5F9", "475569"
    return None


def delete_sk_ctkt(db: Session, record_id: str, actor: str, role: str) -> None:
    record = db.get(SKCTKTModel, record_id)
    if record is None:
        raise KeyError("SK-CTKT not found")
    can_delete = role == Role.ADMIN.value
    if role in AUTHOR_ROLES:
        can_delete = is_author_or_submitter(record, actor) and record.status in OWNER_DELETABLE_STATUSES
    if not can_delete:
        raise PermissionError("Chỉ tác giả/người gửi hộ được xóa SK mới trình hoặc cần bổ sung; Admin được xóa SK")
    images = list(db.execute(select(SKImageModel).where(SKImageModel.sk_ctkt_id == record_id)).scalars())
    audit(db, actor, "SK_CTKT", record_id, "delete", {"record": model_to_dict(record), "image_count": len(images)})
    for image in images:
        try:
            Path(image.file_path).unlink(missing_ok=True)
        except OSError:
            pass
        db.delete(image)
    db.delete(record)
    db.commit()
