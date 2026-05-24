from dataclasses import asdict, dataclass
from datetime import datetime
from pathlib import Path
import re
from typing import Any

from openpyxl import load_workbook


SHEET_TEAM = {"TBCH": "TBCH", "TBĐ": "TBĐL", "TBHTĐK": "TBHTĐK", "TC- ĐK": "TCĐK"}
SHEET_KHMT_COLUMN = {"TBĐ": 14}
SHEET_LEADER_CONCLUSION_COLUMN: dict[str, int | None] = {"TBĐ": None}
APPROVED_STATUSES = {"Approved", "Completed"}


@dataclass
class BM01PreviewRow:
    source_sheet: str
    source_row: int
    team: str
    author_name: str
    title: str
    content_description: str
    completion_plan: str
    raw_conclusion: str
    workshop_leader_conclusion: str
    khmt_raw: str
    status: str
    registration_month: int | None
    registration_year: int | None
    khmt_month: int | None
    khmt_year: int | None
    consider_for_khmt: bool
    warnings: list[str]

    def to_dict(self) -> dict[str, Any]:
        return asdict(self)


def _clean(value: Any) -> str:
    return str(value or "").strip()


def _parse_month_year(value: str) -> tuple[int | None, int | None]:
    text = _clean(value).lower()
    if not text:
        return None, None
    match = re.search(
        r"(?:tháng|thang|t)?\s*(1[0-2]|0?[1-9])\s*[./-]\s*(20\d{2})",
        text,
        re.IGNORECASE,
    )
    if match:
        return int(match.group(1)), int(match.group(2))
    match = re.search(r"(?:tháng|thang|t)\s*(1[0-2]|0?[1-9])\b", text, re.IGNORECASE)
    if match:
        return int(match.group(1)), None
    if re.fullmatch(r"1[0-2]|0?[1-9]", text):
        return int(text), None
    return None, None


def _status_from_review(value: str) -> tuple[str, str | None]:
    lowered = value.lower().strip()
    if not lowered:
        return "Submitted", "Missing approval status"
    rejected_markers = [
        "không đồng ý",
        "khong dong y",
        "khồng đồng ý",
        "khong dong ý",
        "không đạt",
        "khong dat",
        "không dat",
    ]
    if any(marker in lowered for marker in rejected_markers):
        return "Rejected", None
    if "xem xét sau" in lowered or "xem xet sau" in lowered:
        return "Deferred", None
    if "đồng ý" in lowered or "dong y" in lowered:
        return "Approved", None
    return "Submitted", "Unclear approval status"


def _has_data(sheet, row: int) -> bool:
    author = _clean(sheet.cell(row, 4).value)
    title = _clean(sheet.cell(row, 5).value)
    content = _clean(sheet.cell(row, 6).value)
    if author.lower() in {"họ và tên tác giả chính", "họ và tên tác giả"}:
        return False
    return bool(author and title and content)


def _cell(sheet, row: int, column: int | None) -> str:
    if column is None:
        return ""
    return _clean(sheet.cell(row, column).value)


def build_bm01_status_history(
    row: dict[str, Any],
    *,
    imported_by: str,
    imported_at: datetime,
) -> list[dict[str, Any]]:
    comments: dict[str, Any] = {
        "registration_month": row["registration_month"],
        "registration_year": row["registration_year"],
        "month_raw": row.get("month_raw"),
        "khmt_raw": row.get("khmt_raw"),
        "khmt_month": row.get("khmt_month"),
        "khmt_year": row.get("khmt_year"),
        "consider_for_khmt": row.get("consider_for_khmt", False),
    }
    history = [
        {
            "from_status": "Legacy",
            "to_status": row["status"],
            "changed_by": imported_by,
            "changed_at": imported_at.isoformat(),
            "reason": row["raw_conclusion"] or "Legacy BM01 chưa có dữ liệu xét duyệt",
            "comments": comments,
        }
    ]
    if row.get("consider_for_khmt"):
        history.append(
            {
                "from_status": row["status"],
                "to_status": row["status"],
                "changed_by": imported_by,
                "changed_at": imported_at.isoformat(),
                "reason": "khmt_legacy_note",
                "comments": {
                    "khmt_month": row.get("khmt_month"),
                    "khmt_year": row.get("khmt_year"),
                    "khmt_raw": row.get("khmt_raw"),
                    "source": "BM01",
                },
            }
        )
    return history


def preview_bm01(path: Path) -> dict[str, Any]:
    workbook = load_workbook(path, read_only=False, data_only=True, keep_links=False)
    rows: list[BM01PreviewRow] = []
    workbook_warnings: list[str] = []
    for sheet_name, team in SHEET_TEAM.items():
        if sheet_name not in workbook.sheetnames:
            workbook_warnings.append(f"Missing sheet {sheet_name}")
            continue
        sheet = workbook[sheet_name]
        current_registration_month: int | None = None
        khmt_column = SHEET_KHMT_COLUMN.get(sheet_name, 15)
        leader_column = SHEET_LEADER_CONCLUSION_COLUMN.get(sheet_name, 14)
        for row in range(1, sheet.max_row + 1):
            month_raw = _cell(sheet, row, 1)
            candidate_month, _ = _parse_month_year(month_raw)
            if candidate_month is not None:
                current_registration_month = candidate_month
            if not _has_data(sheet, row):
                continue
            raw_conclusion = _cell(sheet, row, 13)
            status, status_warning = _status_from_review(raw_conclusion)
            khmt_raw = _cell(sheet, row, khmt_column)
            khmt_month, khmt_year = _parse_month_year(khmt_raw)
            note_month, _ = _parse_month_year(_cell(sheet, row, 12))
            registration_month = candidate_month or khmt_month or note_month or current_registration_month
            if registration_month is not None:
                current_registration_month = registration_month
            registration_year = 2026
            khmt_year = khmt_year or (registration_year if khmt_month else None)
            consider_for_khmt = status in APPROVED_STATUSES and khmt_month is not None
            warnings = []
            if registration_month is None:
                warnings.append("Missing or ambiguous registration month")
            if status_warning:
                warnings.append(status_warning)
            if sheet.row_dimensions[row].hidden:
                warnings.append("Source row is hidden")
            rows.append(
                BM01PreviewRow(
                    source_sheet=sheet_name,
                    source_row=row,
                    team=team,
                    author_name=_cell(sheet, row, 4),
                    title=_cell(sheet, row, 5),
                    content_description=_cell(sheet, row, 6),
                    completion_plan=_cell(sheet, row, 11),
                    raw_conclusion=raw_conclusion,
                    workshop_leader_conclusion=_cell(sheet, row, leader_column),
                    khmt_raw=khmt_raw,
                    status=status,
                    registration_month=registration_month,
                    registration_year=registration_year,
                    khmt_month=khmt_month,
                    khmt_year=khmt_year,
                    consider_for_khmt=consider_for_khmt,
                    warnings=warnings,
                )
            )
    workbook.close()
    return {
        "source_file": str(path),
        "rows": [row.to_dict() for row in rows],
        "warnings": workbook_warnings,
        "row_count": len(rows),
    }
